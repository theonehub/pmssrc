"""
Unified Taxation Controller
Handles all taxation-related operations and calculations for both comprehensive and employee-specific use cases
"""

from typing import List, Dict, Any, Optional, Union, Tuple
from datetime import datetime, date
from decimal import Decimal
from fastapi import HTTPException, status

# Import centralized logger
from app.utils.logger import get_logger

# Excel-related imports
import pandas as pd
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# Import all DTOs from both controllers
from app.application.dto.taxation_dto import (
    # Basic DTOs
    UpdateSalaryIncomeRequest,
    UpdateDeductionsRequest,
    UpdateResponse,
    
    # Comprehensive DTOs
    PerquisitesDTO,
    HousePropertyIncomeDTO,
    CapitalGainsIncomeDTO,
    RetirementBenefitsDTO,
    OtherIncomeDTO,

    ErrorResponse,
    
    # Employee Selection DTOs
    EmployeeSelectionQuery,
    EmployeeSelectionResponse,
    
    # Individual Component Update DTOs
    UpdateSalaryComponentRequest,
    UpdatePerquisitesComponentRequest,
    UpdateDeductionsComponentRequest,
    UpdateHousePropertyComponentRequest,
    UpdateCapitalGainsComponentRequest,
    UpdateRetirementBenefitsComponentRequest,
    UpdateOtherIncomeComponentRequest,
    UpdateMonthlyPayrollComponentRequest,
    UpdateRegimeComponentRequest,
    IsRegimeUpdateAllowedRequest,
    IsRegimeUpdateAllowedResponse,
    ComponentUpdateResponse,
    ComponentResponse,
    TaxationRecordStatusResponse,
    FlatRetirementBenefitsDTO,
    
    # Monthly Salary DTOs
    MonthlySalaryComputeRequestDTO,
    MonthlySalaryResponseDTO,
)

# Import domain entities and value objects
from app.domain.value_objects.money import Money
from app.domain.value_objects.tax_regime import TaxRegime, TaxRegimeType
from app.domain.value_objects.employee_id import EmployeeId
from app.domain.value_objects.tax_year import TaxYear
from app.domain.entities.taxation.salary_income import SalaryIncome
from app.domain.entities.taxation.deductions import TaxDeductions
from app.domain.entities.taxation.deductions import (
    DeductionSection80C, DeductionSection80D, DeductionSection80E, 
    DeductionSection80G, DeductionSection80TTA_TTB, OtherDeductions
)
from app.domain.entities.taxation.perquisites import (
    Perquisites, AccommodationPerquisite, CarPerquisite, AccommodationType,
    CityPopulation, CarUseType, AssetType, LTAPerquisite,
    InterestFreeConcessionalLoan, ESOPPerquisite, UtilitiesPerquisite,
    FreeEducationPerquisite, MovableAssetUsage, MovableAssetTransfer,
    LunchRefreshmentPerquisite, GiftVoucherPerquisite, MonetaryBenefitsPerquisite,
    ClubExpensesPerquisite, DomesticHelpPerquisite
)
from app.domain.entities.taxation.house_property_income import HousePropertyIncome
from app.domain.entities.taxation.capital_gains import CapitalGainsIncome
from app.domain.entities.taxation.retirement_benefits import (
    RetirementBenefits, LeaveEncashment, Gratuity, VRS, Pension, RetrenchmentCompensation
)
from app.domain.entities.taxation.other_income import OtherIncome, InterestIncome
from app.domain.entities.taxation.taxation_record import SalaryPackageRecord

# Import services
from app.domain.services.taxation.tax_calculation_service import (
    TaxCalculationService, TaxCalculationInput, TaxCalculationResult
)
from app.application.use_cases.taxation.get_employees_for_selection_use_case import GetEmployeesForSelectionUseCase
from app.application.use_cases.taxation.compute_monthly_salary_use_case import ComputeMonthlySalaryUseCase

logger = get_logger(__name__)


class UnifiedTaxationController:
    """
    Unified production-ready controller for all taxation operations.
    
    Combines basic taxation record management with comprehensive calculations
    including all income types:
    - Basic taxation record CRUD operations
    - Salary income (simple and periodic)
    - Perquisites (all 15+ types)
    - House property income
    - Capital gains
    - Retirement benefits  
    - Other income sources
    - Monthly payroll with LWP
    - Enhanced tax calculations and analytics
    - Mid-year scenarios (joiners, increments)
    - Scenario comparisons and planning
    """
    
    def __init__(self,
                 user_repository,
                 salary_package_repository,
                 monthly_salary_repository=None,
                 organisation_repository=None):
        self.user_repository = user_repository
        self.salary_package_repository = salary_package_repository
        self.tax_calculation_service = TaxCalculationService(salary_package_repository)
        from app.config.dependency_container import get_dependency_container
        container = get_dependency_container()
        user_service = container.get_user_service()
        if monthly_salary_repository is None:
            monthly_salary_repository = container.get_monthly_salary_repository()
        self.monthly_salary_repository = monthly_salary_repository
        self.get_employees_for_selection_use_case = GetEmployeesForSelectionUseCase(
            user_query_service=user_service,
            salary_package_repository=salary_package_repository
        )
        self.compute_monthly_salary_use_case = ComputeMonthlySalaryUseCase(
            salary_package_repository=salary_package_repository,
            user_repository=user_repository,
            monthly_salary_repository=monthly_salary_repository,
            tax_calculation_service=self.tax_calculation_service
        )
        # Add organisation repository
        if organisation_repository is None:
            self.organisation_repository = container.get_organisation_repository()
        else:
            self.organisation_repository = organisation_repository
    
    # =============================================================================
    # EMPLOYEE SELECTION OPERATIONS
    # =============================================================================
    
    async def get_employees_for_selection(
        self,
        query: EmployeeSelectionQuery,
        current_user
    ) -> EmployeeSelectionResponse:
        """
        Get employees for taxation selection.
        
        This method leverages the user service to fetch employee data
        and enriches it with tax record information for admin selection interface.
        
        Args:
            query: Query parameters for filtering and pagination
            current_user: Current authenticated user with organization context
            
        Returns:
            Employee selection response with enriched data
        """
        
        logger.info(f"Getting employees for selection with query: {query}")
        
        try:
            # Execute the use case to get employees
            response = await self.get_employees_for_selection_use_case.execute(
                query, current_user
            )
            
            logger.info(f"Retrieved {len(response.employees)} employees for selection")
            return response
            
        except Exception as e:
            logger.error(f"Error getting employees for selection: {str(e)}")
            raise
    
    # =============================================================================
    # COMPONENT-SPECIFIC CALCULATION METHODS
    # =============================================================================
    
    async def calculate_perquisites_only(self,
                                       perquisites: PerquisitesDTO,
                                       regime_type: str,
                                       organization_id: str) -> Dict[str, Any]:
        """Calculate only perquisites tax impact."""
        try:
            # Convert DTO to entity
            perquisites_entity = self._convert_perquisites_dto_to_entity(perquisites)
            regime = TaxRegime(TaxRegimeType.OLD if regime_type.lower() == "old" else TaxRegimeType.NEW)
            
            # Calculate perquisites
            total_perquisites = perquisites_entity.calculate_total_perquisites(regime)
            breakdown = perquisites_entity.get_perquisites_breakdown(regime)
            
            return {
                "total_perquisites_value": total_perquisites.to_float(),
                "perquisites_breakdown": breakdown,
                "regime_used": regime_type,
                "perquisites_applicable": regime.regime_type == TaxRegimeType.OLD
            }
            
        except Exception as e:
            logger.error(f"Error calculating perquisites for org {organization_id}: {str(e)}")
            raise
    
    async def calculate_house_property_income_only(self,
                                          house_property_income: HousePropertyIncomeDTO,
                                          regime_type: str,
                                          organization_id: str) -> Dict[str, Any]:
        """Calculate only house property income."""
        try:
            # Convert DTO to entity
            house_property_entity = self._convert_house_property_income_dto_to_entity(house_property_income)
            regime = TaxRegime(TaxRegimeType.OLD if regime_type.lower() == "old" else TaxRegimeType.NEW)
            
            # Calculate house property income
            net_income = house_property_entity.calculate_net_income_from_house_property_income(regime)
            breakdown = house_property_entity.get_house_property_income_breakdown(regime)
            
            return {
                "net_house_property_income": net_income.to_float(),
                "house_property_income_breakdown": breakdown,
                "regime_used": regime_type
            }
            
        except Exception as e:
            logger.error(f"Error calculating house property for org {organization_id}: {str(e)}")
            raise
    
    async def calculate_capital_gains_only(self,
                                         capital_gains: CapitalGainsIncomeDTO,
                                         regime_type: str,
                                         organization_id: str) -> Dict[str, Any]:
        """Calculate only capital gains tax."""
        try:
            # Convert DTO to entity
            capital_gains_entity = self._convert_capital_gains_dto_to_entity(capital_gains)
            regime = TaxRegime(TaxRegimeType.OLD if regime_type.lower() == "old" else TaxRegimeType.NEW)
            
            # Calculate capital gains
            total_tax = capital_gains_entity.calculate_total_capital_gains_tax()
            breakdown = capital_gains_entity.get_capital_gains_breakdown(regime)
            
            return {
                "total_capital_gains_tax": total_tax.to_float(),
                "capital_gains_breakdown": breakdown,
                "regime_used": regime_type
            }
            
        except Exception as e:
            logger.error(f"Error calculating capital gains for org {organization_id}: {str(e)}")
            raise
    
    async def calculate_retirement_benefits_only(self,
                                               retirement_benefits: RetirementBenefitsDTO,
                                               regime_type: str,
                                               organization_id: str) -> Dict[str, Any]:
        """Calculate only retirement benefits tax."""
        try:
            # Convert DTO to entity
            retirement_entity = self._convert_retirement_benefits_dto_to_entity(retirement_benefits)
            regime = TaxRegime(TaxRegimeType.OLD if regime_type.lower() == "old" else TaxRegimeType.NEW)
            
            # Calculate retirement benefits
            total_income = retirement_entity.calculate_total_retirement_income(regime)
            breakdown = retirement_entity.get_retirement_benefits_breakdown(regime)
            
            return {
                "total_retirement_income": total_income.to_float(),
                "retirement_benefits_breakdown": breakdown,
                "regime_used": regime_type
            }
            
        except Exception as e:
            logger.error(f"Error calculating retirement benefits for org {organization_id}: {str(e)}")
            raise
    
    # =============================================================================
    # DTO TO DOMAIN ENTITY CONVERSION METHODS
    # =============================================================================
    
    def _create_default_salary_income(self):
        """Create default salary income entity."""
        from app.domain.entities.taxation.salary_income import SalaryIncome, SpecificAllowances
        from app.domain.value_objects.money import Money
        from datetime import datetime
        from app.domain.value_objects.tax_year import TaxYear
        
        current_tax_year = TaxYear.current()
        effective_from = datetime.combine(current_tax_year.get_start_date(), datetime.min.time())
        effective_till = datetime.combine(current_tax_year.get_end_date(), datetime.min.time())
        
        return SalaryIncome(
            effective_from=effective_from,
            effective_till=effective_till,
            basic_salary=Money.zero(),
            dearness_allowance=Money.zero(),
            hra_provided=Money.zero(),
            special_allowance=Money.zero(),
            bonus=Money.zero(),
            commission=Money.zero(),
            arrears=Money.zero(),
            specific_allowances=SpecificAllowances()
        )
    
    def _create_default_perquisites(self):
        """Create default perquisites entity."""
        from app.domain.entities.taxation.perquisites import Perquisites
        
        return Perquisites()
    
    def _create_default_house_property_income(self):
        """Create default house property income entity."""
        from app.domain.entities.taxation.house_property_income import HousePropertyIncome, PropertyType
        from app.domain.value_objects.money import Money
        
        return HousePropertyIncome(
            property_type=PropertyType.SELF_OCCUPIED,
            address="",
            annual_rent_received=Money.zero(),
            municipal_taxes_paid=Money.zero(),
            home_loan_interest=Money.zero(),
            pre_construction_interest=Money.zero()
        )
    
    def _create_default_capital_gains(self):
        """Create default capital gains entity."""
        from app.domain.entities.taxation.capital_gains import CapitalGainsIncome
        from app.domain.value_objects.money import Money
        
        return CapitalGainsIncome(
            stcg_111a_equity_stt=Money.zero(),
            stcg_other_assets=Money.zero(),
            ltcg_112a_equity_stt=Money.zero(),
            ltcg_other_assets=Money.zero(),
            ltcg_debt_mf=Money.zero()
        )
    
    def _create_default_retirement_benefits(self):
        """Create default retirement benefits entity."""
        from app.domain.entities.taxation.retirement_benefits import RetirementBenefits
        
        return RetirementBenefits()
    
    def _create_default_other_income(self):
        """Create default other income entity."""
        from app.domain.entities.taxation.other_income import OtherIncome, InterestIncome
        from app.domain.entities.taxation.capital_gains import CapitalGainsIncome
        from app.domain.value_objects.money import Money
        
        # Create default interest income
        default_interest_income = InterestIncome(
            savings_account_interest=Money.zero(),
            fixed_deposit_interest=Money.zero(),
            recurring_deposit_interest=Money.zero(),
            post_office_interest=Money.zero()
        )
        
        # Create default capital gains income
        default_capital_gains_income = CapitalGainsIncome(
            stcg_111a_equity_stt=Money.zero(),
            stcg_other_assets=Money.zero(),
            stcg_debt_mf=Money.zero(),
            ltcg_112a_equity_stt=Money.zero(),
            ltcg_other_assets=Money.zero(),
            ltcg_debt_mf=Money.zero()
        )
        
        return OtherIncome(
            interest_income=default_interest_income,
            capital_gains_income=default_capital_gains_income,
            dividend_income=Money.zero(),
            gifts_received=Money.zero(),
            business_professional_income=Money.zero(),
            other_miscellaneous_income=Money.zero()
        )
    
    def _create_default_deductions(self):
        """Create default tax deductions entity."""
        from app.domain.entities.taxation.deductions import TaxDeductions
        
        return TaxDeductions()
    
    def _convert_salary_dto_to_entity(self, salary_dto) -> SalaryIncome:
        """Convert salary DTO to domain entity."""
        from datetime import datetime
        
        # Handle effective dates - convert from date to datetime if provided
        # If not provided, use default values for the current tax year
        if hasattr(salary_dto, 'effective_from') and salary_dto.effective_from:
            effective_from = datetime.combine(salary_dto.effective_from, datetime.min.time())
        else:
            # Default to start of current tax year
            from app.domain.value_objects.tax_year import TaxYear
            current_tax_year = TaxYear.current()
            effective_from = datetime.combine(current_tax_year.get_start_date(), datetime.min.time())
        
        if hasattr(salary_dto, 'effective_till') and salary_dto.effective_till:
            effective_till = datetime.combine(salary_dto.effective_till, datetime.min.time())
        else:
            # Default to end of current tax year
            from app.domain.value_objects.tax_year import TaxYear
            current_tax_year = TaxYear.current()
            effective_till = datetime.combine(current_tax_year.get_end_date(), datetime.min.time())
        
        # Create SpecificAllowances from DTO fields
        from app.domain.entities.taxation.salary_income import SpecificAllowances
        specific_allowances = SpecificAllowances(
            monthly_hills_allowance=Money.from_decimal(getattr(salary_dto, 'hills_high_altd_allowance', 0)),
            monthly_hills_exemption_limit=Money.from_decimal(getattr(salary_dto, 'hills_high_altd_exemption_limit', 0)),
            monthly_border_allowance=Money.from_decimal(getattr(salary_dto, 'border_remote_allowance', 0)),
            monthly_border_exemption_limit=Money.from_decimal(getattr(salary_dto, 'border_remote_exemption_limit', 0)),
            transport_employee_allowance=Money.from_decimal(getattr(salary_dto, 'transport_employee_allowance', 0)),
            children_education_allowance=Money.from_decimal(getattr(salary_dto, 'children_education_allowance', 0)),
            children_education_count=getattr(salary_dto, 'children_education_count', 0),
            hostel_allowance=Money.from_decimal(getattr(salary_dto, 'hostel_allowance', 0)),
            children_hostel_count=getattr(salary_dto, 'children_hostel_count', 0),
            disabled_transport_allowance=Money.from_decimal(getattr(salary_dto, 'disabled_transport_allowance', 0)),
            is_disabled=getattr(salary_dto, 'is_disabled', False),
            underground_mines_allowance=Money.from_decimal(getattr(salary_dto, 'underground_mines_allowance', 0)),
            mine_work_months=getattr(salary_dto, 'mine_work_months', 0),
            government_entertainment_allowance=Money.from_decimal(getattr(salary_dto, 'govt_employee_entertainment_allowance', 0)),
            city_compensatory_allowance=Money.from_decimal(getattr(salary_dto, 'city_compensatory_allowance', 0)),
            rural_allowance=Money.from_decimal(getattr(salary_dto, 'rural_allowance', 0)),
            proctorship_allowance=Money.from_decimal(getattr(salary_dto, 'proctorship_allowance', 0)),
            wardenship_allowance=Money.from_decimal(getattr(salary_dto, 'wardenship_allowance', 0)),
            project_allowance=Money.from_decimal(getattr(salary_dto, 'project_allowance', 0)),
            deputation_allowance=Money.from_decimal(getattr(salary_dto, 'deputation_allowance', 0)),
            overtime_allowance=Money.from_decimal(getattr(salary_dto, 'overtime_allowance', 0)),
            interim_relief=Money.from_decimal(getattr(salary_dto, 'interim_relief', 0)),
            tiffin_allowance=Money.from_decimal(getattr(salary_dto, 'tiffin_allowance', 0)),
            fixed_medical_allowance=Money.from_decimal(getattr(salary_dto, 'fixed_medical_allowance', 0)),
            servant_allowance=Money.from_decimal(getattr(salary_dto, 'servant_allowance', 0)),
            any_other_allowance=Money.from_decimal(getattr(salary_dto, 'any_other_allowance', 0)),
            any_other_allowance_exemption=Money.from_decimal(getattr(salary_dto, 'any_other_allowance_exemption', 0)),
            govt_employees_outside_india_allowance=Money.from_decimal(getattr(salary_dto, 'govt_employees_outside_india_allowance', 0)),
            supreme_high_court_judges_allowance=Money.from_decimal(getattr(salary_dto, 'supreme_high_court_judges_allowance', 0)),
            judge_compensatory_allowance=Money.from_decimal(getattr(salary_dto, 'judge_compensatory_allowance', 0)),
            section_10_14_special_allowances=Money.from_decimal(getattr(salary_dto, 'section_10_14_special_allowances', 0)),
            travel_on_tour_allowance=Money.from_decimal(getattr(salary_dto, 'travel_on_tour_allowance', 0)),
            tour_daily_charge_allowance=Money.from_decimal(getattr(salary_dto, 'tour_daily_charge_allowance', 0)),
            conveyance_in_performace_of_duties=Money.from_decimal(getattr(salary_dto, 'conveyance_in_performace_of_duties', 0)),
            helper_in_performace_of_duties=Money.from_decimal(getattr(salary_dto, 'helper_in_performace_of_duties', 0)),
            academic_research=Money.from_decimal(getattr(salary_dto, 'academic_research', 0)),
            uniform_allowance=Money.from_decimal(getattr(salary_dto, 'uniform_allowance', 0))
        )
        
        return SalaryIncome(
            effective_from=effective_from,
            effective_till=effective_till,
            basic_salary=Money.from_decimal(salary_dto.basic_salary),
            dearness_allowance=Money.from_decimal(salary_dto.dearness_allowance),
            hra_provided=Money.from_decimal(salary_dto.hra_provided),
            special_allowance=Money.from_decimal(salary_dto.special_allowance),
            # Optional components with defaults
            bonus=Money.from_decimal(getattr(salary_dto, 'bonus', 0)),
            commission=Money.from_decimal(getattr(salary_dto, 'commission', 0)),
            arrears=Money.from_decimal(getattr(salary_dto, 'arrears', 0)),
            specific_allowances=specific_allowances
        )
    
    def _convert_deductions_dto_to_entity(self, deductions_dto) -> TaxDeductions:
        """Convert deductions DTO to domain entity."""
        
        # Initialize with default values
        deductions = TaxDeductions()
        
        if not deductions_dto:
            logger.debug("Deductions DTO is None or empty, returning default deductions")
            return deductions

        try:
            # Handle both nested DTO structure and flat dictionary structure from frontend
            if hasattr(deductions_dto, 'dict'):
                # If it's a Pydantic model, convert to dict
                deductions_data = deductions_dto.dict()
            elif isinstance(deductions_dto, dict):
                # If it's already a dictionary (from frontend)
                deductions_data = deductions_dto
            else:
                # If it's an object with attributes
                deductions_data = deductions_dto.__dict__ if hasattr(deductions_dto, '__dict__') else {}
            
            logger.info(f"Processing deductions data: {type(deductions_data)} with keys: {list(deductions_data.keys()) if isinstance(deductions_data, dict) else 'N/A'}")
            logger.debug(f"Full deductions data received: {deductions_data}")
        except Exception as e:
            logger.error(f"Error processing deductions DTO structure: {str(e)}")
            return deductions

        # Helper function to safely get values
        def safe_get(data, key, default=0):
            if isinstance(data, dict):
                return data.get(key, default)
            return getattr(data, key, default) if hasattr(data, key) else default
        
        # Helper function to safely convert values to Money
        def safe_money_from_value(value, default=0):
            """Safely convert a value to Money, handling various edge cases."""
            import decimal
            try:
                # Handle None or empty values
                if value is None or value == "":
                    return Money.from_decimal(default)
                
                # Handle string values that might be invalid
                if isinstance(value, str):
                    # Strip whitespace and handle common invalid values
                    value = value.strip()
                    if value.lower() in ['null', 'undefined', 'nan', 'none', '']:
                        return Money.from_decimal(default)
                
                # Try to convert to float first, then to decimal
                if isinstance(value, (int, float)):
                    return Money.from_decimal(float(value))
                elif isinstance(value, str):
                    return Money.from_decimal(float(value))
                else:
                    # For any other type, try direct conversion
                    return Money.from_decimal(value)
                    
            except (ValueError, TypeError, decimal.InvalidOperation, decimal.ConversionSyntax) as e:
                logger.warning(f"Failed to convert value '{value}' to Money, using default {default}: {str(e)}")
                return Money.from_decimal(default)

        # Map Section 80C fields - handle both nested and flat structures
        try:
            section_80c_data = safe_get(deductions_data, 'section_80c', {})
            section_80c_keys = [
                'life_insurance_premium', 'epf_contribution', 'ppf_contribution', 
                'nsc_investment', 'tax_saving_fd', 'elss_investment', 'home_loan_principal',
                'tuition_fees', 'ulip_premium', 'sukanya_samriddhi', 'stamp_duty_property',
                'senior_citizen_savings', 'other_80c_investments'
            ]
            
            if section_80c_data or any(key in deductions_data for key in section_80c_keys):
                logger.debug(f"Processing Section 80C data. Nested: {section_80c_data}, Flat keys present: {[k for k in section_80c_keys if k in deductions_data]}")
                
                # Handle nested structure
                if section_80c_data:
                    deductions.section_80c.life_insurance_premium = safe_money_from_value(safe_get(section_80c_data, 'life_insurance_premium', 0))
                    deductions.section_80c.epf_contribution = safe_money_from_value(safe_get(section_80c_data, 'epf_contribution', 0))
                    deductions.section_80c.ppf_contribution = safe_money_from_value(safe_get(section_80c_data, 'ppf_contribution', 0))
                    deductions.section_80c.nsc_investment = safe_money_from_value(safe_get(section_80c_data, 'nsc_investment', 0))
                    deductions.section_80c.tax_saving_fd = safe_money_from_value(safe_get(section_80c_data, 'tax_saving_fd', 0))
                    deductions.section_80c.elss_investment = safe_money_from_value(safe_get(section_80c_data, 'elss_investment', 0))
                    deductions.section_80c.home_loan_principal = safe_money_from_value(safe_get(section_80c_data, 'home_loan_principal', 0))
                    deductions.section_80c.tuition_fees = safe_money_from_value(safe_get(section_80c_data, 'tuition_fees', 0))
                    deductions.section_80c.ulip_premium = safe_money_from_value(safe_get(section_80c_data, 'ulip_premium', 0))
                    deductions.section_80c.sukanya_samriddhi = safe_money_from_value(safe_get(section_80c_data, 'sukanya_samriddhi', 0))
                    deductions.section_80c.stamp_duty_property = safe_money_from_value(safe_get(section_80c_data, 'stamp_duty_property', 0))
                    deductions.section_80c.senior_citizen_savings = safe_money_from_value(safe_get(section_80c_data, 'senior_citizen_savings', 0))
                    deductions.section_80c.other_80c_investments = safe_money_from_value(safe_get(section_80c_data, 'other_80c_investments', 0))
                else:
                    # Handle flat structure (from frontend)
                    deductions.section_80c.life_insurance_premium = safe_money_from_value(safe_get(deductions_data, 'life_insurance_premium', 0))
                    deductions.section_80c.epf_contribution = safe_money_from_value(safe_get(deductions_data, 'epf_contribution', 0))
                    deductions.section_80c.ppf_contribution = safe_money_from_value(safe_get(deductions_data, 'ppf_contribution', 0))
                    deductions.section_80c.nsc_investment = safe_money_from_value(safe_get(deductions_data, 'nsc_investment', 0))
                    deductions.section_80c.tax_saving_fd = safe_money_from_value(safe_get(deductions_data, 'tax_saving_fd', 0))
                    deductions.section_80c.elss_investment = safe_money_from_value(safe_get(deductions_data, 'elss_investment', 0))
                    deductions.section_80c.home_loan_principal = safe_money_from_value(safe_get(deductions_data, 'home_loan_principal', 0))
                    deductions.section_80c.tuition_fees = safe_money_from_value(safe_get(deductions_data, 'tuition_fees', 0))
                    deductions.section_80c.ulip_premium = safe_money_from_value(safe_get(deductions_data, 'ulip_premium', 0))
                    deductions.section_80c.sukanya_samriddhi = safe_money_from_value(safe_get(deductions_data, 'sukanya_samriddhi', 0))
                    deductions.section_80c.stamp_duty_property = safe_money_from_value(safe_get(deductions_data, 'stamp_duty_property', 0))
                    deductions.section_80c.senior_citizen_savings = safe_money_from_value(safe_get(deductions_data, 'senior_citizen_savings', 0))
                    deductions.section_80c.other_80c_investments = safe_money_from_value(safe_get(deductions_data, 'other_80c_investments', 0))
                
                logger.debug(f"Section 80C total after conversion: {deductions.section_80c.calculate_total_investment()}")
        except Exception as e:
            logger.error(f"Error processing Section 80C deductions: {str(e)}")

        # Map Section 80D fields - handle both nested and flat structures
        try:
            section_80d_data = safe_get(deductions_data, 'section_80d', {})
            section_80d_keys = ['self_family_premium', 'parent_premium', 'preventive_health_checkup']
            
            if section_80d_data or any(key in deductions_data for key in section_80d_keys):
                logger.debug(f"Processing Section 80D data. Nested: {section_80d_data}, Flat keys present: {[k for k in section_80d_keys if k in deductions_data]}")
                
                # Handle nested structure
                if section_80d_data:
                    deductions.section_80d.self_family_premium = safe_money_from_value(safe_get(section_80d_data, 'self_family_premium', 0))
                    deductions.section_80d.parent_premium = safe_money_from_value(safe_get(section_80d_data, 'parent_premium', 0))
                    deductions.section_80d.preventive_health_checkup = safe_money_from_value(safe_get(section_80d_data, 'preventive_health_checkup', 0))
                    deductions.section_80d.parent_age = safe_get(section_80d_data, 'parent_age', 55)
                else:
                    # Handle flat structure (from frontend)
                    deductions.section_80d.self_family_premium = safe_money_from_value(safe_get(deductions_data, 'self_family_premium', 0))
                    deductions.section_80d.parent_premium = safe_money_from_value(safe_get(deductions_data, 'parent_premium', 0))
                    deductions.section_80d.preventive_health_checkup = safe_money_from_value(safe_get(deductions_data, 'preventive_health_checkup', 0))
                
                logger.debug(f"Section 80D total after conversion: {deductions.section_80d.self_family_premium.add(deductions.section_80d.parent_premium).add(deductions.section_80d.preventive_health_checkup)}")
        except Exception as e:
            logger.error(f"Error processing Section 80D deductions: {str(e)}")

        # Map HRA Exemption fields - handle both nested and flat structures
        try:
            hra_exemption_data = safe_get(deductions_data, 'hra_exemption', {})
            hra_keys = ['actual_rent_paid', 'hra_city_type']
            
            if hra_exemption_data or any(key in deductions_data for key in hra_keys):
                logger.debug(f"Processing HRA exemption data. Nested: {hra_exemption_data}, Flat keys present: {[k for k in hra_keys if k in deductions_data]}")
                
                from app.domain.entities.taxation.deductions import HRAExemption
                if hra_exemption_data:
                    actual_rent_paid = safe_get(hra_exemption_data, 'actual_rent_paid', 0)
                    hra_city_type = safe_get(hra_exemption_data, 'hra_city_type', 'non_metro')
                else:
                    # Handle flat structure (from frontend)
                    actual_rent_paid = safe_get(deductions_data, 'actual_rent_paid', 0)
                    hra_city_type = safe_get(deductions_data, 'hra_city_type', 'non_metro')
                
                deductions.hra_exemption = HRAExemption(
                    actual_rent_paid=safe_money_from_value(actual_rent_paid),
                    hra_city_type=hra_city_type
                )
                logger.debug(f"HRA exemption created with actual_rent_paid: {actual_rent_paid}, city_type: {hra_city_type}")
        except Exception as e:
            logger.error(f"Error processing HRA exemption deductions: {str(e)}")

        # Map Other Deductions fields - handle both nested and flat structures
        try:
            other_deductions_data = safe_get(deductions_data, 'other_deductions', {})
            other_deduction_keys = [
                'other_deductions', 'ev_loan_interest', 'political_party_contribution'
            ]
            
            # Check if we have any other deduction data (either nested or flat)
            has_nested_data = other_deductions_data and isinstance(other_deductions_data, dict) and len(other_deductions_data) > 0
            has_flat_data = any(key in deductions_data for key in other_deduction_keys)
            
            if has_nested_data or has_flat_data:
                logger.debug(f"Processing other deductions data. Nested: {other_deductions_data}, Flat keys present: {[k for k in other_deduction_keys if k in deductions_data]}")
                
                # Handle nested structure
                if has_nested_data:
                    deductions.other_deductions.other_deductions = safe_money_from_value(safe_get(other_deductions_data, 'other_deductions', 0))
                else:
                    # Handle flat structure (from frontend) - map field names correctly
                    # Map other deductions
                    deductions.other_deductions.other_deductions = safe_money_from_value(safe_get(deductions_data, 'other_deductions', 0))
                
                # Use the correct method name for total calculation
                logger.debug(f"Other deductions total after conversion: {deductions.other_deductions.calculate_total()}")
        except Exception as e:
            logger.error(f"Error processing other deductions: {str(e)}")

        # Log final deductions total
        try:
            from app.domain.value_objects.tax_regime import TaxRegime, TaxRegimeType
            from app.domain.value_objects.money import Money
            regime = TaxRegime(TaxRegimeType.NEW)  # Default to new regime for calculation
            # Use default values for age and gross_income since we don't have them in this context
            total_deductions = deductions.calculate_total_deductions(regime, 30, Money.zero())
            logger.info(f"Final deductions total after conversion: {total_deductions}")
        except Exception as e:
            logger.error(f"Error calculating total deductions: {str(e)}")
        
        return deductions
    
    def _convert_comprehensive_deductions_dto_to_entity(self, comp_deductions_dto) -> TaxDeductions:
        """Convert comprehensive deductions DTO to entity."""
        # Add defensive check for None input
        if comp_deductions_dto is None:
            logger.warning("Received None for comprehensive deductions DTO, creating default deductions")
            return self._create_default_deductions()
        
        # This method would handle the comprehensive deductions conversion
        # For now, we'll use the existing deductions conversion logic
        return self._convert_deductions_dto_to_entity(comp_deductions_dto)
    
    def _convert_perquisites_dto_to_entity(self, perquisites_dto) -> Perquisites:
        """Convert perquisites DTO to entity."""
        from app.domain.entities.taxation.perquisites import (
            AccommodationPerquisite, CarPerquisite, 
            LTAPerquisite, InterestFreeConcessionalLoan, ESOPPerquisite,
            UtilitiesPerquisite, FreeEducationPerquisite, MovableAssetUsage, MovableAssetTransfer,
            LunchRefreshmentPerquisite, GiftVoucherPerquisite, 
            MonetaryBenefitsPerquisite, ClubExpensesPerquisite, DomesticHelpPerquisite,
            AccommodationType, CityPopulation, CarUseType, AssetType
        )
        
        # Helper function to safely convert to Money
        def safe_money_from_value(value, default=0):
            if value is None:
                return Money.from_decimal(Decimal(str(default)))
            return Money.from_decimal(Decimal(str(value)))
        
        # Helper function to safely convert enum values
        def safe_enum_from_value(enum_class, value, default_value):
            """Safely convert a value to enum, handling empty strings and None."""
            if value is None or value == "":
                return default_value
            try:
                return enum_class(value)
            except ValueError:
                logger.warning(f"Invalid value '{value}' for {enum_class.__name__}, using default: {default_value}")
                return default_value
        
        # Convert accommodation perquisite
        accommodation = AccommodationPerquisite(
            accommodation_type=safe_enum_from_value(AccommodationType, perquisites_dto.accommodation_type, AccommodationType.EMPLOYER_OWNED),
            city_population=safe_enum_from_value(CityPopulation, perquisites_dto.city_population, CityPopulation.BELOW_15_LAKHS),
            license_fees=safe_money_from_value(perquisites_dto.license_fees),
            employee_rent_payment=safe_money_from_value(perquisites_dto.employee_rent_payment),
            rent_paid_by_employer=safe_money_from_value(perquisites_dto.rent_paid_by_employer),
            hotel_charges=safe_money_from_value(perquisites_dto.hotel_charges),
            stay_days=perquisites_dto.stay_days,
            furniture_cost=safe_money_from_value(perquisites_dto.furniture_cost),
            furniture_employee_payment=safe_money_from_value(perquisites_dto.furniture_employee_payment),
            is_furniture_owned_by_employer=perquisites_dto.is_furniture_owned_by_employer
        )
        # Note: basic_salary and dearness_allowance are not fields of AccommodationPerquisite; use them at calculation time if needed.
        
        # Convert car perquisite
        car = CarPerquisite(
            car_use_type=safe_enum_from_value(CarUseType, perquisites_dto.car_use_type, CarUseType.PERSONAL),
            engine_capacity_cc=perquisites_dto.engine_capacity_cc,
            months_used=perquisites_dto.months_used,
            months_used_other_vehicle=perquisites_dto.months_used_other_vehicle,
            car_cost_to_employer=safe_money_from_value(perquisites_dto.car_cost_to_employer),
            other_vehicle_cost=safe_money_from_value(perquisites_dto.other_vehicle_cost),
            driver_cost=safe_money_from_value(perquisites_dto.driver_cost),
            has_expense_reimbursement=perquisites_dto.has_expense_reimbursement,
            driver_provided=perquisites_dto.driver_provided
        )
        
        # Convert LTA
        lta = LTAPerquisite(
            is_monthly_paid=perquisites_dto.is_monthly_paid,
            lta_allocated_yearly=safe_money_from_value(perquisites_dto.lta_allocated_yearly),
            lta_amount_claimed=safe_money_from_value(perquisites_dto.lta_amount_claimed),
            lta_claimed_count=perquisites_dto.lta_claimed_count,
            public_transport_cost=safe_money_from_value(perquisites_dto.public_transport_cost),
            travel_mode=perquisites_dto.travel_mode
        )
        
        # Convert interest free loan
        interest_free_loan = InterestFreeConcessionalLoan(
            loan_amount=safe_money_from_value(perquisites_dto.loan_amount),
            emi_amount=safe_money_from_value(perquisites_dto.emi_amount),
            outstanding_amount=safe_money_from_value(perquisites_dto.loan_amount),  # Assuming same as loan amount
            company_interest_rate=Decimal(str(perquisites_dto.company_interest_rate)),
            sbi_interest_rate=Decimal(str(perquisites_dto.sbi_interest_rate)),
            loan_type=perquisites_dto.loan_type,
            loan_start_date=perquisites_dto.loan_start_date
        )
        
        # Convert ESOP
        esop = ESOPPerquisite(
            shares_exercised=perquisites_dto.esop_shares_exercised,
            exercise_price=safe_money_from_value(perquisites_dto.esop_exercise_value),
            allotment_price=safe_money_from_value(perquisites_dto.esop_fair_market_value)
        )
        
        # Convert utilities
        utilities = UtilitiesPerquisite(
            gas_paid_by_employer=safe_money_from_value(perquisites_dto.gas_paid_by_employer),
            electricity_paid_by_employer=safe_money_from_value(perquisites_dto.electricity_paid_by_employer),
            water_paid_by_employer=safe_money_from_value(perquisites_dto.water_paid_by_employer),
            gas_paid_by_employee=safe_money_from_value(perquisites_dto.gas_paid_by_employee),
            electricity_paid_by_employee=safe_money_from_value(perquisites_dto.electricity_paid_by_employee),
            water_paid_by_employee=safe_money_from_value(perquisites_dto.water_paid_by_employee),
            is_gas_manufactured_by_employer=perquisites_dto.is_gas_manufactured_by_employer,
            is_electricity_manufactured_by_employer=perquisites_dto.is_electricity_manufactured_by_employer,
            is_water_manufactured_by_employer=perquisites_dto.is_water_manufactured_by_employer
        )
        
        # Convert free education
        free_education = FreeEducationPerquisite(
            monthly_expenses_child1=safe_money_from_value(perquisites_dto.monthly_expenses_child1),
            monthly_expenses_child2=safe_money_from_value(perquisites_dto.monthly_expenses_child2),
            months_child1=perquisites_dto.months_child1,
            months_child2=perquisites_dto.months_child2,
            employer_maintained_1st_child=perquisites_dto.employer_maintained_1st_child,
            employer_maintained_2nd_child=perquisites_dto.employer_maintained_2nd_child
        )
        
        # Convert movable asset usage
        movable_asset_usage = MovableAssetUsage(
            asset_type=safe_enum_from_value(AssetType, perquisites_dto.movable_asset_type, AssetType.OTHERS),
            asset_value=safe_money_from_value(perquisites_dto.movable_asset_usage_value),
            hire_cost=safe_money_from_value(perquisites_dto.movable_asset_hire_cost),
            employee_payment=safe_money_from_value(perquisites_dto.movable_asset_employee_payment),
            is_employer_owned=perquisites_dto.movable_asset_is_employer_owned
        )
        
        # Convert movable asset transfer
        movable_asset_transfer = MovableAssetTransfer(
            asset_type=safe_enum_from_value(AssetType, perquisites_dto.movable_asset_transfer_type, AssetType.OTHERS),
            asset_cost=safe_money_from_value(perquisites_dto.movable_asset_transfer_cost),
            years_of_use=perquisites_dto.movable_asset_years_of_use,
            employee_payment=safe_money_from_value(perquisites_dto.movable_asset_transfer_employee_payment)
        )
        
        # Convert lunch refreshment
        lunch_refreshment = LunchRefreshmentPerquisite(
            employer_cost=safe_money_from_value(perquisites_dto.lunch_employer_cost),
            employee_payment=safe_money_from_value(perquisites_dto.lunch_employee_payment),
            meal_days_per_year=perquisites_dto.lunch_meal_days_per_year
        )
        
        # Convert domestic help
        domestic_help = DomesticHelpPerquisite(
            domestic_help_paid_by_employer=safe_money_from_value(perquisites_dto.domestic_help_paid_by_employer),
            domestic_help_paid_by_employee=safe_money_from_value(perquisites_dto.domestic_help_paid_by_employee)
        )
        
        # Convert gift voucher
        gift_voucher = GiftVoucherPerquisite(
            gift_voucher_amount=Money.zero()
        )
        
        # Convert monetary benefits
        monetary_benefits = MonetaryBenefitsPerquisite(
            monetary_amount_paid_by_employer=safe_money_from_value(perquisites_dto.monetary_amount_paid_by_employer),
            expenditure_for_official_purpose=safe_money_from_value(perquisites_dto.expenditure_for_official_purpose),
            amount_paid_by_employee=safe_money_from_value(perquisites_dto.amount_paid_by_employee)
        )
        
        # Convert club expenses
        club_expenses = ClubExpensesPerquisite(
            club_expenses_paid_by_employer=safe_money_from_value(perquisites_dto.club_expenses_paid_by_employer),
            club_expenses_paid_by_employee=safe_money_from_value(perquisites_dto.club_expenses_paid_by_employee),
            club_expenses_for_official_purpose=safe_money_from_value(perquisites_dto.club_expenses_for_official_purpose)
        )
        
        return Perquisites(
            accommodation=accommodation,
            car=car,
            lta=lta,
            interest_free_loan=interest_free_loan,
            esop=esop,
            utilities=utilities,
            free_education=free_education,
            movable_asset_usage=movable_asset_usage,
            movable_asset_transfer=movable_asset_transfer,
            lunch_refreshment=lunch_refreshment,
            gift_voucher=gift_voucher,
            monetary_benefits=monetary_benefits,
            club_expenses=club_expenses,
            domestic_help=domestic_help
        )
    
    def _convert_house_property_income_dto_to_entity(self, house_property_income_dto) -> HousePropertyIncome:
        """Convert house property DTO to entity."""
        from app.domain.entities.taxation.house_property_income import PropertyType
        
        # Map property type to enum
        property_type_mapping = {
            "Self-Occupied": PropertyType.SELF_OCCUPIED,
            "Let-Out": PropertyType.LET_OUT
        }
        property_type = property_type_mapping.get(house_property_income_dto.property_type, PropertyType.SELF_OCCUPIED)
        
        return HousePropertyIncome(
            property_type=property_type,
            address=getattr(house_property_income_dto, 'address', ''),
            annual_rent_received=Money.from_decimal(house_property_income_dto.annual_rent_received),
            municipal_taxes_paid=Money.from_decimal(house_property_income_dto.municipal_taxes_paid),
            home_loan_interest=Money.from_decimal(house_property_income_dto.home_loan_interest),
            pre_construction_interest=Money.from_decimal(house_property_income_dto.pre_construction_interest)
        )
    
    def _convert_capital_gains_dto_to_entity(self, capital_gains_dto) -> CapitalGainsIncome:
        """Convert capital gains DTO to entity."""
        return CapitalGainsIncome(
            stcg_111a_equity_stt=Money.from_decimal(capital_gains_dto.stcg_111a_equity_stt),
            stcg_other_assets=Money.from_decimal(capital_gains_dto.stcg_other_assets),
            stcg_debt_mf=Money.from_decimal(capital_gains_dto.stcg_debt_mf),
            ltcg_112a_equity_stt=Money.from_decimal(capital_gains_dto.ltcg_112a_equity_stt),
            ltcg_other_assets=Money.from_decimal(capital_gains_dto.ltcg_other_assets),
            ltcg_debt_mf=Money.from_decimal(capital_gains_dto.ltcg_debt_mf)
        )
    
    def _convert_retirement_benefits_dto_to_entity(self, retirement_dto) -> RetirementBenefits:
        """Convert retirement benefits DTO to entity."""
        # Handle flat structure (from frontend)
        if hasattr(retirement_dto, 'gratuity_amount'):
            # Convert flat structure to nested structure first
            flat_dto = FlatRetirementBenefitsDTO(
                gratuity_amount=retirement_dto.gratuity_amount,
                leave_encashment_amount=retirement_dto.leave_encashment_amount,
                vrs_amount=retirement_dto.vrs_amount,
                pension_amount=retirement_dto.pension_amount,
                commuted_pension_amount=retirement_dto.commuted_pension_amount,
                other_retirement_benefits=retirement_dto.other_retirement_benefits
            )
            retirement_dto = flat_dto.to_nested_structure()
        
        # Convert leave encashment
        leave_encashment = None
        if retirement_dto.leave_encashment:
            le_dto = retirement_dto.leave_encashment
            leave_encashment = LeaveEncashment(
                leave_encashment_amount=Money.from_decimal(le_dto.leave_encashment_amount),
                average_monthly_salary=Money.from_decimal(le_dto.average_monthly_salary),
                leave_days_encashed=le_dto.leave_days_encashed,
                is_deceased=le_dto.is_deceased,
                during_employment=le_dto.during_employment
            )
        
        # Convert gratuity
        gratuity = None
        if retirement_dto.gratuity:
            gr_dto = retirement_dto.gratuity
            gratuity = Gratuity(
                gratuity_amount=Money.from_decimal(gr_dto.gratuity_amount),
                monthly_salary=Money.from_decimal(gr_dto.monthly_salary),
                service_years=Decimal(str(gr_dto.service_years)),
            )
        
        # Convert VRS
        vrs = None
        if retirement_dto.vrs:
            vrs_dto = retirement_dto.vrs
            vrs = VRS(
                vrs_amount=Money.from_decimal(vrs_dto.vrs_amount),
                monthly_salary=Money.from_decimal(vrs_dto.monthly_salary),
                service_years=Decimal(str(vrs_dto.service_years))
            )
        
        # Convert pension
        pension = None
        if retirement_dto.pension:
            pen_dto = retirement_dto.pension
            pension = Pension(
                regular_pension=Money.from_decimal(pen_dto.regular_pension),
                commuted_pension=Money.from_decimal(pen_dto.commuted_pension),
                total_pension=Money.from_decimal(pen_dto.total_pension),
                gratuity_received=pen_dto.gratuity_received
            )
        
        # Convert retrenchment compensation
        retrenchment_compensation = None
        if retirement_dto.retrenchment_compensation:
            rc_dto = retirement_dto.retrenchment_compensation
            retrenchment_compensation = RetrenchmentCompensation(
                retrenchment_amount=Money.from_decimal(rc_dto.retrenchment_amount),
                monthly_salary=Money.from_decimal(rc_dto.monthly_salary),
                service_years=Decimal(str(rc_dto.service_years))
            )
        
        return RetirementBenefits(
            leave_encashment=leave_encashment,
            gratuity=gratuity,
            vrs=vrs,
            pension=pension,
            retrenchment_compensation=retrenchment_compensation
        )
    
    def _convert_other_income_dto_to_entity(self, other_income_dto) -> OtherIncome:
        """Convert other income DTO to entity."""
        # Convert interest income
        interest_income = None
        if other_income_dto.interest_income:
            int_dto = other_income_dto.interest_income
            # Handle field name mapping between frontend and backend
            # Frontend uses: savings_interest, fd_interest, rd_interest, post_office_interest
            # Backend expects: savings_account_interest, fixed_deposit_interest, recurring_deposit_interest, post_office_interest
            
            # Get values with fallback to handle both naming conventions
            savings_interest = getattr(int_dto, 'savings_account_interest', None)
            if savings_interest is None:
                savings_interest = getattr(int_dto, 'savings_interest', 0)
            
            fd_interest = getattr(int_dto, 'fixed_deposit_interest', None)
            if fd_interest is None:
                fd_interest = getattr(int_dto, 'fd_interest', 0)
            
            rd_interest = getattr(int_dto, 'recurring_deposit_interest', None)
            if rd_interest is None:
                rd_interest = getattr(int_dto, 'rd_interest', 0)
            
            post_office_interest = getattr(int_dto, 'post_office_interest', 0)
            
            interest_income = InterestIncome(
                savings_account_interest=Money.from_decimal(savings_interest),
                fixed_deposit_interest=Money.from_decimal(fd_interest),
                recurring_deposit_interest=Money.from_decimal(rd_interest),
                post_office_interest=Money.from_decimal(post_office_interest)
            )
        
        # Convert house property income if present
        house_property_income = None
        if hasattr(other_income_dto, 'house_property_income') and other_income_dto.house_property_income:
            house_property_income = self._convert_house_property_income_dto_to_entity(other_income_dto.house_property_income)
        
        # Convert capital gains income if present
        capital_gains_income = None
        if hasattr(other_income_dto, 'capital_gains_income') and other_income_dto.capital_gains_income:
            capital_gains_income = self._convert_capital_gains_dto_to_entity(other_income_dto.capital_gains_income)
        
        return OtherIncome(
            interest_income=interest_income,
            house_property_income=house_property_income,
            capital_gains_income=capital_gains_income,
            dividend_income=Money.from_decimal(other_income_dto.dividend_income),
            gifts_received=Money.from_decimal(other_income_dto.gifts_received),
            business_professional_income=Money.from_decimal(other_income_dto.business_professional_income),
            other_miscellaneous_income=Money.from_decimal(other_income_dto.other_miscellaneous_income)
        )
    
    async def _get_or_create_salary_package_record(
        self,
        employee_id: str,
        tax_year: str,
        organization_id: str,
        salary_income: SalaryIncome = None
    ) -> Tuple[SalaryPackageRecord, bool]:
        """Get existing salary package record or create a new one with defaults."""
        
        # Try to get existing record
        salary_package_record = await self.salary_package_repository.get_salary_package_record(
            employee_id, tax_year, organization_id
        )
        
        if salary_package_record:
            return salary_package_record, True
        
        # Create new record with defaults
        employee_id_vo = EmployeeId(employee_id)
        tax_year_vo = TaxYear.from_string(tax_year)
        
        # Create default salary income
        if salary_income:
            default_salary_income = salary_income
        else:
            # Use tax year boundaries for effective dates
            tax_year_start = tax_year_vo.get_start_date()
            tax_year_end = tax_year_vo.get_end_date()
            effective_from = datetime.combine(tax_year_start, datetime.min.time())
            effective_till = datetime.combine(tax_year_end, datetime.min.time())
            
            default_salary_income = SalaryIncome(
                effective_from=effective_from,
                effective_till=effective_till,
                basic_salary=Money.zero(),
                dearness_allowance=Money.zero(),
                hra_provided=Money.zero(),
                special_allowance=Money.zero(),
                bonus=Money.zero(),
                commission=Money.zero()
            )
        
        # Create default deductions
        default_deductions = TaxDeductions(
            section_80c=DeductionSection80C(),
            section_80d=DeductionSection80D(),
            section_80e=DeductionSection80E(),
            section_80g=DeductionSection80G(),
            section_80tta_ttb=DeductionSection80TTA_TTB(),
            other_deductions=OtherDeductions()
        )
        
        # Create default perquisites
        default_perquisites = self._create_default_perquisites()
        
        # Create default retirement benefits
        default_retirement_benefits = self._create_default_retirement_benefits()
        
        # Create default other income
        default_other_income = self._create_default_other_income()
        
        # Create new salary package record with correct parameters
        salary_package_record = SalaryPackageRecord(
            employee_id=employee_id_vo,
            tax_year=tax_year_vo,
            age=25,  # Default age
            regime=TaxRegime.old_regime(),  # Default to old regime
            salary_incomes=[default_salary_income],
            deductions=default_deductions,
            perquisites=default_perquisites,  # Add default perquisites
            retirement_benefits=default_retirement_benefits,  # Add default retirement benefits
            other_income=default_other_income,  # Add default other income
            organization_id=organization_id
        )
        
        return salary_package_record, False
    
    # =============================================================================
    # INDIVIDUAL COMPONENT UPDATE METHODS
    # =============================================================================
    
    async def update_salary_component(
        self,
        request: "UpdateSalaryComponentRequest",
        organization_id: str
    ) -> "ComponentUpdateResponse":
        """Update salary component individually using SalaryPackageRecord."""
        
        try:
            # Convert DTO to entity
            salary_income = self._convert_salary_dto_to_entity(request.salary_income)

            # Get or create salary package record
            salary_package_record, found_record = await self._get_or_create_salary_package_record(
                request.employee_id, request.tax_year, organization_id, salary_income
            )
            
            # Handle different modes: new revision vs update existing
            if request.force_new_revision:
                if found_record:
                    # Force create new salary revision (always add new entry)
                    salary_package_record.add_salary_income(salary_income)
            else:
                # Update mode: update existing or create first entry
                if not salary_package_record.salary_incomes:
                    # If no salary incomes exist, add the first one
                    salary_package_record.add_salary_income(salary_income)
                else:
                    # If salary incomes exist, update the latest one
                    salary_package_record.update_latest_salary_income(salary_income)
            
            salary_package_record.updated_at = datetime.utcnow()
            
            # Save to database using salary package repository
            await self.salary_package_repository.save(salary_package_record, organization_id)
            
            return ComponentUpdateResponse(
                taxation_id=salary_package_record.salary_package_id,
                employee_id=request.employee_id,
                tax_year=request.tax_year,
                component_type="salary_income",
                status="success",
                message="New salary revision created successfully" if request.force_new_revision else "Salary component updated successfully",
                updated_at=salary_package_record.updated_at,
                notes=request.notes
            )
            
        except Exception as e:
            logger.error(f"Failed to update salary component: {str(e)}")
            raise
    
    async def update_perquisites_component(
        self,
        request: "UpdatePerquisitesComponentRequest",
        organization_id: str
    ) -> "ComponentUpdateResponse":
        """Update perquisites component individually using SalaryPackageRecord."""
        
        try:
            # Convert DTO to entity
            perquisites = self._convert_perquisites_dto_to_entity(request.perquisites)
            
            # Get or create salary package record (it should ideally be present)
            salary_package_record, found_record = await self._get_or_create_salary_package_record(
                request.employee_id, request.tax_year, organization_id
            )
            
            if not found_record:
                logger.warning(f"Salary package record not found for employee {request.employee_id} in {request.tax_year}, created new one")
            
            # Update perquisites in salary package record
            salary_package_record.perquisites = perquisites
            salary_package_record.updated_at = datetime.utcnow()
            
            # Save to database using salary package repository
            await self.salary_package_repository.save(salary_package_record, organization_id)
            
            return ComponentUpdateResponse(
                taxation_id=salary_package_record.salary_package_id,
                employee_id=request.employee_id,
                tax_year=request.tax_year,
                component_type="perquisites",
                status="success",
                message="Perquisites component updated successfully",
                updated_at=salary_package_record.updated_at,
                notes=request.notes
            )
            
        except Exception as e:
            logger.error(f"Failed to update perquisites component: {str(e)}")
            raise
    
    async def update_deductions_component(
        self,
        request: "UpdateDeductionsComponentRequest",
        organization_id: str
    ) -> "ComponentUpdateResponse":
        """Update deductions component individually using SalaryPackageRecord."""
        
        try:
            logger.info(f"Starting deductions component update for employee {request.employee_id}, tax_year {request.tax_year}")
            logger.info(f"Deductions data received: {request.deductions}")
            logger.debug(f"Full request data: {request}")
            
            # Convert DTO to entity
            if request.deductions is None:
                logger.info("No deductions data provided, creating default deductions")
                deductions = self._create_default_deductions()
            else:
                logger.info("Converting deductions DTO to entity")
                deductions = self._convert_comprehensive_deductions_dto_to_entity(request.deductions)
                logger.debug(f"Converted deductions entity - Section 80C total: {deductions.section_80c.calculate_total_investment()}")

            # Get or create salary package record (it should ideally be present)
            logger.info(f"Getting or creating salary package record for employee {request.employee_id}")
            salary_package_record, found_record = await self._get_or_create_salary_package_record(
                request.employee_id, request.tax_year, organization_id
            )
            
            if not found_record:
                logger.warning(f"Salary package record not found for employee {request.employee_id} in {request.tax_year}, created new one")
            else:
                logger.info(f"Found existing salary package record with ID: {salary_package_record.salary_package_id}")

            # Log existing deductions before update
            # Calculate gross income for deduction calculations
            gross_income = salary_package_record.calculate_gross_income()
            existing_deductions_total = salary_package_record.deductions.calculate_total_deductions(salary_package_record.regime, salary_package_record.age, gross_income)
            logger.info(f"Existing deductions total before update: {existing_deductions_total}")
            
            # Update deductions using the salary package record's method
            logger.info("Updating deductions on salary package record")
            salary_package_record.update_deductions(deductions)
            salary_package_record.updated_at = datetime.utcnow()
            
            # Log new deductions after update
            new_deductions_total = salary_package_record.deductions.calculate_total_deductions(salary_package_record.regime, salary_package_record.age, gross_income) 
            logger.info(f"New deductions total after update: {new_deductions_total}")
            
            # Save to database using salary package repository
            logger.info(f"Saving salary package record to database for organization {organization_id}")
            saved_record = await self.salary_package_repository.save(salary_package_record, organization_id)
            logger.info(f"Successfully saved salary package record. Returned record ID: {saved_record.salary_package_id}")
            
            # Verify the save by attempting to read back
            logger.info("Verifying save operation by reading back the record")
            verification_record = await self.salary_package_repository.get_salary_package_record(
                request.employee_id, request.tax_year, organization_id
            )
            if verification_record:
                # Calculate gross income for verification record
                verify_gross_income = verification_record.calculate_gross_income()
                verify_deductions_total = verification_record.deductions.calculate_total_deductions(verification_record.regime, verification_record.age, verify_gross_income)
                logger.info(f"Verification successful - Deductions total in database: {verify_deductions_total}")
                if verify_deductions_total != new_deductions_total:
                    logger.error(f"MISMATCH: Expected {new_deductions_total}, but found {verify_deductions_total} in database")
            else:
                logger.error("VERIFICATION FAILED: Could not read back the saved record from database")
            
            return ComponentUpdateResponse(
                taxation_id=salary_package_record.salary_package_id,
                employee_id=request.employee_id,
                tax_year=request.tax_year,
                component_type="deductions",
                status="success",
                message="Deductions component updated successfully",
                updated_at=salary_package_record.updated_at,
                notes=request.notes
            )
            
        except Exception as e:
            logger.error(f"Failed to update deductions component: {str(e)}")
            raise
    
    async def update_house_property_component(
        self,
        request: "UpdateHousePropertyComponentRequest",
        organization_id: str
    ) -> "ComponentUpdateResponse":
        """Update house property component individually using SalaryPackageRecord."""
        
        try:
            # Convert DTO to entity
            house_property_income = self._convert_house_property_income_dto_to_entity(request.house_property_income)
            
            # Get or create salary package record (it should ideally be present)
            salary_package_record, found_record = await self._get_or_create_salary_package_record(
                request.employee_id, request.tax_year, organization_id
            )
            
            if not found_record:
                logger.warning(f"Salary package record not found for employee {request.employee_id} in {request.tax_year}, created new one")
            
            # Create or update other_income
            if not salary_package_record.other_income:
                salary_package_record.other_income = self._create_default_other_income()
            
            # Update house property in other income
            salary_package_record.other_income.house_property_income = house_property_income
            salary_package_record.updated_at = datetime.utcnow()
            
            # Save to database using salary package repository
            await self.salary_package_repository.save(salary_package_record, organization_id)
            
            return ComponentUpdateResponse(
                taxation_id=salary_package_record.salary_package_id,
                employee_id=request.employee_id,
                tax_year=request.tax_year,
                component_type="house_property_income",
                status="success",
                message="House property component updated successfully",
                updated_at=salary_package_record.updated_at,
                notes=request.notes
            )
            
        except Exception as e:
            logger.error(f"Failed to update house property component: {str(e)}")
            raise
    
    async def update_capital_gains_component(
        self,
        request: "UpdateCapitalGainsComponentRequest",
        organization_id: str
    ) -> "ComponentUpdateResponse":
        """Update capital gains component individually using SalaryPackageRecord."""
        
        try:
            # Convert DTO to entity
            capital_gains = self._convert_capital_gains_dto_to_entity(request.capital_gains_income)
            
            # Get or create salary package record (it should ideally be present)
            salary_package_record, found_record = await self._get_or_create_salary_package_record(
                request.employee_id, request.tax_year, organization_id
            )
            
            if not found_record:
                logger.warning(f"Salary package record not found for employee {request.employee_id} in {request.tax_year}, created new one")
            
            # Create or update other_income
            if not salary_package_record.other_income:
                salary_package_record.other_income = self._create_default_other_income()
            
            # Update capital gains in other income
            salary_package_record.other_income.capital_gains_income = capital_gains
            salary_package_record.updated_at = datetime.utcnow()
            
            # Save to database using salary package repository
            await self.salary_package_repository.save(salary_package_record, organization_id)
            
            return ComponentUpdateResponse(
                taxation_id=salary_package_record.salary_package_id,
                employee_id=request.employee_id,
                tax_year=request.tax_year,
                component_type="capital_gains_income",
                status="success",
                message="Capital gains component updated successfully",
                updated_at=salary_package_record.updated_at,
                notes=request.notes
            )
            
        except Exception as e:
            logger.error(f"Failed to update capital gains component: {str(e)}")
            raise
    
    async def update_retirement_benefits_component(
        self,
        request: "UpdateRetirementBenefitsComponentRequest",
        organization_id: str
    ) -> "ComponentUpdateResponse":
        """Update retirement benefits component individually using SalaryPackageRecord."""
        
        try:
            # Convert DTO to entity
            retirement_benefits = self._convert_retirement_benefits_dto_to_entity(request.retirement_benefits)
            
            # Get or create salary package record (it should ideally be present)
            salary_package_record, found_record = await self._get_or_create_salary_package_record(
                request.employee_id, request.tax_year, organization_id
            )
            
            if not found_record:
                logger.warning(f"Salary package record not found for employee {request.employee_id} in {request.tax_year}, created new one")
            
            # Update retirement benefits in salary package record
            salary_package_record.retirement_benefits = retirement_benefits
            salary_package_record.updated_at = datetime.utcnow()
            
            # Save to database using salary package repository
            await self.salary_package_repository.save(salary_package_record, organization_id)
            
            return ComponentUpdateResponse(
                taxation_id=salary_package_record.salary_package_id,
                employee_id=request.employee_id,
                tax_year=request.tax_year,
                component_type="retirement_benefits",
                status="success",
                message="Retirement benefits component updated successfully",
                updated_at=salary_package_record.updated_at,
                notes=request.notes
            )
            
        except Exception as e:
            logger.error(f"Failed to update retirement benefits component: {str(e)}")
            raise
    
    async def update_other_income_component(
        self,
        request: "UpdateOtherIncomeComponentRequest",
        organization_id: str
    ) -> "ComponentUpdateResponse":
        """Update other income component individually using SalaryPackageRecord."""
        
        try:
            # Convert DTO to entity
            other_income = self._convert_other_income_dto_to_entity(request.other_income)
            
            # Get or create salary package record (it should ideally be present)
            salary_package_record, found_record = await self._get_or_create_salary_package_record(
                request.employee_id, request.tax_year, organization_id
            )
            
            if not found_record:
                logger.warning(f"Salary package record not found for employee {request.employee_id} in {request.tax_year}, created new one")
            
            # Update other income in salary package record
            salary_package_record.other_income = other_income
            salary_package_record.updated_at = datetime.utcnow()
            
            # Save to database using salary package repository
            await self.salary_package_repository.save(salary_package_record, organization_id)
            
            return ComponentUpdateResponse(
                taxation_id=salary_package_record.salary_package_id,
                employee_id=request.employee_id,
                tax_year=request.tax_year,
                component_type="other_income",
                status="success",
                message="Other income component updated successfully",
                updated_at=salary_package_record.updated_at,
                notes=request.notes
            )
            
        except Exception as e:
            logger.error(f"Failed to update other income component: {str(e)}")
            raise
    
    async def update_monthly_payroll_component(
        self,
        request: "UpdateMonthlyPayrollComponentRequest",
        organization_id: str
    ) -> "ComponentUpdateResponse":
        """Update monthly payroll component individually."""
        
        try:
            # Get or create taxation record
            taxation_record = await self._get_or_create_taxation_record(
                request.employee_id, request.tax_year, organization_id
            )
            
            # Convert DTO to entity
            monthly_payroll = self._convert_monthly_payroll_dto_to_entity(request.monthly_payroll)
            
            # Update the record
            taxation_record.monthly_payroll = monthly_payroll
            taxation_record.updated_at = datetime.utcnow()
            
            # Save to database
            await self.taxation_repository.save(taxation_record, organization_id)
            
            return ComponentUpdateResponse(
                taxation_id=taxation_record.taxation_id,
                employee_id=request.employee_id,
                tax_year=request.tax_year,
                component_type="monthly_payroll",
                status="success",
                message="Monthly payroll component updated successfully",
                updated_at=taxation_record.updated_at,
                notes=request.notes
            )
            
        except Exception as e:
            logger.error(f"Failed to update monthly payroll component: {str(e)}")
            raise
    

    async def is_regime_update_allowed(
        self,
        request: "IsRegimeUpdateAllowedRequest",
        organization_id: str
    ) -> "IsRegimeUpdateAllowedResponse":
        """Check if regime update is allowed."""

        try:
            # Get or create salary package record (it should ideally be present)
            salary_package_record, found_record = await self._get_or_create_salary_package_record(
                request.employee_id, request.tax_year, organization_id
            )
            
            return IsRegimeUpdateAllowedResponse(
                is_allowed=salary_package_record.is_regime_update_allowed,
                regime_type=salary_package_record.regime.regime_type.value,
                message="Regime update is allowed" if salary_package_record.is_regime_update_allowed else "Regime update is not allowed"
            )
            
        except Exception as e:
            logger.error(f"Failed to check if regime update is allowed: {str(e)}")
            raise
    
    async def update_regime_component(
        self,
        request: "UpdateRegimeComponentRequest",
        organization_id: str
    ) -> "ComponentUpdateResponse":
        """Update tax regime component individually."""
        
        try:
            # Get or create salary package record (it should ideally be present)
            salary_package_record, found_record = await self._get_or_create_salary_package_record(
                request.employee_id, request.tax_year, organization_id
            )

            if not salary_package_record.is_regime_update_allowed:
                raise ValueError("Regime update is not allowed")
            
            # Update regime and age
            regime_type = TaxRegimeType(request.regime_type)
            salary_package_record.regime = TaxRegime(regime_type)
            salary_package_record.age = request.age
            salary_package_record.updated_at = datetime.utcnow()
            salary_package_record.is_regime_update_allowed = False
            
            # Save to database
            await self.salary_package_repository.save(salary_package_record, organization_id)

            return ComponentUpdateResponse( 
                taxation_id=salary_package_record.salary_package_id,
                employee_id=request.employee_id,
                tax_year=request.tax_year,
                component_type="regime",
                status="success",
                message="Tax regime component updated successfully",
                updated_at=salary_package_record.updated_at,
                notes=request.notes
            )
            
        except Exception as e:
            logger.error(f"Failed to update regime component: {str(e)}")
            raise
    
    async def get_component(
        self,
        employee_id: str,
        tax_year: str,
        component_type: str,
        organization_id: str
    ) -> "ComponentResponse":
        """Get a specific component from salary package record."""
        
        try:
            # Get or create salary package record instead of taxation record
            salary_package_record, found_record = await self._get_or_create_salary_package_record(
                employee_id, tax_year, organization_id
            )
            
            # Extract component data based on type
            component_data = self._extract_component_data_from_salary_package(salary_package_record, component_type)
            
            return ComponentResponse(
                taxation_id=salary_package_record.salary_package_id,
                employee_id=employee_id,
                tax_year=tax_year,
                component_type=component_type,
                component_data=component_data,
                last_updated=salary_package_record.updated_at,
                notes=None  # Could be enhanced to store component-specific notes
            )
            
        except Exception as e:
            logger.error(f"Failed to get component {component_type}: {str(e)}")
            raise
    
    # async def get_taxation_record_status(
    #     self,
    #     employee_id: str,
    #     tax_year: str,
    #     organization_id: str
    # ) -> "TaxationRecordStatusResponse":
    #     """Get status of all components in a taxation record."""
        
    #     try:
    #         # Get taxation record
    #         taxation_record = await self.taxation_repository.get_taxation_record(
    #             employee_id, tax_year, organization_id
    #         )
            
    #         if not taxation_record:
    #             raise ValueError(f"Taxation record not found for employee {employee_id} and tax year {tax_year}")
            
    #         # Build components status
    #         components_status = self._build_components_status(taxation_record)
            
    #         # Determine overall status
    #         overall_status = self._determine_overall_status(components_status)
            
    #         return TaxationRecordStatusResponse(
    #             taxation_id=taxation_record.taxation_id,
    #             employee_id=employee_id,
    #             tax_year=tax_year,
    #             regime_type=taxation_record.regime.regime_type.value,
    #             age=taxation_record.age,
    #             components_status=components_status,
    #             overall_status=overall_status,
    #             last_updated=taxation_record.updated_at,
    #             is_final=taxation_record.is_final
    #         )
            
    #     except Exception as e:
    #         logger.error(f"Failed to get taxation record status: {str(e)}")
    #         raise
    
    def _extract_component_data_from_salary_package(self, salary_package_record: SalaryPackageRecord, component_type: str) -> Dict[str, Any]:
        """Extract component data from salary package record."""
        
        # Handle frontend component type aliases
        component_type_mapping = {
            "salary": "salary_income",
            "house_property_income": "house_property_income",
            "capital_gains": "capital_gains_income",
            "retirement_benefits": "retirement_benefits",
            "other_income": "other_income",
            "monthly_payroll": "monthly_payroll"
        }
        
        # Map frontend component type to backend component type
        mapped_component_type = component_type_mapping.get(component_type, component_type)
        
        if mapped_component_type == "salary_income":
            # For salary package record, we use the latest salary income
            latest_salary = salary_package_record.get_latest_salary_income()
            return self._serialize_salary_income_with_history(salary_package_record)
        elif mapped_component_type == "perquisites":
            return self._serialize_perquisites(salary_package_record.perquisites) if salary_package_record.perquisites else {}
        elif mapped_component_type == "deductions":
            return self._serialize_deductions(salary_package_record.deductions)
        elif mapped_component_type == "house_property_income":
            if salary_package_record.other_income and salary_package_record.other_income.house_property_income:
                return self._serialize_house_property_income(salary_package_record.other_income.house_property_income)
            return {}
        elif mapped_component_type == "capital_gains_income":
            if salary_package_record.other_income and salary_package_record.other_income.capital_gains_income:
                return self._serialize_capital_gains_income(salary_package_record.other_income.capital_gains_income)
            return {}
        elif mapped_component_type == "retirement_benefits":
            return self._serialize_retirement_benefits(salary_package_record.retirement_benefits) if salary_package_record.retirement_benefits else {}
        elif mapped_component_type == "other_income":
            return self._serialize_other_income(salary_package_record.other_income) if salary_package_record.other_income else {}
        elif mapped_component_type == "monthly_payroll":
            # Monthly payroll is not supported in salary package record, return empty
            return {}
        elif mapped_component_type == "regime":
            return {
                "regime_type": salary_package_record.regime.regime_type.value,
                "age": salary_package_record.age
            }
        else:
            raise ValueError(f"Unknown component type: {component_type}")
    
    def _serialize_salary_income_with_history(self, salary_package_record: SalaryPackageRecord) -> Dict[str, Any]:
        """Serialize salary income with history from salary package record."""
        latest_salary = salary_package_record.get_latest_salary_income()
        
        # Get basic salary serialization
        salary_data = self._serialize_salary_income(latest_salary)
        
        # Add salary history information
        salary_data["salary_history"] = [
            {
                "index": i,
                "gross_salary": salary.calculate_gross_salary().to_float(),
                "basic_salary": salary.basic_salary.to_float(),
                "special_allowance": salary.special_allowance.to_float(),
                "hra_provided": salary.hra_provided.to_float(),
                "bonus": salary.bonus.to_float(),
                "commission": salary.commission.to_float()
            }
            for i, salary in enumerate(salary_package_record.salary_incomes)
        ]
        
        salary_data["salary_incomes_count"] = len(salary_package_record.salary_incomes)
        salary_data["is_latest"] = True
        
        return salary_data
    
    def _determine_overall_status(self, components_status: Dict[str, Dict[str, Any]]) -> str:
        """Determine overall status based on component statuses."""
        
        required_components = ["salary", "deductions", "regime"]
        optional_components = ["perquisites", "house_property_income", "capital-gains", 
                             "retirement-benefits", "other-income", "monthly-payroll"]
        
        # Check required components
        for component in required_components:
            if component not in components_status or components_status[component]["status"] != "complete":
                return "incomplete"
        
        # Count optional components with data
        optional_with_data = sum(
            1 for component in optional_components
            if component in components_status and components_status[component]["status"] == "complete"
        )
        
        if optional_with_data == 0:
            return "basic_complete"
        elif optional_with_data >= 3:
            return "comprehensive_complete"
        else:
            return "partial_complete"
    
    # Serialization helper methods for components
    def _serialize_salary_income(self, salary_income: SalaryIncome) -> Dict[str, Any]:
        """Serialize salary income to dict."""
        if not salary_income:
            return {}
        
        # Base salary components
        result = {
            "basic_salary": float(salary_income.basic_salary.amount),
            "dearness_allowance": float(salary_income.dearness_allowance.amount),
            "hra_provided": float(salary_income.hra_provided.amount),
            "special_allowance": float(salary_income.special_allowance.amount),
            "bonus": float(salary_income.bonus.amount),
            "commission": float(salary_income.commission.amount),
            # HRA details are now in deductions module, not salary
            "hra_city_type": "metro",  # Default value for frontend compatibility
            "actual_rent_paid": 0.0,   # Default value for frontend compatibility
        }
        
        # Add specific allowances if available
        if salary_income.specific_allowances:
            specific_allowances_data = {
                "city_compensatory_allowance": float(salary_income.specific_allowances.city_compensatory_allowance.amount),
                "rural_allowance": float(salary_income.specific_allowances.rural_allowance.amount),
                "proctorship_allowance": float(salary_income.specific_allowances.proctorship_allowance.amount),
                "wardenship_allowance": float(salary_income.specific_allowances.wardenship_allowance.amount),
                "project_allowance": float(salary_income.specific_allowances.project_allowance.amount),
                "deputation_allowance": float(salary_income.specific_allowances.deputation_allowance.amount),
                "interim_relief": float(salary_income.specific_allowances.interim_relief.amount),
                "tiffin_allowance": float(salary_income.specific_allowances.tiffin_allowance.amount),
                "overtime_allowance": float(salary_income.specific_allowances.overtime_allowance.amount),
                "servant_allowance": float(salary_income.specific_allowances.servant_allowance.amount),
                "hills_high_altd_allowance": float(salary_income.specific_allowances.hills_allowance.amount),
                "border_remote_allowance": float(salary_income.specific_allowances.border_allowance.amount),
                "transport_employee_allowance": float(salary_income.specific_allowances.transport_employee_allowance.amount),
                "children_education_allowance": float(salary_income.specific_allowances.children_education_allowance.amount),
                "hostel_allowance": float(salary_income.specific_allowances.hostel_allowance.amount),
                "underground_mines_allowance": float(salary_income.specific_allowances.underground_mines_allowance.amount),
                "govt_employee_entertainment_allowance": float(salary_income.specific_allowances.government_entertainment_allowance.amount),
                "supreme_high_court_judges_allowance": float(salary_income.specific_allowances.supreme_high_court_judges_allowance.amount),
                "judge_compensatory_allowance": float(salary_income.specific_allowances.judge_compensatory_allowance.amount),
                "section_10_14_special_allowances": float(salary_income.specific_allowances.section_10_14_special_allowances.amount),
                "travel_on_tour_allowance": float(salary_income.specific_allowances.travel_on_tour_allowance.amount),
                "tour_daily_charge_allowance": float(salary_income.specific_allowances.tour_daily_charge_allowance.amount),
                "conveyance_in_performace_of_duties": float(salary_income.specific_allowances.conveyance_in_performace_of_duties.amount),
                "helper_in_performace_of_duties": float(salary_income.specific_allowances.helper_in_performace_of_duties.amount),
                "academic_research": float(salary_income.specific_allowances.academic_research.amount),
                "uniform_allowance": float(salary_income.specific_allowances.uniform_allowance.amount),
                "any_other_allowance_exemption": float(salary_income.specific_allowances.any_other_allowance_exemption.amount),
                "any_other_allowance": float(salary_income.specific_allowances.any_other_allowance.amount),
                "hills_exemption_limit": float(salary_income.specific_allowances.hills_exemption_limit.amount),
                "border_exemption_limit": float(salary_income.specific_allowances.border_exemption_limit.amount),
                "children_count": salary_income.specific_allowances.children_count,
                "disabled_transport_allowance": float(salary_income.specific_allowances.disabled_transport_allowance.amount),
                "is_disabled": salary_income.specific_allowances.is_disabled,
                "mine_work_months": salary_income.specific_allowances.mine_work_months,
                "fixed_medical_allowance": float(salary_income.specific_allowances.fixed_medical_allowance.amount),
                "govt_employees_outside_india_allowance": float(salary_income.specific_allowances.govt_employees_outside_india_allowance.amount)
            }
            result.update(specific_allowances_data)
        else:
            # Add default zero values for all specific allowances if not available
            default_allowances = {
                "city_compensatory_allowance": 0.0,
                "rural_allowance": 0.0,
                "proctorship_allowance": 0.0,
                "wardenship_allowance": 0.0,
                "project_allowance": 0.0,
                "deputation_allowance": 0.0,
                "interim_relief": 0.0,
                "tiffin_allowance": 0.0,
                "overtime_allowance": 0.0,
                "servant_allowance": 0.0,
                "hills_high_altd_allowance": 0.0,
                "border_remote_allowance": 0.0,
                "transport_employee_allowance": 0.0,
                "children_education_allowance": 0.0,
                "hostel_allowance": 0.0,
                "underground_mines_allowance": 0.0,
                "govt_employee_entertainment_allowance": 0.0,
                "supreme_high_court_judges_allowance": 0.0,
                "judge_compensatory_allowance": 0.0,
                "section_10_14_special_allowances": 0.0,
                "travel_on_tour_allowance": 0.0,
                "tour_daily_charge_allowance": 0.0,
                "conveyance_in_performace_of_duties": 0.0,
                "helper_in_performace_of_duties": 0.0,
                "academic_research": 0.0,
                "uniform_allowance": 0.0,
                "any_other_allowance_exemption": 0.0,
                "any_other_allowance": 0.0,
                "hills_exemption_limit": 0.0,
                "border_exemption_limit": 0.0,
                "children_count": 0,
                "disabled_transport_allowance": 0.0,
                "is_disabled": False,
                "mine_work_months": 0,
                "fixed_medical_allowance": 0.0,
                "govt_employees_outside_india_allowance": 0.0
            }
            result.update(default_allowances)
        
        return result
    
    def _serialize_perquisites(self, perquisites: Perquisites) -> Dict[str, Any]:
        """Serialize perquisites to dict."""
        if not perquisites:
            return {}
        
        result = {}
        
        # Serialize accommodation
        if perquisites.accommodation:
            acc = perquisites.accommodation
            result["accommodation"] = {
                "accommodation_type": acc.accommodation_type.value,
                "city_population": acc.city_population.value,
                "license_fees": float(acc.license_fees.amount),
                "employee_rent_payment": float(acc.employee_rent_payment.amount),
                "rent_paid_by_employer": float(acc.rent_paid_by_employer.amount),
                "hotel_charges": float(acc.hotel_charges.amount),
                "stay_days": acc.stay_days,
                "furniture_cost": float(acc.furniture_cost.amount),
                "furniture_employee_payment": float(acc.furniture_employee_payment.amount),
                "is_furniture_owned_by_employer": acc.is_furniture_owned_by_employer
            }
        
        # Serialize car
        if perquisites.car:
            car = perquisites.car
            result["car"] = {
                "car_use_type": car.car_use_type.value,
                "engine_capacity_cc": car.engine_capacity_cc,
                "months_used": car.months_used,
                "months_used_other_vehicle": car.months_used_other_vehicle,
                "car_cost_to_employer": float(car.car_cost_to_employer.amount),
                "other_vehicle_cost": float(car.other_vehicle_cost.amount),
                "driver_cost": float(car.driver_cost.amount),
                "has_expense_reimbursement": car.has_expense_reimbursement,
                "driver_provided": car.driver_provided
            }
        
        # Serialize LTA
        if perquisites.lta:
            lta = perquisites.lta
            result["lta"] = {
                "lta_allocated_yearly": float(lta.lta_allocated_yearly.amount),
                "lta_amount_claimed": float(lta.lta_amount_claimed.amount),
                "lta_claimed_count": lta.lta_claimed_count,
                "public_transport_cost": float(lta.public_transport_cost.amount),
                "travel_mode": lta.travel_mode,
                "is_monthly_paid": lta.is_monthly_paid
            }
        
        # Serialize interest free loan
        if perquisites.interest_free_loan:
            loan = perquisites.interest_free_loan
            result["interest_free_loan"] = {
                "loan_amount": float(loan.loan_amount.amount),
                "emi_amount": float(loan.emi_amount.amount),
                "outstanding_amount": float(loan.outstanding_amount.amount),
                "company_interest_rate": float(loan.company_interest_rate),
                "sbi_interest_rate": float(loan.sbi_interest_rate),
                "loan_start_date": loan.loan_start_date.isoformat() if loan.loan_start_date else None,
            }
        
        # Serialize ESOP
        if perquisites.esop:
            esop = perquisites.esop
            result["esop"] = {
                "shares_exercised": esop.shares_exercised,
                "exercise_price": float(esop.exercise_price.amount),
                "allotment_price": float(esop.allotment_price.amount)
            }
        
        # Serialize utilities
        if perquisites.utilities:
            util = perquisites.utilities
            result["utilities"] = {
                "gas_paid_by_employer": float(util.gas_paid_by_employer.amount),
                "electricity_paid_by_employer": float(util.electricity_paid_by_employer.amount),
                "water_paid_by_employer": float(util.water_paid_by_employer.amount),
                "gas_paid_by_employee": float(util.gas_paid_by_employee.amount),
                "electricity_paid_by_employee": float(util.electricity_paid_by_employee.amount),
                "water_paid_by_employee": float(util.water_paid_by_employee.amount),
                "is_gas_manufactured_by_employer": util.is_gas_manufactured_by_employer,
                "is_electricity_manufactured_by_employer": util.is_electricity_manufactured_by_employer,
                "is_water_manufactured_by_employer": util.is_water_manufactured_by_employer
            }
        
        # Serialize free education
        if perquisites.free_education:
            edu = perquisites.free_education
            result["free_education"] = {
                "monthly_expenses_child1": float(edu.monthly_expenses_child1.amount),
                "monthly_expenses_child2": float(edu.monthly_expenses_child2.amount),
                "months_child1": edu.months_child1,
                "months_child2": edu.months_child2,
                "employer_maintained_1st_child": edu.employer_maintained_1st_child,
                "employer_maintained_2nd_child": edu.employer_maintained_2nd_child
            }
        
        # Serialize lunch refreshment
        if perquisites.lunch_refreshment:
            lunch = perquisites.lunch_refreshment
            result["lunch_refreshment"] = {
                "employer_cost": float(lunch.employer_cost.amount),
                "employee_payment": float(lunch.employee_payment.amount),
                "meal_days_per_year": lunch.meal_days_per_year
            }
        
        # Serialize domestic help
        if perquisites.domestic_help:
            help_obj = perquisites.domestic_help
            result["domestic_help"] = {
                "domestic_help_paid_by_employer": float(help_obj.domestic_help_paid_by_employer.amount),
                "domestic_help_paid_by_employee": float(help_obj.domestic_help_paid_by_employee.amount)
            }
        
        return result
    
    def _serialize_deductions(self, deductions: TaxDeductions) -> Dict[str, Any]:
        """Serialize deductions to dict with comprehensive breakdown."""
        if not deductions:
            return {
                "hra_exemption": {"actual_rent_paid": 0.0, "hra_city_type": "non_metro"},
                "section_80c": {
                    "life_insurance_premium": 0.0, "epf_contribution": 0.0, "ppf_contribution": 0.0,
                    "nsc_investment": 0.0, "tax_saving_fd": 0.0, "elss_investment": 0.0,
                    "home_loan_principal": 0.0, "tuition_fees": 0.0, "ulip_premium": 0.0,
                    "sukanya_samriddhi": 0.0, "stamp_duty_property": 0.0, "senior_citizen_savings": 0.0,
                    "other_80c_investments": 0.0, "total_invested": 0.0, "limit": 150000, "remaining_limit": 150000
                },
                "section_80ccc": {"pension_fund_contribution": 0.0},
                "section_80ccd": {
                    "employee_nps_contribution": 0.0, "additional_nps_contribution": 0.0,
                    "employer_nps_contribution": 0.0, "limit_80ccd_1b": 50000
                },
                "section_80d": {
                    "self_family_premium": 0.0, "parent_premium": 0.0, "preventive_health_checkup": 0.0,
                    "parent_age": 55, "self_family_limit": 25000, "parent_limit": 25000, "preventive_limit": 5000
                },
                "section_80dd": {"relation": None, "disability_percentage": None, "eligible_deduction": 0.0},
                "section_80ddb": {"dependent_age": 0, "medical_expenses": 0.0, "relation": None, "eligible_deduction": 0.0},
                "section_80e": {"education_loan_interest": 0.0, "relation": None},
                "section_80eeb": {"ev_loan_interest": 0.0, "ev_purchase_date": None, "eligible_deduction": 0.0},
                "section_80g": {
                    "pm_relief_fund": 0.0, "national_defence_fund": 0.0, "national_foundation_communal_harmony": 0.0,
                    "zila_saksharta_samiti": 0.0, "national_illness_assistance_fund": 0.0, "national_blood_transfusion_council": 0.0,
                    "national_trust_autism_fund": 0.0, "national_sports_fund": 0.0, "national_cultural_fund": 0.0,
                    "technology_development_fund": 0.0, "national_children_fund": 0.0, "cm_relief_fund": 0.0,
                    "army_naval_air_force_funds": 0.0, "swachh_bharat_kosh": 0.0, "clean_ganga_fund": 0.0,
                    "drug_abuse_control_fund": 0.0, "other_100_percent_wo_limit": 0.0, "jn_memorial_fund": 0.0,
                    "pm_drought_relief": 0.0, "indira_gandhi_memorial_trust": 0.0, "rajiv_gandhi_foundation": 0.0,
                    "other_50_percent_wo_limit": 0.0, "family_planning_donation": 0.0, "indian_olympic_association": 0.0,
                    "other_100_percent_w_limit": 0.0, "govt_charitable_donations": 0.0, "housing_authorities_donations": 0.0,
                    "religious_renovation_donations": 0.0, "other_charitable_donations": 0.0, "other_50_percent_w_limit": 0.0,
                    "total_donations": 0.0
                },
                "section_80ggc": {"political_party_contribution": 0.0},
                "section_80u": {"disability_percentage": None, "eligible_deduction": 0.0},
                "section_80tta_ttb": {
                    "savings_interest": 0.0, "fd_interest": 0.0, "rd_interest": 0.0, "post_office_interest": 0.0,
                    "age": 25, "applicable_section": "80TTA", "exemption_limit": 10000, "eligible_exemption": 0.0
                },
                "other_deductions": {
                    "education_loan_interest": 0.0, "charitable_donations": 0.0, "savings_interest": 0.0,
                    "nps_contribution": 0.0, "other_deductions": 0.0, "total": 0.0
                },
                "summary": {"total_deductions": 0.0, "total_interest_exemptions": 0.0, "combined_80c_80ccc_80ccd1": 0.0, "regime_applicable": "old"}
            }
        
        # Create comprehensive deductions structure
        result = {
            # HRA Exemption
            "hra_exemption": {
                "actual_rent_paid": float(deductions.hra_exemption.actual_rent_paid.amount) if deductions.hra_exemption else 0.0,
                "hra_city_type": deductions.hra_exemption.hra_city_type if deductions.hra_exemption else "non_metro"
            },
            
            # Section 80C - Detailed breakdown
            "section_80c": self._serialize_section_80c(deductions.section_80c),
            
            # Section 80CCC
            "section_80ccc": {
                "pension_fund_contribution": float(deductions.section_80ccc.pension_fund_contribution.amount) if deductions.section_80ccc else 0.0
            },
            
            # Section 80CCD - NPS contributions
            "section_80ccd": {
                "employee_nps_contribution": float(deductions.section_80ccd.employee_nps_contribution.amount) if deductions.section_80ccd else 0.0,
                "additional_nps_contribution": float(deductions.section_80ccd.additional_nps_contribution.amount) if deductions.section_80ccd else 0.0,
                "employer_nps_contribution": float(deductions.section_80ccd.employer_nps_contribution.amount) if deductions.section_80ccd else 0.0,
                "limit_80ccd_1b": 50000
            },
            
            # Section 80D - Health insurance
            "section_80d": self._serialize_section_80d(deductions.section_80d),
            
            # Section 80DD - Disability (Dependent)
            "section_80dd": self._serialize_section_80dd(deductions.section_80dd),
            
            # Section 80DDB - Medical treatment
            "section_80ddb": self._serialize_section_80ddb(deductions.section_80ddb),
            
            # Section 80E - Education loan interest
            "section_80e": {
                "education_loan_interest": float(deductions.section_80e.education_loan_interest.amount) if deductions.section_80e else 0.0,
                "relation": deductions.section_80e.relation.value if deductions.section_80e and deductions.section_80e.relation else None
            },
            
            # Section 80EEB - Electric vehicle loan interest
            "section_80eeb": self._serialize_section_80eeb(deductions.section_80eeb),
            
            # Section 80G - Charitable donations
            "section_80g": self._serialize_section_80g(deductions.section_80g),
            
            # Section 80GGC - Political party contributions
            "section_80ggc": {
                "political_party_contribution": float(deductions.section_80ggc.political_party_contribution.amount) if deductions.section_80ggc else 0.0
            },
            
            # Section 80U - Self disability
            "section_80u": {
                "disability_percentage": deductions.section_80u.disability_percentage.value if deductions.section_80u and deductions.section_80u.disability_percentage else None,
                "eligible_deduction": self._calculate_80u_deduction(deductions.section_80u)
            },
            
            # Section 80TTA/TTB - Interest exemptions
            "section_80tta_ttb": self._serialize_section_80tta_ttb(deductions.section_80tta_ttb),
            
            # Other deductions
            "other_deductions": {
                "education_loan_interest": float(deductions.education_loan_interest.amount) if deductions.education_loan_interest else 0.0,
                "charitable_donations": float(deductions.donations_80g.amount) if deductions.donations_80g else 0.0,
                "savings_interest": float(deductions.savings_account_interest.amount) if deductions.savings_account_interest else 0.0,
                "nps_contribution": float(deductions.section_80ccd.employee_nps_contribution.amount) if deductions.section_80ccd else 0.0,
                "other_deductions": float(deductions.other_deductions.other_deductions.amount) if deductions.other_deductions else 0.0,
                "total": float(deductions.other_deductions.calculate_total().amount) if deductions.other_deductions else 0.0
            },
            
            # Summary totals
            "summary": self._calculate_deductions_summary(deductions)
        }
        
        return result
    
    def _serialize_section_80c(self, section_80c) -> Dict[str, Any]:
        """Serialize Section 80C details."""
        if not section_80c:
            return {
                "life_insurance_premium": 0.0, "epf_contribution": 0.0, "ppf_contribution": 0.0,
                "nsc_investment": 0.0, "tax_saving_fd": 0.0, "elss_investment": 0.0,
                "home_loan_principal": 0.0, "tuition_fees": 0.0, "ulip_premium": 0.0,
                "sukanya_samriddhi": 0.0, "stamp_duty_property": 0.0, "senior_citizen_savings": 0.0,
                "other_80c_investments": 0.0, "total_invested": 0.0, "limit": 150000, "remaining_limit": 150000
            }
        
        total_invested = float(section_80c.calculate_total_investment().amount)
        limit = 150000
        remaining_limit = max(0, limit - total_invested)
        
        return {
            "life_insurance_premium": float(section_80c.life_insurance_premium.amount),
            "epf_contribution": float(section_80c.epf_contribution.amount),
            "ppf_contribution": float(section_80c.ppf_contribution.amount),
            "nsc_investment": float(section_80c.nsc_investment.amount),
            "tax_saving_fd": float(section_80c.tax_saving_fd.amount),
            "elss_investment": float(section_80c.elss_investment.amount),
            "home_loan_principal": float(section_80c.home_loan_principal.amount),
            "tuition_fees": float(section_80c.tuition_fees.amount),
            "ulip_premium": float(section_80c.ulip_premium.amount),
            "sukanya_samriddhi": float(section_80c.sukanya_samriddhi.amount),
            "stamp_duty_property": float(section_80c.stamp_duty_property.amount),
            "senior_citizen_savings": float(section_80c.senior_citizen_savings.amount),
            "other_80c_investments": float(section_80c.other_80c_investments.amount),
            "total_invested": total_invested,
            "limit": limit,
            "remaining_limit": remaining_limit
        }
    
    def _serialize_section_80d(self, section_80d) -> Dict[str, Any]:
        """Serialize Section 80D details."""
        if not section_80d:
            return {
                "self_family_premium": 0.0, "parent_premium": 0.0, "preventive_health_checkup": 0.0,
                "parent_age": 55, "self_family_limit": 25000, "parent_limit": 25000, "preventive_limit": 5000
            }
        
        return {
            "self_family_premium": float(section_80d.self_family_premium.amount),
            "parent_premium": float(section_80d.parent_premium.amount),
            "preventive_health_checkup": float(section_80d.preventive_health_checkup.amount),
            "parent_age": section_80d.parent_age,
            "self_family_limit": float(section_80d.calculate_self_family_limit(30).amount),
            "parent_limit": float(section_80d.calculate_parent_limit().amount),
            "preventive_limit": 5000.0
        }
    
    def _serialize_section_80dd(self, section_80dd) -> Dict[str, Any]:
        """Serialize Section 80DD details."""
        if not section_80dd:
            return {"relation": None, "disability_percentage": None, "eligible_deduction": 0.0}
        
        from app.domain.value_objects.tax_regime import TaxRegime, TaxRegimeType
        regime = TaxRegime(TaxRegimeType.OLD)
        
        return {
            "relation": section_80dd.relation.value if section_80dd.relation else None,
            "disability_percentage": section_80dd.disability_percentage.value if section_80dd.disability_percentage else None,
            "eligible_deduction": float(section_80dd.calculate_eligible_deduction(regime).amount)
        }
    
    def _serialize_section_80ddb(self, section_80ddb) -> Dict[str, Any]:
        """Serialize Section 80DDB details."""
        if not section_80ddb:
            return {"dependent_age": 0, "medical_expenses": 0.0, "relation": None, "eligible_deduction": 0.0}
        
        from app.domain.value_objects.tax_regime import TaxRegime, TaxRegimeType
        regime = TaxRegime(TaxRegimeType.OLD)
        
        return {
            "dependent_age": section_80ddb.dependent_age,
            "medical_expenses": float(section_80ddb.medical_expenses.amount),
            "relation": section_80ddb.relation.value if section_80ddb.relation else None,
            "eligible_deduction": float(section_80ddb.calculate_eligible_deduction(regime).amount)
        }
    
    def _serialize_section_80eeb(self, section_80eeb) -> Dict[str, Any]:
        """Serialize Section 80EEB details."""
        if not section_80eeb:
            return {"ev_loan_interest": 0.0, "ev_purchase_date": None, "eligible_deduction": 0.0}
        
        from app.domain.value_objects.tax_regime import TaxRegime, TaxRegimeType
        regime = TaxRegime(TaxRegimeType.OLD)
        
        return {
            "ev_loan_interest": float(section_80eeb.ev_loan_interest.amount),
            "ev_purchase_date": section_80eeb.ev_purchase_date.isoformat() if section_80eeb.ev_purchase_date else None,
            "eligible_deduction": float(section_80eeb.calculate_eligible_deduction(regime).amount)
        }
    
    def _serialize_section_80g(self, section_80g) -> Dict[str, Any]:
        """Serialize Section 80G details."""
        if not section_80g:
            return {
                "pm_relief_fund": 0.0, "national_defence_fund": 0.0, "national_foundation_communal_harmony": 0.0,
                "zila_saksharta_samiti": 0.0, "national_illness_assistance_fund": 0.0, "national_blood_transfusion_council": 0.0,
                "national_trust_autism_fund": 0.0, "national_sports_fund": 0.0, "national_cultural_fund": 0.0,
                "technology_development_fund": 0.0, "national_children_fund": 0.0, "cm_relief_fund": 0.0,
                "army_naval_air_force_funds": 0.0, "swachh_bharat_kosh": 0.0, "clean_ganga_fund": 0.0,
                "drug_abuse_control_fund": 0.0, "other_100_percent_wo_limit": 0.0, "jn_memorial_fund": 0.0,
                "pm_drought_relief": 0.0, "indira_gandhi_memorial_trust": 0.0, "rajiv_gandhi_foundation": 0.0,
                "other_50_percent_wo_limit": 0.0, "family_planning_donation": 0.0, "indian_olympic_association": 0.0,
                "other_100_percent_w_limit": 0.0, "govt_charitable_donations": 0.0, "housing_authorities_donations": 0.0,
                "religious_renovation_donations": 0.0, "other_charitable_donations": 0.0, "other_50_percent_w_limit": 0.0,
                "total_donations": 0.0
            }
        
        return {
            "pm_relief_fund": float(section_80g.pm_relief_fund.amount),
            "national_defence_fund": float(section_80g.national_defence_fund.amount),
            "national_foundation_communal_harmony": float(section_80g.national_foundation_communal_harmony.amount),
            "zila_saksharta_samiti": float(section_80g.zila_saksharta_samiti.amount),
            "national_illness_assistance_fund": float(section_80g.national_illness_assistance_fund.amount),
            "national_blood_transfusion_council": float(section_80g.national_blood_transfusion_council.amount),
            "national_trust_autism_fund": float(section_80g.national_trust_autism_fund.amount),
            "national_sports_fund": float(section_80g.national_sports_fund.amount),
            "national_cultural_fund": float(section_80g.national_cultural_fund.amount),
            "technology_development_fund": float(section_80g.technology_development_fund.amount),
            "national_children_fund": float(section_80g.national_children_fund.amount),
            "cm_relief_fund": float(section_80g.cm_relief_fund.amount),
            "army_naval_air_force_funds": float(section_80g.army_naval_air_force_funds.amount),
            "swachh_bharat_kosh": float(section_80g.swachh_bharat_kosh.amount),
            "clean_ganga_fund": float(section_80g.clean_ganga_fund.amount),
            "drug_abuse_control_fund": float(section_80g.drug_abuse_control_fund.amount),
            "other_100_percent_wo_limit": float(section_80g.other_100_percent_wo_limit.amount),
            "jn_memorial_fund": float(section_80g.jn_memorial_fund.amount),
            "pm_drought_relief": float(section_80g.pm_drought_relief.amount),
            "indira_gandhi_memorial_trust": float(section_80g.indira_gandhi_memorial_trust.amount),
            "rajiv_gandhi_foundation": float(section_80g.rajiv_gandhi_foundation.amount),
            "other_50_percent_wo_limit": float(section_80g.other_50_percent_wo_limit.amount),
            "family_planning_donation": float(section_80g.family_planning_donation.amount),
            "indian_olympic_association": float(section_80g.indian_olympic_association.amount),
            "other_100_percent_w_limit": float(section_80g.other_100_percent_w_limit.amount),
            "govt_charitable_donations": float(section_80g.govt_charitable_donations.amount),
            "housing_authorities_donations": float(section_80g.housing_authorities_donations.amount),
            "religious_renovation_donations": float(section_80g.religious_renovation_donations.amount),
            "other_charitable_donations": float(section_80g.other_charitable_donations.amount),
            "other_50_percent_w_limit": float(section_80g.other_50_percent_w_limit.amount),
            "total_donations": float(section_80g.calculate_total_donations().amount)
        }
    
    def _serialize_section_80tta_ttb(self, section_80tta_ttb) -> Dict[str, Any]:
        """Serialize Section 80TTA/TTB details."""
        if not section_80tta_ttb:
            return {
                "savings_interest": 0.0, "fd_interest": 0.0, "rd_interest": 0.0, "post_office_interest": 0.0,
                "age": 25, "applicable_section": "80TTA", "exemption_limit": 10000, "eligible_exemption": 0.0
            }
        
        from app.domain.value_objects.tax_regime import TaxRegime, TaxRegimeType
        regime = TaxRegime(TaxRegimeType.OLD)
        
        breakdown = section_80tta_ttb.get_exemption_breakdown(regime)
        
        return {
            "savings_interest": float(section_80tta_ttb.savings_interest.amount),
            "fd_interest": float(section_80tta_ttb.fd_interest.amount),
            "rd_interest": float(section_80tta_ttb.rd_interest.amount),
            "post_office_interest": float(section_80tta_ttb.post_office_interest.amount),
            "age": section_80tta_ttb.age,
            "applicable_section": breakdown.get("applicable_section", "80TTA"),
            "exemption_limit": breakdown.get("exemption_limit", 10000),
            "eligible_exemption": breakdown.get("eligible_exemption", 0.0)
        }
    
    def _calculate_80u_deduction(self, section_80u) -> float:
        """Calculate Section 80U deduction."""
        if not section_80u:
            return 0.0
        
        from app.domain.value_objects.tax_regime import TaxRegime, TaxRegimeType
        regime = TaxRegime(TaxRegimeType.OLD)
        
        return float(section_80u.calculate_eligible_deduction(regime).amount)
    
    def _calculate_deductions_summary(self, deductions: TaxDeductions) -> Dict[str, Any]:
        """Calculate summary totals for deductions."""
        from app.domain.value_objects.tax_regime import TaxRegime, TaxRegimeType
        
        # Use old regime for calculation (deductions are primarily for old regime)
        regime = TaxRegime(TaxRegimeType.OLD)
        
        # Use default values for age and gross_income since we don't have them in this context
        total_deductions = float(deductions.calculate_total_deductions(regime, 30, Money.zero()).amount)
        total_interest_exemptions = float(deductions.calculate_interest_exemptions(regime).amount)
        combined_80c_80ccc_80ccd1 = float(deductions.calculate_combined_80c_80ccc_80ccd1_deduction(regime).amount)
        
        return {
            "total_deductions": total_deductions,
            "total_interest_exemptions": total_interest_exemptions,
            "combined_80c_80ccc_80ccd1": combined_80c_80ccc_80ccd1,
            "regime_applicable": "old"
        }
    
    def _serialize_house_property_income(self, house_property_income: HousePropertyIncome) -> Dict[str, Any]:
        """Serialize house property income to dict."""
        return {
            "property_type": house_property_income.property_type.value,
            "address": house_property_income.address,
            "annual_rent_received": float(house_property_income.annual_rent_received.amount),
            "municipal_taxes_paid": float(house_property_income.municipal_taxes_paid.amount),
            "home_loan_interest": float(house_property_income.home_loan_interest.amount),
            "pre_construction_interest": float(house_property_income.pre_construction_interest.amount)
        }
    
    def _serialize_capital_gains_income(self, capital_gains: CapitalGainsIncome) -> Dict[str, Any]:
        """Serialize capital gains income to dict."""
        return {
            "stcg_111a_equity_stt": float(capital_gains.stcg_111a_equity_stt.amount),
            "stcg_other_assets": float(capital_gains.stcg_other_assets.amount),
            "stcg_debt_mf": float(capital_gains.stcg_debt_mf.amount),
            "ltcg_112a_equity_stt": float(capital_gains.ltcg_112a_equity_stt.amount),
            "ltcg_other_assets": float(capital_gains.ltcg_other_assets.amount),
            "ltcg_debt_mf": float(capital_gains.ltcg_debt_mf.amount)
        }
    
    def _serialize_retirement_benefits(self, retirement_benefits: RetirementBenefits) -> Dict[str, Any]:
        """Serialize retirement benefits to dict."""
        if not retirement_benefits:
            return {}
        
        result = {}
        
        # Serialize leave encashment
        if retirement_benefits.leave_encashment:
            le = retirement_benefits.leave_encashment
            result["leave_encashment"] = {
                "leave_encashment_amount": float(le.leave_encashment_amount.amount),
                "average_monthly_salary": float(le.average_monthly_salary.amount),
                "leave_days_encashed": le.leave_days_encashed,
                "is_deceased": le.is_deceased,
                "during_employment": le.during_employment
            }
        
        # Serialize gratuity
        if retirement_benefits.gratuity:
            gr = retirement_benefits.gratuity
            result["gratuity"] = {
                "gratuity_amount": float(gr.gratuity_amount.amount),
                "monthly_salary": float(gr.monthly_salary.amount),
                "service_years": float(gr.service_years)
            }
        
        # Serialize VRS
        if retirement_benefits.vrs:
            vrs = retirement_benefits.vrs
            result["vrs"] = {
                "vrs_amount": float(vrs.vrs_amount.amount),
                "monthly_salary": float(vrs.monthly_salary.amount),
                "service_years": float(vrs.service_years)
            }
        
        # Serialize pension
        if retirement_benefits.pension:
            pen = retirement_benefits.pension
            result["pension"] = {
                "regular_pension": float(pen.regular_pension.amount),
                "commuted_pension": float(pen.commuted_pension.amount),
                "total_pension": float(pen.total_pension.amount),
                "gratuity_received": pen.gratuity_received
            }
        
        # Serialize retrenchment compensation
        if retirement_benefits.retrenchment_compensation:
            rc = retirement_benefits.retrenchment_compensation
            result["retrenchment_compensation"] = {
                "retrenchment_amount": float(rc.retrenchment_amount.amount),
                "monthly_salary": float(rc.monthly_salary.amount),
                "service_years": float(rc.service_years)
            }
        
        return result
    
    def _serialize_other_income(self, other_income: OtherIncome) -> Dict[str, Any]:
        """Serialize other income to dict."""
        if not other_income:
            return {}
        
        result = {
            "dividend_income": float(other_income.dividend_income.amount),
            "gifts_received": float(other_income.gifts_received.amount),
            "business_professional_income": float(other_income.business_professional_income.amount),
            "other_miscellaneous_income": float(other_income.other_miscellaneous_income.amount)
        }
        
        # Serialize interest income
        if other_income.interest_income:
            interest = other_income.interest_income
            result["interest_income"] = {
                "savings_interest": float(interest.savings_account_interest.amount),
                "fd_interest": float(interest.fixed_deposit_interest.amount),
                "rd_interest": float(interest.recurring_deposit_interest.amount),
                "post_office_interest": float(interest.post_office_interest.amount)
            }
        
        # Serialize house property income
        if other_income.house_property_income:
            result["house_property_income"] = self._serialize_house_property_income(other_income.house_property_income)
        
        # Serialize capital gains income
        if other_income.capital_gains_income:
            result["capital_gains_income"] = self._serialize_capital_gains_income(other_income.capital_gains_income)
        
        return result

    async def compute_monthly_tax(self, employee_id: str, organization_id: str) -> Dict[str, Any]:
        """
        Compute monthly tax for an employee based on their salary package record.
        
        Args:
            employee_id: Employee ID as string
            organization_id: Organization ID for database segregation
            
        Returns:
            Dict[str, Any]: Monthly tax computation result with details
            
        Raises:
            ValueError: If employee or salary data not found
            RuntimeError: If computation fails
        """
        logger.debug(f"UnifiedTaxationController.compute_monthly_tax: Starting for employee {employee_id}, organization {organization_id}")
        
        try:
            
            # Check if enhanced_tax_service is available
            if not self.tax_calculation_service:
                logger.error("UnifiedTaxationController.compute_monthly_tax: Enhanced tax service not configured")
                raise RuntimeError("Enhanced tax service not configured")
            
            # Use the enhanced tax service to compute monthly tax with details
            result = await self.tax_calculation_service.compute_monthly_tax_with_details(
                employee_id, organization_id
            )
            
            logger.debug(f"UnifiedTaxationController.compute_monthly_tax: Successfully received result from enhanced tax service")
            logger.debug(f"UnifiedTaxationController.compute_monthly_tax: Result keys: {list(result.keys())}")
            logger.debug(f"UnifiedTaxationController.compute_monthly_tax: Monthly tax liability: {result.get('monthly_tax_liability', 'Not found')}")
            
            return result
            
        except ValueError as e:
            logger.error(f"UnifiedTaxationController.compute_monthly_tax: Validation error for employee {employee_id}: {str(e)}")
            # Re-raise validation errors
            raise e
        except Exception as e:
            logger.error(f"UnifiedTaxationController.compute_monthly_tax: Unexpected error for employee {employee_id}: {str(e)}", exc_info=True)
            # Wrap other errors
            raise RuntimeError(f"Failed to compute monthly tax for employee {employee_id}: {str(e)}")
    
    async def compute_monthly_salary(
        self, 
        request: MonthlySalaryComputeRequestDTO,
        organization_id: str
    ) -> MonthlySalaryResponseDTO:
        """
        Compute monthly salary for an employee.
        
        This method:
        1. Uses the monthly salary computation use case
        2. Creates a MonthlySalary entity with current month's data
        3. Computes monthly salary components including deductions
        4. Calculates tax liability for the month
        5. Returns the computed monthly salary details
        
        Args:
            request: Monthly salary computation request
            organization_id: Organization ID for multi-tenancy
            
        Returns:
            MonthlySalaryResponseDTO: Computed monthly salary details
            
        Raises:
            ValueError: If employee or salary package not found
            RuntimeError: If computation fails
        """
        
        logger.debug(f"Computing monthly salary for employee {request.employee_id}")
        logger.debug(f"Month: {request.month}, Year: {request.year}, Tax Year: {request.tax_year}")
        
        try:
            # Use the monthly salary computation use case
            result = await self.compute_monthly_salary_use_case.execute(
                request, organization_id
            )
            
            logger.info(f"Successfully computed monthly salary for employee {request.employee_id}")
            return result
            
        except Exception as e:
            logger.error(f"Failed to compute monthly salary for employee {request.employee_id}: {str(e)}")
            raise
    
    async def export_salary_package_to_excel(
        self, 
        employee_id: str, 
        tax_year: Optional[str], 
        hostname: str
    ) -> bytes:
        """
        Export comprehensive SalaryPackageRecord data to Excel format.
        
        Args:
            employee_id: Employee ID to export data for
            tax_year: Tax year (optional, defaults to current)
            hostname: Organization hostname
            
        Returns:
            bytes: Excel file content as bytes
        """
        try:
            logger.info(f"Exporting salary package to Excel for employee: {employee_id}, tax_year: {tax_year}")
            
            # Get or create the salary package record
            salary_package_record = await self._get_or_create_salary_package_record(
                employee_id, tax_year, hostname
            )
            
            # Create Excel workbook
            wb = openpyxl.Workbook()
            
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            subheader_font = Font(bold=True, color="000000")
            subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin")
            )
            
            # 1. Summary Sheet
            summary_ws = wb.create_sheet("Summary")
            self._create_summary_sheet(summary_ws, salary_package_record, header_font, header_fill, border)
            
            # 2. Salary History Sheet
            salary_history_ws = wb.create_sheet("Salary History")
            self._create_salary_history_sheet(salary_history_ws, salary_package_record, header_font, header_fill, subheader_font, subheader_fill, border)
            
            # 3. Annual Salary Components Sheet
            annual_salary_ws = wb.create_sheet("Annual Salary")
            self._create_annual_salary_sheet(annual_salary_ws, salary_package_record, header_font, header_fill, border)
            
            # 4. Specific Allowances Sheet
            allowances_ws = wb.create_sheet("Specific Allowances")
            self._create_specific_allowances_sheet(allowances_ws, salary_package_record, header_font, header_fill, border)
            
            # 5. Perquisites Sheet (if available)
            if salary_package_record.perquisites:
                perquisites_ws = wb.create_sheet("Perquisites")
                self._create_perquisites_sheet(perquisites_ws, salary_package_record, header_font, header_fill, border)
            
            # 6. Other Income Sheet (if available)
            if salary_package_record.other_income:
                other_income_ws = wb.create_sheet("Other Income")
                self._create_other_income_sheet(other_income_ws, salary_package_record, header_font, header_fill, border)
            
            # 7. Deductions Sheet
            deductions_ws = wb.create_sheet("Deductions")
            self._create_deductions_sheet(deductions_ws, salary_package_record, header_font, header_fill, subheader_font, subheader_fill, border)
            
            # 8. Tax Calculation Sheet (if calculated)
            if salary_package_record.calculation_result:
                tax_calc_ws = wb.create_sheet("Tax Calculation")
                self._create_tax_calculation_sheet(tax_calc_ws, salary_package_record, header_font, header_fill, border)
            
            # 9. Raw Data Sheet (for debugging/reference)
            raw_data_ws = wb.create_sheet("Raw Data")
            self._create_raw_data_sheet(raw_data_ws, salary_package_record, header_font, header_fill, border)
            
            # Save workbook to bytes
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            logger.info(f"Successfully exported salary package to Excel for employee: {employee_id}")
            return excel_buffer.getvalue()
            
        except Exception as e:
            logger.error(f"Error exporting salary package to Excel: {e}")
            raise RuntimeError(f"Failed to export salary package to Excel: {str(e)}")
    
    def _create_summary_sheet(self, ws, salary_package_record, header_font, header_fill, border):
        """Create summary sheet with key information."""
        ws.title = "Summary"
        
        # Headers
        headers = ["Field", "Value"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
        
        # Data rows
        summary_data = [
            ("Employee ID", str(salary_package_record.employee_id)),
            ("Tax Year", str(salary_package_record.tax_year)),
            ("Age", salary_package_record.age),
            ("Tax Regime", salary_package_record.regime.regime_type.value),
            ("Organization ID", salary_package_record.organization_id or "N/A"),
            ("Is Government Employee", "Yes" if salary_package_record.is_government_employee else "No"),
            ("Is Final", "Yes" if salary_package_record.is_final else "No"),
            ("Salary Revisions Count", len(salary_package_record.salary_incomes)),
            ("Has Perquisites", "Yes" if salary_package_record.perquisites else "No"),
            ("Has Other Income", "Yes" if salary_package_record.other_income else "No"),
            ("Has Retirement Benefits", "Yes" if salary_package_record.retirement_benefits else "No"),
            ("Created At", salary_package_record.created_at.strftime("%Y-%m-%d %H:%M:%S")),
            ("Updated At", salary_package_record.updated_at.strftime("%Y-%m-%d %H:%M:%S")),
            ("Last Calculated", salary_package_record.last_calculated_at.strftime("%Y-%m-%d %H:%M:%S") if salary_package_record.last_calculated_at else "Never"),
        ]
        
        if salary_package_record.calculation_result:
            summary_data.extend([
                ("", ""),  # Empty row
                ("TAX CALCULATION RESULTS", ""),
                ("Gross Income", f"₹{salary_package_record.calculation_result.gross_income.to_float():,.2f}"),
                ("Total Exemptions", f"₹{salary_package_record.calculation_result.total_exemptions.to_float():,.2f}"),
                ("Total Deductions", f"₹{salary_package_record.calculation_result.total_deductions.to_float():,.2f}"),
                ("Taxable Income", f"₹{salary_package_record.calculation_result.taxable_income.to_float():,.2f}"),
                ("Tax Liability", f"₹{salary_package_record.calculation_result.tax_liability.to_float():,.2f}"),
                ("Monthly Tax", f"₹{salary_package_record.calculation_result.tax_liability.divide(12).to_float():,.2f}"),
            ])
        
        for row, (field, value) in enumerate(summary_data, 2):
            ws.cell(row=row, column=1, value=field).border = border
            ws.cell(row=row, column=2, value=value).border = border
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_salary_history_sheet(self, ws, salary_package_record, header_font, header_fill, subheader_font, subheader_fill, border):
        """Create salary history sheet showing all salary revisions."""
        ws.title = "Salary History"
        
        # Headers
        headers = [
            "Revision #", "Effective From", "Effective Till", "Months Applicable",
            "Basic Salary", "DA", "HRA", "Special Allowance", "Bonus", "Commission",
            "Total Specific Allowances", "Monthly Gross", "Period Total"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
        
        # Get salary breakdown
        breakdown = salary_package_record.get_annual_salary_breakdown()
        
        # Add salary periods
        row = 2
        for period in breakdown.get("salary_periods", []):
            components = period.get("salary_components", {})
            specific_breakdown = components.get("specific_allowances_breakdown", {})
            total_specific = sum(specific_breakdown.values())
            
            data = [
                period.get("period_index"),
                period.get("effective_from", "N/A"),
                period.get("effective_till", "N/A"),
                period.get("months_applicable", 0),
                components.get("basic_salary", 0),
                components.get("dearness_allowance", 0),
                components.get("hra_provided", 0),
                components.get("special_allowance", 0),
                components.get("bonus", 0),
                components.get("commission", 0),
                total_specific,
                period.get("monthly_gross_salary", 0),
                period.get("total_for_period", 0)
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                # Format currency columns
                if col >= 5:  # Currency columns
                    if isinstance(value, (int, float)) and value != 0:
                        cell.value = f"₹{value:,.2f}"
            row += 1
        
        # Add projections if any
        if breakdown.get("projections"):
            # Add separator row
            ws.cell(row=row, column=1, value="PROJECTIONS").font = subheader_font
            ws.cell(row=row, column=1).fill = subheader_fill
            row += 1
            
            for projection in breakdown["projections"]:
                data = [
                    "Projection",
                    projection.get("projection_start", "N/A"),
                    projection.get("projection_end", "N/A"),
                    projection.get("uncovered_months", 0),
                    "", "", "", "", "", "",  # Empty cells for individual components
                    "",  # Empty for total specific allowances
                    projection.get("monthly_gross_salary", 0),
                    projection.get("projected_amount", 0)
                ]
                
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = border
                    # Format currency columns
                    if col in [12, 13] and isinstance(value, (int, float)) and value != 0:
                        cell.value = f"₹{value:,.2f}"
                row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_annual_salary_sheet(self, ws, salary_package_record, header_font, header_fill, border):
        """Create annual salary components sheet."""
        ws.title = "Annual Salary"
        
        annual_salary = salary_package_record.get_annual_salary_income()
        
        # Headers
        headers = ["Component", "Annual Amount", "Monthly Equivalent"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
        
        # Core salary components
        core_components = [
            ("Basic Salary", annual_salary.basic_salary.to_float()),
            ("Dearness Allowance", annual_salary.dearness_allowance.to_float()),
            ("HRA Provided", annual_salary.hra_provided.to_float()),
            ("Special Allowance", annual_salary.special_allowance.to_float()),
            ("Bonus", annual_salary.bonus.to_float()),
            ("Commission", annual_salary.commission.to_float()),
        ]
        
        row = 2
        for component, annual_amount in core_components:
            monthly_equivalent = annual_amount / 12 if annual_amount > 0 else 0
            
            ws.cell(row=row, column=1, value=component).border = border
            ws.cell(row=row, column=2, value=f"₹{annual_amount:,.2f}").border = border
            ws.cell(row=row, column=3, value=f"₹{monthly_equivalent:,.2f}").border = border
            row += 1
        
        # Add total row
        total_annual = annual_salary.calculate_gross_salary().to_float()
        total_monthly = total_annual / 12
        
        ws.cell(row=row, column=1, value="TOTAL GROSS SALARY").border = border
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2, value=f"₹{total_annual:,.2f}").border = border
        ws.cell(row=row, column=2).font = Font(bold=True)
        ws.cell(row=row, column=3, value=f"₹{total_monthly:,.2f}").border = border
        ws.cell(row=row, column=3).font = Font(bold=True)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_specific_allowances_sheet(self, ws, salary_package_record, header_font, header_fill, border):
        """Create specific allowances breakdown sheet."""
        ws.title = "Specific Allowances"
        
        annual_salary = salary_package_record.get_annual_salary_income()
        
        if not annual_salary.specific_allowances:
            ws.cell(row=1, column=1, value="No specific allowances data available")
            return
        
        # Headers
        headers = ["Allowance Type", "Annual Amount", "Monthly Equivalent", "Exemption Limit", "Taxable Amount"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
        
        specific_allowances = annual_salary.specific_allowances
        
        # Define allowance fields with their display names
        allowance_fields = [
            ("hills_allowance", "Hills Allowance", "hills_exemption_limit"),
            ("border_allowance", "Border Allowance", "border_exemption_limit"),
            ("transport_employee_allowance", "Transport Allowance", None),
            ("children_education_allowance", "Children Education Allowance", None),
            ("hostel_allowance", "Hostel Allowance", None),
            ("disabled_transport_allowance", "Disabled Transport Allowance", None),
            ("underground_mines_allowance", "Underground Mines Allowance", None),
            ("government_entertainment_allowance", "Government Entertainment Allowance", None),
            ("city_compensatory_allowance", "City Compensatory Allowance", None),
            ("rural_allowance", "Rural Allowance", None),
            ("proctorship_allowance", "Proctorship Allowance", None),
            ("wardenship_allowance", "Wardenship Allowance", None),
            ("project_allowance", "Project Allowance", None),
            ("deputation_allowance", "Deputation Allowance", None),
            ("overtime_allowance", "Overtime Allowance", None),
            ("any_other_allowance", "Any Other Allowance", "any_other_allowance_exemption"),
        ]
        
        row = 2
        total_allowances = 0
        total_exemptions = 0
        
        for field_name, display_name, exemption_field in allowance_fields:
            if hasattr(specific_allowances, field_name):
                annual_amount = getattr(specific_allowances, field_name).to_float()
                
                if annual_amount > 0:  # Only show allowances with values
                    monthly_equivalent = annual_amount / 12
                    
                    # Get exemption limit if available
                    exemption_limit = 0
                    if exemption_field and hasattr(specific_allowances, exemption_field):
                        exemption_limit = getattr(specific_allowances, exemption_field).to_float()
                    
                    # Calculate taxable amount (simplified - actual calculation is more complex)
                    taxable_amount = max(0, annual_amount - exemption_limit)
                    
                    ws.cell(row=row, column=1, value=display_name).border = border
                    ws.cell(row=row, column=2, value=f"₹{annual_amount:,.2f}").border = border
                    ws.cell(row=row, column=3, value=f"₹{monthly_equivalent:,.2f}").border = border
                    ws.cell(row=row, column=4, value=f"₹{exemption_limit:,.2f}" if exemption_limit > 0 else "N/A").border = border
                    ws.cell(row=row, column=5, value=f"₹{taxable_amount:,.2f}").border = border
                    
                    total_allowances += annual_amount
                    total_exemptions += exemption_limit
                    row += 1
        
        # Add total row
        if total_allowances > 0:
            total_taxable = total_allowances - total_exemptions
            ws.cell(row=row, column=1, value="TOTAL").border = border
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=2, value=f"₹{total_allowances:,.2f}").border = border
            ws.cell(row=row, column=2).font = Font(bold=True)
            ws.cell(row=row, column=3, value=f"₹{total_allowances/12:,.2f}").border = border
            ws.cell(row=row, column=3).font = Font(bold=True)
            ws.cell(row=row, column=4, value=f"₹{total_exemptions:,.2f}").border = border
            ws.cell(row=row, column=4).font = Font(bold=True)
            ws.cell(row=row, column=5, value=f"₹{total_taxable:,.2f}").border = border
            ws.cell(row=row, column=5).font = Font(bold=True)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 25)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_perquisites_sheet(self, ws, salary_package_record, header_font, header_fill, border):
        """Create perquisites sheet if available."""
        ws.title = "Perquisites"
        
        if not salary_package_record.perquisites:
            ws.cell(row=1, column=1, value="No perquisites data available")
            return
        
        # Headers
        headers = ["Perquisite Type", "Value", "Taxable Amount"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
        
        # Add perquisites data (this would need to be expanded based on actual perquisites structure)
        row = 2
        ws.cell(row=row, column=1, value="Accommodation").border = border
        ws.cell(row=row, column=2, value="Data available in perquisites object").border = border
        ws.cell(row=row, column=3, value="Calculated based on regime").border = border
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_other_income_sheet(self, ws, salary_package_record, header_font, header_fill, border):
        """Create other income sheet if available."""
        ws.title = "Other Income"
        
        if not salary_package_record.other_income:
            ws.cell(row=1, column=1, value="No other income data available")
            return
        
        # Headers
        headers = ["Income Type", "Amount", "Tax Treatment"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
        
        # Add other income data
        other_income = salary_package_record.other_income
        row = 2
        
        # Dividend income
        dividend_amount = other_income.dividend_income.to_float() if other_income.dividend_income else 0
        if dividend_amount > 0:
            ws.cell(row=row, column=1, value="Dividend Income").border = border
            ws.cell(row=row, column=2, value=f"₹{dividend_amount:,.2f}").border = border
            ws.cell(row=row, column=3, value="Added to total income").border = border
            row += 1
        
        # Business/Professional income
        business_amount = other_income.business_professional_income.to_float() if other_income.business_professional_income else 0
        if business_amount > 0:
            ws.cell(row=row, column=1, value="Business/Professional Income").border = border
            ws.cell(row=row, column=2, value=f"₹{business_amount:,.2f}").border = border
            ws.cell(row=row, column=3, value="Added to total income").border = border
            row += 1
        
        # House property income
        if other_income.house_property_income:
            hpi_amount = other_income.house_property_income.calculate_net_income_from_house_property(salary_package_record.regime).to_float()
            ws.cell(row=row, column=1, value="House Property Income").border = border
            ws.cell(row=row, column=2, value=f"₹{hpi_amount:,.2f}").border = border
            ws.cell(row=row, column=3, value="Net income after deductions").border = border
            row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_deductions_sheet(self, ws, salary_package_record, header_font, header_fill, subheader_font, subheader_fill, border):
        """Create deductions sheet."""
        ws.title = "Deductions"
        
        # Headers
        headers = ["Deduction Section", "Item", "Amount"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
        
        deductions = salary_package_record.deductions
        row = 2
        
        # Section 80C
        ws.cell(row=row, column=1, value="Section 80C").font = subheader_font
        ws.cell(row=row, column=1).fill = subheader_fill
        ws.cell(row=row, column=1).border = border
        row += 1
        
        section_80c_fields = [
            ("epf", "EPF"),
            ("ppf", "PPF"),
            ("life_insurance_premium", "Life Insurance Premium"),
            ("elss", "ELSS"),
            ("nsc", "NSC"),
            ("home_loan_principal", "Home Loan Principal"),
            ("tuition_fees", "Tuition Fees"),
            ("tax_saver_fd", "Tax Saver FD"),
            ("sukanya_samriddhi", "Sukanya Samriddhi"),
            ("ulip", "ULIP"),
            ("other_80c", "Other 80C Investments"),
        ]
        
        for field_name, display_name in section_80c_fields:
            if hasattr(deductions.section_80c, field_name):
                amount = getattr(deductions.section_80c, field_name).to_float()
                if amount > 0:
                    ws.cell(row=row, column=2, value=display_name).border = border
                    ws.cell(row=row, column=3, value=f"₹{amount:,.2f}").border = border
                    row += 1
        
        # Section 80D
        row += 1  # Empty row
        ws.cell(row=row, column=1, value="Section 80D").font = subheader_font
        ws.cell(row=row, column=1).fill = subheader_fill
        ws.cell(row=row, column=1).border = border
        row += 1
        
        section_80d_fields = [
            ("health_insurance_self", "Health Insurance - Self"),
            ("health_insurance_parents", "Health Insurance - Parents"),
            ("preventive_health_checkup", "Preventive Health Checkup"),
        ]
        
        for field_name, display_name in section_80d_fields:
            if hasattr(deductions.section_80d, field_name):
                amount = getattr(deductions.section_80d, field_name).to_float()
                if amount > 0:
                    ws.cell(row=row, column=2, value=display_name).border = border
                    ws.cell(row=row, column=3, value=f"₹{amount:,.2f}").border = border
                    row += 1
        
        # Other Deductions
        row += 1  # Empty row
        ws.cell(row=row, column=1, value="Other Deductions").font = subheader_font
        ws.cell(row=row, column=1).fill = subheader_fill
        ws.cell(row=row, column=1).border = border
        row += 1
        
        other_deduction_fields = [
            ("section_80e", "Section 80E - Education Loan"),
            ("section_80g", "Section 80G - Donations"),
            ("section_24b", "Section 24B - Home Loan Interest"),
        ]
        
        for field_name, display_name in other_deduction_fields:
            if hasattr(deductions, field_name):
                amount = getattr(deductions, field_name).to_float()
                if amount > 0:
                    ws.cell(row=row, column=2, value=display_name).border = border
                    ws.cell(row=row, column=3, value=f"₹{amount:,.2f}").border = border
                    row += 1
        
        # Total deductions
        row += 1  # Empty row
        # Calculate gross income for deduction calculations
        gross_income = salary_package_record.calculate_gross_income()
        total_deductions = deductions.calculate_total_deductions(salary_package_record.regime, salary_package_record.age, gross_income).to_float()
        ws.cell(row=row, column=2, value="TOTAL DEDUCTIONS").border = border
        ws.cell(row=row, column=2).font = Font(bold=True)
        ws.cell(row=row, column=3, value=f"₹{total_deductions:,.2f}").border = border
        ws.cell(row=row, column=3).font = Font(bold=True)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_tax_calculation_sheet(self, ws, salary_package_record, header_font, header_fill, border):
        """Create tax calculation sheet if calculation result is available."""
        ws.title = "Tax Calculation"
        
        if not salary_package_record.calculation_result:
            ws.cell(row=1, column=1, value="No tax calculation data available")
            return
        
        # Headers
        headers = ["Description", "Amount"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
        
        calc_result = salary_package_record.calculation_result
        
        # Tax calculation data
        calc_data = [
            ("Gross Income", calc_result.gross_income.to_float()),
            ("Total Exemptions", calc_result.total_exemptions.to_float()),
            ("Income After Exemptions", calc_result.gross_income.subtract(calc_result.total_exemptions).to_float()),
            ("Total Deductions", calc_result.total_deductions.to_float()),
            ("Taxable Income", calc_result.taxable_income.to_float()),
            ("Tax Liability", calc_result.tax_liability.to_float()),
            ("Monthly Tax", calc_result.tax_liability.divide(12).to_float()),
        ]
        
        # Add effective tax rate
        gross_income_float = calc_result.gross_income.to_float()
        tax_liability_float = calc_result.tax_liability.to_float()
        effective_rate = (tax_liability_float / gross_income_float * 100) if gross_income_float > 0 else 0
        calc_data.append(("Effective Tax Rate", f"{effective_rate:.2f}%"))
        
        row = 2
        for description, amount in calc_data:
            ws.cell(row=row, column=1, value=description).border = border
            if isinstance(amount, str):  # For percentage
                ws.cell(row=row, column=2, value=amount).border = border
            else:
                ws.cell(row=row, column=2, value=f"₹{amount:,.2f}").border = border
            row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_raw_data_sheet(self, ws, salary_package_record, header_font, header_fill, border):
        """Create raw data sheet for debugging/reference."""
        ws.title = "Raw Data"
        
        # Headers
        headers = ["Field Path", "Value", "Type"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
        
        # Get detailed breakdown
        detailed_breakdown = salary_package_record.get_detailed_breakdown()
        
        def flatten_dict(d, parent_key='', sep='.'):
            """Flatten nested dictionary for raw data view."""
            items = []
            for k, v in d.items():
                new_key = f"{parent_key}{sep}{k}" if parent_key else k
                if isinstance(v, dict):
                    items.extend(flatten_dict(v, new_key, sep=sep).items())
                elif isinstance(v, list):
                    for i, item in enumerate(v):
                        if isinstance(item, dict):
                            items.extend(flatten_dict(item, f"{new_key}[{i}]", sep=sep).items())
                        else:
                            items.append((f"{new_key}[{i}]", str(item)))
                else:
                    items.append((new_key, v))
            return dict(items)
        
        # Flatten the data
        flat_data = flatten_dict(detailed_breakdown)
        
        row = 2
        for field_path, value in flat_data.items():
            ws.cell(row=row, column=1, value=field_path).border = border
            ws.cell(row=row, column=2, value=str(value)).border = border
            ws.cell(row=row, column=3, value=type(value).__name__).border = border
            row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    async def export_salary_package_single_sheet(
        self, 
        employee_id: str, 
        tax_year: Optional[str], 
        hostname: str
    ) -> bytes:
        """
        Export comprehensive SalaryPackageRecord data to Excel format in a single sheet.
        
        Args:
            employee_id: Employee ID to export data for
            tax_year: Tax year (optional, defaults to current)
            hostname: Organization hostname
            
        Returns:
            bytes: Excel file content as bytes
        """
        try:
            logger.info(f"Exporting salary package to single Excel sheet for employee: {employee_id}, tax_year: {tax_year}")
            
            # Get or create the salary package record
            salary_package_record = await self._get_or_create_salary_package_record(
                employee_id, tax_year, hostname
            )
            
            # Create Excel workbook with single sheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Complete Salary Package"
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            section_font = Font(bold=True, color="000000", size=12)
            section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            subsection_font = Font(bold=True, color="000000", size=10)
            border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin")
            )
            
            current_row = 1
            
            # 1. Employee Summary Section
            current_row = self._add_employee_summary_section(ws, salary_package_record, current_row, section_font, section_fill, border)
            current_row += 2  # Add spacing
            
            # 2. Salary Components Section
            current_row = self._add_salary_components_section(ws, salary_package_record, current_row, section_font, section_fill, subsection_font, border)
            current_row += 2  # Add spacing
            
            # 3. Specific Allowances Section
            current_row = self._add_specific_allowances_section(ws, salary_package_record, current_row, section_font, section_fill, border)
            current_row += 2  # Add spacing
            
            # 4. Perquisites Section (if available)
            if salary_package_record.perquisites:
                current_row = self._add_perquisites_section(ws, salary_package_record, current_row, section_font, section_fill, border)
                current_row += 2  # Add spacing
            
            # 5. Other Income Section (if available)
            if salary_package_record.other_income:
                current_row = self._add_other_income_section(ws, salary_package_record, current_row, section_font, section_fill, border)
                current_row += 2  # Add spacing
            
            # 6. Deductions Section
            current_row = self._add_deductions_section(ws, salary_package_record, current_row, section_font, section_fill, subsection_font, border)
            current_row += 2  # Add spacing
            
            # 7. Tax Calculation Section (if calculated)
            if salary_package_record.calculation_result:
                current_row = self._add_tax_calculation_section(ws, salary_package_record, current_row, section_font, section_fill, border)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save workbook to bytes
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            logger.info(f"Successfully exported salary package to single Excel sheet for employee: {employee_id}")
            return excel_buffer.getvalue()
            
        except Exception as e:
            logger.error(f"Error exporting salary package to single Excel sheet: {e}")
            raise RuntimeError(f"Failed to export salary package to single Excel sheet: {str(e)}")
    
    def _add_employee_summary_section(self, ws, salary_package_record, start_row, section_font, section_fill, border):
        """Add employee summary section to single sheet."""
        # Section header
        ws.cell(row=start_row, column=1, value="EMPLOYEE SUMMARY").font = section_font
        ws.cell(row=start_row, column=1).fill = section_fill
        ws.merge_cells(f'A{start_row}:B{start_row}')
        
        current_row = start_row + 1
        
        # Employee data
        summary_data = [
            ("Employee ID", str(salary_package_record.employee_id)),
            ("Tax Year", str(salary_package_record.tax_year)),
            ("Age", salary_package_record.age),
            ("Tax Regime", salary_package_record.regime.regime_type.value),
            ("Is Government Employee", "Yes" if salary_package_record.is_government_employee else "No"),
            ("Salary Revisions", len(salary_package_record.salary_incomes)),
            ("Record Status", "Final" if salary_package_record.is_final else "Draft"),
        ]
        
        for field, value in summary_data:
            ws.cell(row=current_row, column=1, value=field).border = border
            ws.cell(row=current_row, column=2, value=value).border = border
            current_row += 1
        
        return current_row
    
    def _add_salary_components_section(self, ws, salary_package_record, start_row, section_font, section_fill, subsection_font, border):
        """Add salary components section to single sheet."""
        # Section header
        ws.cell(row=start_row, column=1, value="SALARY COMPONENTS").font = section_font
        ws.cell(row=start_row, column=1).fill = section_fill
        ws.merge_cells(f'A{start_row}:C{start_row}')
        
        current_row = start_row + 1
        
        # Headers
        headers = ["Component", "Annual Amount", "Monthly Amount"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = subsection_font
            cell.border = border
        current_row += 1
        
        # Get annual salary
        annual_salary = salary_package_record.get_annual_salary_income()
        
        # Core components
        components = [
            ("Basic Salary", annual_salary.basic_salary.to_float()),
            ("Dearness Allowance", annual_salary.dearness_allowance.to_float()),
            ("HRA Provided", annual_salary.hra_provided.to_float()),
            ("Special Allowance", annual_salary.special_allowance.to_float()),
            ("Bonus", annual_salary.bonus.to_float()),
            ("Commission", annual_salary.commission.to_float()),
        ]
        
        total_annual = 0
        for component, annual_amount in components:
            if annual_amount > 0:
                monthly_amount = annual_amount / 12
                total_annual += annual_amount
                
                ws.cell(row=current_row, column=1, value=component).border = border
                ws.cell(row=current_row, column=2, value=f"₹{annual_amount:,.2f}").border = border
                ws.cell(row=current_row, column=3, value=f"₹{monthly_amount:,.2f}").border = border
                current_row += 1
        
        # Total row
        ws.cell(row=current_row, column=1, value="TOTAL GROSS SALARY").border = border
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        ws.cell(row=current_row, column=2, value=f"₹{total_annual:,.2f}").border = border
        ws.cell(row=current_row, column=2).font = Font(bold=True)
        ws.cell(row=current_row, column=3, value=f"₹{total_annual/12:,.2f}").border = border
        ws.cell(row=current_row, column=3).font = Font(bold=True)
        current_row += 1
        
        return current_row
    
    def _add_specific_allowances_section(self, ws, salary_package_record, start_row, section_font, section_fill, border):
        """Add specific allowances section to single sheet."""
        annual_salary = salary_package_record.get_annual_salary_income()
        
        if not annual_salary.specific_allowances:
            return start_row
        
        # Section header
        ws.cell(row=start_row, column=1, value="SPECIFIC ALLOWANCES").font = section_font
        ws.cell(row=start_row, column=1).fill = section_fill
        ws.merge_cells(f'A{start_row}:C{start_row}')
        
        current_row = start_row + 1
        
        # Headers
        headers = ["Allowance Type", "Annual Amount", "Monthly Amount"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.border = border
        current_row += 1
        
        specific_allowances = annual_salary.specific_allowances
        
        # Define allowance fields with their display names
        allowance_fields = [
            ("hills_allowance", "Hills Allowance"),
            ("border_allowance", "Border Allowance"),
            ("transport_employee_allowance", "Transport Allowance"),
            ("children_education_allowance", "Children Education Allowance"),
            ("hostel_allowance", "Hostel Allowance"),
            ("city_compensatory_allowance", "City Compensatory Allowance"),
            ("rural_allowance", "Rural Allowance"),
            ("project_allowance", "Project Allowance"),
            ("overtime_allowance", "Overtime Allowance"),
            ("any_other_allowance", "Any Other Allowance"),
        ]
        
        total_specific = 0
        for field_name, display_name in allowance_fields:
            if hasattr(specific_allowances, field_name):
                annual_amount = getattr(specific_allowances, field_name).to_float()
                if annual_amount > 0:
                    monthly_amount = annual_amount / 12
                    total_specific += annual_amount
                    
                    ws.cell(row=current_row, column=1, value=display_name).border = border
                    ws.cell(row=current_row, column=2, value=f"₹{annual_amount:,.2f}").border = border
                    ws.cell(row=current_row, column=3, value=f"₹{monthly_amount:,.2f}").border = border
                    current_row += 1
        
        # Total row if any allowances exist
        if total_specific > 0:
            ws.cell(row=current_row, column=1, value="TOTAL SPECIFIC ALLOWANCES").border = border
            ws.cell(row=current_row, column=1).font = Font(bold=True)
            ws.cell(row=current_row, column=2, value=f"₹{total_specific:,.2f}").border = border
            ws.cell(row=current_row, column=2).font = Font(bold=True)
            ws.cell(row=current_row, column=3, value=f"₹{total_specific/12:,.2f}").border = border
            ws.cell(row=current_row, column=3).font = Font(bold=True)
            current_row += 1
        
        return current_row
    
    def _add_perquisites_section(self, ws, salary_package_record, start_row, section_font, section_fill, border):
        """Add perquisites section to single sheet."""
        # Section header
        ws.cell(row=start_row, column=1, value="PERQUISITES").font = section_font
        ws.cell(row=start_row, column=1).fill = section_fill
        ws.merge_cells(f'A{start_row}:B{start_row}')
        
        current_row = start_row + 1
        
        # For now, add placeholder data - can be expanded based on actual perquisites structure
        ws.cell(row=current_row, column=1, value="Perquisites Available").border = border
        ws.cell(row=current_row, column=2, value="Yes").border = border
        current_row += 1
        
        ws.cell(row=current_row, column=1, value="Details").border = border
        ws.cell(row=current_row, column=2, value="Available in detailed view").border = border
        current_row += 1
        
        return current_row
    
    def _add_other_income_section(self, ws, salary_package_record, start_row, section_font, section_fill, border):
        """Add other income section to single sheet."""
        # Section header
        ws.cell(row=start_row, column=1, value="OTHER INCOME").font = section_font
        ws.cell(row=start_row, column=1).fill = section_fill
        ws.merge_cells(f'A{start_row}:B{start_row}')
        
        current_row = start_row + 1
        
        other_income = salary_package_record.other_income
        
        # Dividend income
        dividend_amount = other_income.dividend_income.to_float() if other_income.dividend_income else 0
        if dividend_amount > 0:
            ws.cell(row=current_row, column=1, value="Dividend Income").border = border
            ws.cell(row=current_row, column=2, value=f"₹{dividend_amount:,.2f}").border = border
            current_row += 1
        
        # Business/Professional income
        business_amount = other_income.business_professional_income.to_float() if other_income.business_professional_income else 0
        if business_amount > 0:
            ws.cell(row=current_row, column=1, value="Business/Professional Income").border = border
            ws.cell(row=current_row, column=2, value=f"₹{business_amount:,.2f}").border = border
            current_row += 1
        
        # House property income
        if other_income.house_property_income:
            hpi_amount = other_income.house_property_income.calculate_net_income_from_house_property(salary_package_record.regime).to_float()
            ws.cell(row=current_row, column=1, value="House Property Income").border = border
            ws.cell(row=current_row, column=2, value=f"₹{hpi_amount:,.2f}").border = border
            current_row += 1
        
        return current_row
    
    def _add_deductions_section(self, ws, salary_package_record, start_row, section_font, section_fill, subsection_font, border):
        """Add deductions section to single sheet."""
        # Section header
        ws.cell(row=start_row, column=1, value="DEDUCTIONS").font = section_font
        ws.cell(row=start_row, column=1).fill = section_fill
        ws.merge_cells(f'A{start_row}:B{start_row}')
        
        current_row = start_row + 1
        
        deductions = salary_package_record.deductions
        
        # Section 80C
        ws.cell(row=current_row, column=1, value="Section 80C").font = subsection_font
        ws.merge_cells(f'A{current_row}:B{current_row}')
        current_row += 1
        
        section_80c_fields = [
            ("epf", "EPF"),
            ("ppf", "PPF"),
            ("life_insurance_premium", "Life Insurance Premium"),
            ("elss", "ELSS"),
            ("nsc", "NSC"),
            ("home_loan_principal", "Home Loan Principal"),
            ("tuition_fees", "Tuition Fees"),
            ("tax_saver_fd", "Tax Saver FD"),
            ("sukanya_samriddhi", "Sukanya Samriddhi"),
            ("ulip", "ULIP"),
            ("other_80c", "Other 80C Investments"),
        ]
        
        for field_name, display_name in section_80c_fields:
            if hasattr(deductions.section_80c, field_name):
                amount = getattr(deductions.section_80c, field_name).to_float()
                if amount > 0:
                    ws.cell(row=current_row, column=1, value=display_name).border = border
                    ws.cell(row=current_row, column=2, value=f"₹{amount:,.2f}").border = border
                    current_row += 1
        
        # Section 80D
        current_row += 1
        ws.cell(row=current_row, column=1, value="Section 80D").font = subsection_font
        ws.merge_cells(f'A{current_row}:B{current_row}')
        current_row += 1
        
        section_80d_fields = [
            ("health_insurance_self", "Health Insurance - Self"),
            ("health_insurance_parents", "Health Insurance - Parents"),
            ("preventive_health_checkup", "Preventive Health Checkup"),
        ]
        
        for field_name, display_name in section_80d_fields:
            if hasattr(deductions.section_80d, field_name):
                amount = getattr(deductions.section_80d, field_name).to_float()
                if amount > 0:
                    ws.cell(row=current_row, column=1, value=display_name).border = border
                    ws.cell(row=current_row, column=2, value=f"₹{amount:,.2f}").border = border
                    current_row += 1
        
        # Total deductions
        current_row += 1
        # Calculate gross income for deduction calculation
        gross_income = salary_package_record.calculate_gross_income()
        total_deductions = deductions.calculate_total_deductions(salary_package_record.regime, salary_package_record.age, gross_income).to_float()
        ws.cell(row=current_row, column=1, value="TOTAL DEDUCTIONS").border = border
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        ws.cell(row=current_row, column=2, value=f"₹{total_deductions:,.2f}").border = border
        ws.cell(row=current_row, column=2).font = Font(bold=True)
        current_row += 1
        
        return current_row
    
    def _add_tax_calculation_section(self, ws, salary_package_record, start_row, section_font, section_fill, border):
        """Add tax calculation section to single sheet."""
        # Section header
        ws.cell(row=start_row, column=1, value="TAX CALCULATION").font = section_font
        ws.cell(row=start_row, column=1).fill = section_fill
        ws.merge_cells(f'A{start_row}:B{start_row}')
        
        current_row = start_row + 1
        
        calc_result = salary_package_record.calculation_result
        
        # Tax calculation data
        calc_data = [
            ("Gross Income", calc_result.gross_income.to_float()),
            ("Total Exemptions", calc_result.total_exemptions.to_float()),
            ("Total Deductions", calc_result.total_deductions.to_float()),
            ("Taxable Income", calc_result.taxable_income.to_float()),
            ("Tax Liability", calc_result.tax_liability.to_float()),
            ("Monthly Tax", calc_result.tax_liability.divide(12).to_float()),
        ]
        
        for description, amount in calc_data:
            ws.cell(row=current_row, column=1, value=description).border = border
            ws.cell(row=current_row, column=2, value=f"₹{amount:,.2f}").border = border
            current_row += 1
        
        # Effective tax rate
        gross_income_float = calc_result.gross_income.to_float()
        tax_liability_float = calc_result.tax_liability.to_float()
        effective_rate = (tax_liability_float / gross_income_float * 100) if gross_income_float > 0 else 0
        
        ws.cell(row=current_row, column=1, value="Effective Tax Rate").border = border
        ws.cell(row=current_row, column=2, value=f"{effective_rate:.2f}%").border = border
        current_row += 1
        
        return current_row

    async def get_monthly_salary(
        self,
        employee_id: str,
        month: int,
        year: int,
        organization_id: str
    ) -> MonthlySalaryResponseDTO:
        """
        Get monthly salary for an employee for a specific month and year.
        
        Args:
            employee_id: Employee ID
            month: Month number (1-12)
            year: Year
            organization_id: Organization ID for multi-tenancy
            
        Returns:
            MonthlySalaryResponseDTO: Monthly salary details
            
        Raises:
            ValueError: If salary not found
        """
        
        logger.info(f"Getting monthly salary for employee {employee_id}, month {month}, year {year}")
        
        try:
            # Get monthly salary from repository
            monthly_salary = await self.monthly_salary_repository.get_by_employee_month_year(
                employee_id, month, year, organization_id
            )
            
            if not monthly_salary:
                raise ValueError(f"Monthly salary not found for employee {employee_id}, month {month}, year {year}")
            
            # Get user details
            from app.domain.value_objects.employee_id import EmployeeId
            employee_id_vo = EmployeeId(employee_id)
            user = await self.user_repository.get_by_id(employee_id_vo, organization_id)
            
            # Convert entity to DTO
            response = self._convert_monthly_salary_entity_to_dto(monthly_salary, user)
            
            logger.info(f"Successfully retrieved monthly salary for employee {employee_id}")
            return response
            
        except Exception as e:
            logger.error(f"Failed to get monthly salary for employee {employee_id}: {str(e)}")
            raise

    async def get_monthly_salaries_for_period(
        self,
        month: int,
        year: int,
        organization_id: str,
        salary_status: Optional[str] = None,
        department: Optional[str] = None,
        skip: int = 0,
        limit: int = 20
    ) -> Dict[str, Any]:
        """
        Get all monthly salaries for a specific month and year with filtering and pagination.
        
        Args:
            month: Month number (1-12)
            year: Year
            organization_id: Organization ID for multi-tenancy
            status: Optional status filter
            department: Optional department filter
            skip: Number of records to skip
            limit: Maximum number of records to return
            
        Returns:
            Dict containing list of monthly salaries and pagination info
        """
        
        logger.info(f"Getting monthly salaries for period: month {month}, year {year}")
        
        try:
            # Get monthly salaries from repository
            monthly_salaries = await self.monthly_salary_repository.get_by_month_year(
                month, year, organization_id, limit, skip
            )
            
            # Apply filters if provided
            filtered_salaries = []
            for salary in monthly_salaries:
                # Get user details for filtering
                from app.domain.value_objects.employee_id import EmployeeId
                employee_id_vo = EmployeeId(salary.employee_id.value)
                user = await self.user_repository.get_by_id(employee_id_vo, organization_id)
                
                # Apply status filter
                if salary_status and salary.status != salary_status:
                    continue
                
                # Apply department filter
                if department and user and user.department != department:
                    continue
                
                filtered_salaries.append(salary)
            
            # Convert entities to DTOs
            items = []
            for salary in filtered_salaries:
                from app.domain.value_objects.employee_id import EmployeeId
                employee_id_vo = EmployeeId(salary.employee_id.value)
                user = await self.user_repository.get_by_id(employee_id_vo, organization_id)
                dto = self._convert_monthly_salary_entity_to_dto(salary, user)
                items.append(dto)
            
            # Get total count for pagination
            total = len(filtered_salaries)
            
            result = {
                "items": items,
                "total": total,
                "skip": skip,
                "limit": limit,
                "has_more": (skip + limit) < total
            }
            
            logger.info(f"Successfully retrieved {len(items)} monthly salaries for period")
            return result
            
        except Exception as e:
            logger.error(f"Failed to get monthly salaries for period: {str(e)}")
            raise

    async def get_monthly_salary_summary(
        self,
        month: int,
        year: int,
        organization_id: str
    ) -> Dict[str, Any]:
        """
        Get summary statistics for monthly salaries in a specific month and year.
        
        Args:
            month: Month number (1-12)
            year: Year
            organization_id: Organization ID for multi-tenancy
            
        Returns:
            Dict containing summary statistics
        """
        
        logger.info(f"Getting monthly salary summary for period: month {month}, year {year}")
        
        try:
            # Get summary from repository
            summary = await self.monthly_salary_repository.get_monthly_summary(
                month, year, organization_id
            )
            
            logger.info(f"Successfully retrieved monthly salary summary")
            return summary
            
        except Exception as e:
            logger.error(f"Failed to get monthly salary summary: {str(e)}")
            raise

    async def delete_monthly_salary(
        self,
        employee_id: str,
        month: int,
        year: int,
        organization_id: str
    ) -> str:
        """
        Delete monthly salary record for an employee.
        
        Args:
            employee_id: Employee ID
            month: Month number (1-12)
            year: Year
            organization_id: Organization ID for multi-tenancy
            
        Returns:
            str: Success message
            
        Raises:
            ValueError: If salary not found
        """
        
        logger.info(f"Deleting monthly salary for employee {employee_id}, month {month}, year {year}")
        
        try:
            # Check if salary exists
            exists = await self.monthly_salary_repository.exists(
                employee_id, month, year, organization_id
            )
            
            if not exists:
                raise ValueError(f"Monthly salary not found for employee {employee_id}, month {month}, year {year}")
            
            # Delete the salary
            deleted = await self.monthly_salary_repository.delete(
                employee_id, month, year, organization_id
            )
            
            if not deleted:
                raise RuntimeError(f"Failed to delete monthly salary for employee {employee_id}")
            
            message = f"Successfully deleted monthly salary for employee {employee_id}, month {month}, year {year}"
            logger.info(message)
            return message
            
        except Exception as e:
            logger.error(f"Failed to delete monthly salary for employee {employee_id}: {str(e)}")
            raise

    def _convert_monthly_salary_entity_to_dto(
        self,
        monthly_salary,
        user
    ) -> MonthlySalaryResponseDTO:
        """
        Convert MonthlySalary entity to MonthlySalaryResponseDTO.
        
        Args:
            monthly_salary: MonthlySalary entity
            user: User entity (can be None)
            
        Returns:
            MonthlySalaryResponseDTO: Response DTO
        """
        
        # Calculate gross salary
        gross_salary = monthly_salary.salary.calculate_gross_salary()
        
        # Calculate deductions
        epf_employee = self._calculate_monthly_epf(gross_salary)
        esi_employee = self._calculate_monthly_esi(gross_salary)
        professional_tax = self._calculate_monthly_professional_tax(gross_salary)
        tds = monthly_salary.tax_amount
        
        # Get loan EMI amount from perquisites payouts
        loan_emi = Money.zero()
        if monthly_salary.perquisites_payouts and monthly_salary.perquisites_payouts.components:
            for component in monthly_salary.perquisites_payouts.components:
                if component.key == "loan":
                    loan_emi = component.value
                    break
        
        total_deductions = epf_employee.add(esi_employee).add(professional_tax).add(tds).add(loan_emi)
        net_salary = self._safe_subtract(gross_salary, total_deductions)
        
        # Get working days info
        working_days_info = self._get_working_days_info(monthly_salary.month, monthly_salary.year)
        
        # Tax details - handle tax_regime robustly
        try:
            if hasattr(monthly_salary.tax_regime, "regime_type"):
                if hasattr(monthly_salary.tax_regime.regime_type, "value"):
                    tax_regime_value = monthly_salary.tax_regime.regime_type.value
                else:
                    # regime_type is a string
                    tax_regime_value = str(monthly_salary.tax_regime.regime_type)
            elif hasattr(monthly_salary.tax_regime, "value"):
                tax_regime_value = monthly_salary.tax_regime.value
            else:
                # tax_regime is a string or other type
                tax_regime_value = str(monthly_salary.tax_regime)
        except Exception as e:
            # Fallback to a default value if anything goes wrong
            logger.warning(f"Error extracting tax_regime value: {e}")
            tax_regime_value = "old_regime"  # Default fallback

        return MonthlySalaryResponseDTO(
            employee_id=monthly_salary.employee_id.value,
            month=monthly_salary.month,
            year=monthly_salary.year,
            tax_year=str(monthly_salary.tax_year),
            
            # Employee details
            employee_name=user.name if user else None,
            employee_email=user.email if user else None,
            department=user.department if user else None,
            designation=user.designation if user else None,
            
            # Salary components
            basic_salary=monthly_salary.salary.basic_salary.to_float(),
            da=monthly_salary.salary.dearness_allowance.to_float(),
            hra=monthly_salary.salary.hra_provided.to_float(),
            special_allowance=monthly_salary.salary.special_allowance.to_float(),
            transport_allowance=0.0,  # Not in current model
            medical_allowance=0.0,  # Not in current model
            bonus=monthly_salary.salary.bonus.to_float(),
            commission=monthly_salary.salary.commission.to_float(),
            other_allowances=0.0,  # Would need to sum specific allowances
            arrears=monthly_salary.salary.arrears.to_float(),
            
            # Deductions
            epf_employee=epf_employee.to_float(),
            esi_employee=esi_employee.to_float(),
            professional_tax=professional_tax.to_float(),
            tds=tds.to_float(),
            advance_deduction=0.0,
            loan_deduction=loan_emi.to_float(),
            other_deductions=0.0,
            
            # Calculated totals
            gross_salary=gross_salary.to_float(),
            total_deductions=total_deductions.to_float(),
            net_salary=net_salary.to_float(),
            
            # Annual projections
            annual_gross_salary=gross_salary.multiply(12).to_float(),
            annual_tax_liability=tds.multiply(12).to_float(),
            
            # Tax details
            tax_regime=tax_regime_value,
            tax_exemptions=0.0,  # Would need to calculate
            standard_deduction=0.0,  # Would need to calculate
            
            # Working days
            total_days_in_month=working_days_info['total_days'],
            working_days_in_period=working_days_info['working_days'],
            lwp_days=working_days_info['lwp_days'],
            effective_working_days=working_days_info['effective_days'],
            
            # Status and metadata
            status=monthly_salary.status if hasattr(monthly_salary, 'status') else "computed",
            computation_date=monthly_salary.created_at.isoformat() if hasattr(monthly_salary, 'created_at') else None,
            notes=None,
            remarks=None,
            created_at=monthly_salary.created_at.isoformat() if hasattr(monthly_salary, 'created_at') else datetime.now().isoformat(),
            updated_at=monthly_salary.updated_at.isoformat() if hasattr(monthly_salary, 'updated_at') else datetime.now().isoformat(),
            created_by=None,
            updated_by=None,
            
            # Computation details
            use_declared_values=False,  # Default value
            computation_mode="actual",  # Default value
            computation_summary={
                "gross_salary": gross_salary.to_float(),
                "total_deductions": total_deductions.to_float(),
                "net_salary": net_salary.to_float(),
                "monthly_tax": tds.to_float()
            }
        )

    def _calculate_monthly_epf(self, gross_salary):
        """Calculate monthly EPF contribution (12% of basic + DA)."""
        from app.domain.value_objects.money import Money
        from decimal import Decimal
        return gross_salary.multiply(Decimal('0.12'))

    def _calculate_monthly_esi(self, gross_salary):
        """Calculate monthly ESI contribution (0.75% of gross)."""
        from app.domain.value_objects.money import Money
        from decimal import Decimal
        return gross_salary.multiply(Decimal('0.0075'))

    def _calculate_monthly_professional_tax(self, gross_salary):
        """Calculate monthly professional tax based on slabs."""
        from app.domain.value_objects.money import Money
        gross_amount = gross_salary.to_float()
        if gross_amount <= 10000:
            return Money.zero()
        elif gross_amount <= 15000:
            return Money.from_float(150.0)
        else:
            return Money.from_float(200.0)

    def _get_working_days_info(self, month: int, year: int) -> Dict[str, int]:
        """Get working days information for the month."""
        import calendar
        
        # Get total days in month
        total_days = calendar.monthrange(year, month)[1]
        
        # Assume 26 working days per month (excluding Sundays)
        working_days = 26
        lwp_days = 0  # Would need to get from attendance system
        effective_days = working_days - lwp_days
        
        return {
            'total_days': total_days,
            'working_days': working_days,
            'lwp_days': lwp_days,
            'effective_days': effective_days
        }

    def _safe_subtract(self, a, b):
        """Safely subtract two Money objects and handle negative results gracefully."""
        try:
            return a.subtract(b)
        except ValueError as e:
            if "Cannot subtract to negative amount" in str(e):
                # If subtraction would result in negative, return zero
                from app.domain.value_objects.money import Money
                return Money.zero()
            else:
                # Re-raise other errors
                raise

    # =============================================================================
    # LOAN PROCESSING OPERATIONS
    # =============================================================================
    
    async def process_loan_schedule(
        self,
        employee_id: str,
        tax_year: str,
        organization_id: str
    ) -> Dict[str, Any]:
        """
        Process loan schedule for an employee.
        
        This method retrieves the employee's loan information and calculates
        the monthly payment schedule, showing outstanding amounts, payments made,
        and interest calculations.
        
        Args:
            employee_id: Employee ID
            tax_year: Tax year
            organization_id: Organization ID
            
        Returns:
            Dict containing loan schedule information
        """
        
        logger.info(f"Processing loan schedule for employee {employee_id} for tax year {tax_year}")
        
        try:
            # Get the salary package record
            salary_package_record, _ = await self._get_or_create_salary_package_record(
                employee_id, tax_year, organization_id
            )
            
            # Extract loan information from perquisites
            loan_info = None
            if salary_package_record.perquisites and salary_package_record.perquisites.interest_free_loan:
                loan = salary_package_record.perquisites.interest_free_loan
                loan_info = {
                    "loan_amount": float(loan.loan_amount.amount),
                    "emi_amount": float(loan.emi_amount.amount),
                    "outstanding_amount": float(loan.outstanding_amount.amount),
                    "company_interest_rate": float(loan.company_interest_rate),
                    "sbi_interest_rate": float(loan.sbi_interest_rate),
                    "loan_type": loan.loan_type,
                    "loan_start_date": loan.loan_start_date.isoformat() if loan.loan_start_date else None
                }
                
                # Calculate monthly payment schedules
                company_schedule, company_interest_paid = loan.calculate_monthly_payment_schedule(loan.company_interest_rate)
                sbi_schedule, sbi_interest_paid = loan.calculate_monthly_payment_schedule(loan.sbi_interest_rate)
                
                # Convert schedules to serializable format
                def serialize_schedule(schedule):
                    return [
                        {
                            "month": item.month,
                            "outstanding_amount": float(item.outstanding_amount.amount),
                            "principal_amount": float(item.principal_amount.amount),
                            "interest_amount": float(item.interest_amount.amount),
                            "emi_deducted": float(item.principal_amount.amount + item.interest_amount.amount)
                        }
                        for item in schedule
                    ]
                
                loan_info.update({
                    "company_schedule": serialize_schedule(company_schedule),
                    "sbi_schedule": serialize_schedule(sbi_schedule),
                    "company_interest_paid": float(company_interest_paid.amount),
                    "sbi_interest_paid": float(sbi_interest_paid.amount),
                    "interest_saved": float(sbi_interest_paid.amount - company_interest_paid.amount),
                    "taxable_benefit": float(loan.calculate_taxable_loan_value().amount)
                })
            
            # Get employee information
            employee_info = {
                "employee_id": employee_id,
                "tax_year": tax_year,
                "has_loan": loan_info is not None
            }
            
            if loan_info:
                # Calculate summary statistics
                total_months = len(loan_info["company_schedule"])
                total_principal_paid = sum(item["principal_amount"] for item in loan_info["company_schedule"])
                total_interest_paid = sum(item["interest_amount"] for item in loan_info["company_schedule"])
                
                loan_info.update({
                    "summary": {
                        "total_months": total_months,
                        "total_principal_paid": total_principal_paid,
                        "total_interest_paid": total_interest_paid,
                        "total_payments_made": total_principal_paid + total_interest_paid,
                        "remaining_principal": loan_info["outstanding_amount"] - total_principal_paid,
                        "average_monthly_payment": (total_principal_paid + total_interest_paid) / total_months if total_months > 0 else 0
                    }
                })
            
            result = {
                "employee_info": employee_info,
                "loan_info": loan_info,
                "processing_date": datetime.now().isoformat()
            }
            
            logger.info(f"Successfully processed loan schedule for employee {employee_id}")
            return result
            
        except Exception as e:
            logger.error(f"Error processing loan schedule for employee {employee_id}: {str(e)}")
            raise

    async def get_employee_salary_history(
        self,
        employee_id: str,
        organization_id: str,
        limit: int = 100,
        offset: int = 0
    ) -> List[MonthlySalaryResponseDTO]:
        """
        Get all monthly salary records for an employee.
        Args:
            employee_id: Employee ID
            organization_id: Organization ID for multi-tenancy
            limit: Maximum number of records to return
            offset: Number of records to skip
        Returns:
            List of MonthlySalaryResponseDTO
        """
        try:
            salary_entities = await self.monthly_salary_repository.get_by_employee(
                employee_id, organization_id, limit=limit, offset=offset
            )
            # Optionally enrich with user info
            from app.domain.value_objects.employee_id import EmployeeId
            items = []
            for salary in salary_entities:
                employee_id_vo = EmployeeId(salary.employee_id.value)
                user = await self.user_repository.get_by_id(employee_id_vo, organization_id)
                dto = self._convert_monthly_salary_entity_to_dto(salary, user)
                items.append(dto)
            return items
        except Exception as e:
            logger.error(f"Failed to get salary history for employee {employee_id}: {str(e)}")
            raise

    async def download_payslip(
        self,
        employee_id: str,
        month: int,
        year: int,
        organization_id: str
    ) -> bytes:
        """
        Download payslip for a specific month.
        ...
        """
        try:
            # Get salary record
            salary_record = await self.monthly_salary_repository.get_by_employee_month_year(
                employee_id=employee_id,
                month=month,
                year=year,
                organization_id=organization_id
            )
            if not salary_record:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail=f"Salary record not found for employee {employee_id} for {month}/{year}"
                )
            # Fetch organisation for logo and details
            from app.domain.value_objects.organisation_id import OrganisationId
            organisation = await self.organisation_repository.get_by_id(OrganisationId(organization_id))
            logo_path = None
            organisation_details = {
                'name': 'COMPANY NAME',
                'address': 'Company Address',
                'contact': 'Phone: +91-XXXXXXXXXX | Email: info@company.com'
            }
            
            if organisation:
                if getattr(organisation, 'logo_path', None):
                    logo_path = organisation.logo_path
                
                # Extract organisation details
                organisation_details['name'] = getattr(organisation, 'name', 'COMPANY NAME')
                
                if hasattr(organisation, 'address') and organisation.address:
                    address_parts = []
                    if getattr(organisation.address, 'street_address', None):
                        address_parts.append(organisation.address.street_address)
                    if getattr(organisation.address, 'city', None):
                        address_parts.append(organisation.address.city)
                    if getattr(organisation.address, 'state', None):
                        address_parts.append(organisation.address.state)
                    if getattr(organisation.address, 'pin_code', None):
                        address_parts.append(organisation.address.pin_code)
                    organisation_details['address'] = ", ".join(address_parts) if address_parts else "Company Address"
                
                if hasattr(organisation, 'contact_info') and organisation.contact_info:
                    contact_parts = []
                    if getattr(organisation.contact_info, 'phone', None):
                        contact_parts.append(f"Phone: {organisation.contact_info.phone}")
                    if getattr(organisation.contact_info, 'email', None):
                        contact_parts.append(f"Email: {organisation.contact_info.email}")
                    organisation_details['contact'] = " | ".join(contact_parts) if contact_parts else "Phone: +91-XXXXXXXXXX | Email: info@company.com"
            
            # Fetch user details for employee information
            from app.domain.value_objects.employee_id import EmployeeId
            user = await self.user_repository.get_by_id(EmployeeId(employee_id), organization_id)
            user_details = {
                'name': 'N/A',
                'department': 'N/A',
                'designation': 'N/A',
                'bank_name': 'N/A',
                'account_no': 'N/A',
                'ifsc_code': 'N/A',
                'branch': 'N/A'
            }
            
            if user:
                user_details['name'] = getattr(user, 'name', 'N/A')
                user_details['department'] = getattr(user, 'department', 'N/A')
                user_details['designation'] = getattr(user, 'designation', 'N/A')
                
                # Extract bank details if available
                if hasattr(user, 'bank_details') and user.bank_details:
                    user_details['bank_name'] = getattr(user.bank_details, 'bank_name', 'N/A')
                    user_details['account_no'] = getattr(user.bank_details, 'account_number', 'N/A')
                    user_details['ifsc_code'] = getattr(user.bank_details, 'ifsc_code', 'N/A')
                    user_details['branch'] = getattr(user.bank_details, 'branch', 'N/A')
            
            # Generate payslip content
            payslip_content = self._generate_payslip_text(salary_record)
            # Convert to PDF bytes, pass logo_path, organisation details, and user details
            pdf_bytes = self._text_to_pdf(payslip_content, logo_path=logo_path, organisation_details=organisation_details, user_details=user_details)
            return pdf_bytes
        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"Error downloading payslip: {str(e)}")
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Failed to download payslip: {str(e)}"
            )

    def _generate_payslip_text(self, salary_record) -> str:
        """Generate payslip text content."""
        month_names = [
            'January', 'February', 'March', 'April', 'May', 'June',
            'July', 'August', 'September', 'October', 'November', 'December'
        ]

        # Access salary components via salary_record.salary
        s = salary_record.salary

        # Calculate gross salary
        gross_salary = s.calculate_gross_salary()

        # Calculate deductions using the same logic as DTO conversion
        epf_employee = self._calculate_monthly_epf(gross_salary)
        esi_employee = self._calculate_monthly_esi(gross_salary)
        professional_tax = self._calculate_monthly_professional_tax(gross_salary)
        tds = salary_record.tax_amount

        total_deductions = epf_employee.add(esi_employee).add(professional_tax).add(tds)
        net_salary = self._safe_subtract(gross_salary, total_deductions)

        # Get working days info
        working_days_info = self._get_working_days_info(salary_record.month, salary_record.year)

        # Calculate specific allowances total
        specific_allowances_total = 0.0
        if s.specific_allowances:
            specific_allowances_total = s.specific_allowances.calculate_total_specific_allowances().to_float()

        payslip = f"""
        COMPANY PAYSLIP
        ===============

        Employee Details:
        ----------------
        Employee ID: {salary_record.employee_id.value}
        Name: {getattr(salary_record, 'employee_name', 'N/A')}
        Department: {getattr(salary_record, 'department', 'N/A')}
        Designation: {getattr(salary_record, 'designation', 'N/A')}

        Pay Period: {month_names[salary_record.month - 1]} {salary_record.year}
        Tax Year: {str(salary_record.tax_year)}

        Earnings:
        ---------
        Basic Salary: ₹{s.basic_salary.to_float():,.2f}
        Dearness Allowance: ₹{s.dearness_allowance.to_float():,.2f}
        House Rent Allowance: ₹{s.hra_provided.to_float():,.2f}
        Special Allowance: ₹{s.special_allowance.to_float():,.2f}
        Transport Allowance: ₹{0.0:,.2f}
        Bonus: ₹{s.bonus.to_float():,.2f}
        Commission: ₹{s.commission.to_float():,.2f}
        Other Allowances: ₹{specific_allowances_total:,.2f}
        Arrears: ₹{s.arrears.to_float():,.2f}

        Total Earnings: ₹{gross_salary.to_float():,.2f}

        Deductions:
        -----------
        EPF Employee: ₹{epf_employee.to_float():,.2f}
        ESI Employee: ₹{esi_employee.to_float():,.2f}
        Professional Tax: ₹{professional_tax.to_float():,.2f}
        TDS: ₹{tds.to_float():,.2f}
        Advance Deduction: ₹{0.0:,.2f}
        Loan Deduction: ₹{0.0:,.2f}
        Other Deductions: ₹{0.0:,.2f}

        Total Deductions: ₹{total_deductions.to_float():,.2f}

        Net Pay: ₹{net_salary.to_float():,.2f}

        Working Days: {working_days_info['effective_days']}/{working_days_info['total_days']}
        LWP Days: {working_days_info['lwp_days']}

        Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
        """

        return payslip

    def _text_to_pdf(self, text: str, logo_path: str = None, organisation_details: dict = None, user_details: dict = None) -> bytes:
        """Convert text to PDF bytes using ReportLab with side-by-side layout and logo."""
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from io import BytesIO
        import os
        from datetime import datetime
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=0.5*inch, rightMargin=0.5*inch)
        
        # Register a Unicode-compatible font for better symbol support
        try:
            # Try to register a system font that supports Unicode
            pdfmetrics.registerFont(TTFont('DejaVuSans', '/System/Library/Fonts/Supplemental/Arial Unicode MS.ttf'))
            default_font = 'DejaVuSans'
        except:
            try:
                # Fallback to another common Unicode font
                pdfmetrics.registerFont(TTFont('DejaVuSans', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'))
                default_font = 'DejaVuSans'
            except:
                # Use default font if Unicode fonts not available
                default_font = 'Helvetica'
        
        styles = getSampleStyleSheet()
        
        # Custom styles with Unicode font
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=20,
            alignment=TA_CENTER,
            textColor=colors.darkblue,
            fontName=default_font
        )
        
        header_style = ParagraphStyle(
            'HeaderStyle',
            parent=styles['Normal'],
            fontSize=12,
            spaceAfter=6,
            alignment=TA_LEFT,
            textColor=colors.darkblue,
            fontName=default_font
        )
        
        normal_style = ParagraphStyle(
            'NormalStyle',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=3,
            fontName=default_font
        )
        
        amount_style = ParagraphStyle(
            'AmountStyle',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=3,
            alignment=TA_RIGHT,
            fontName=default_font
        )
        
        story = []
        
        # Add logo if available
        if logo_path and os.path.isfile(logo_path):
            try:
                img = Image(logo_path)
                img.drawHeight = 0.8*inch
                img.drawWidth = 0.8*inch
                img.hAlign = 'CENTER'
                story.append(img)
                story.append(Spacer(1, 10))
            except Exception as e:
                story.append(Paragraph("[Logo could not be loaded]", normal_style))
        
        # Company header - use actual organisation details
        if organisation_details:
            organisation_name = organisation_details.get('name', 'COMPANY NAME')
            organisation_address = organisation_details.get('address', 'Company Address')
            organisation_contact = organisation_details.get('contact', 'Phone: +91-XXXXXXXXXX | Email: info@company.com')
        else:
            organisation_name = "COMPANY NAME"
            organisation_address = "Company Address"
            organisation_contact = "Phone: +91-XXXXXXXXXX | Email: info@company.com"
        
        story.append(Paragraph(organisation_name, title_style))
        story.append(Paragraph(organisation_address, header_style))
        story.append(Paragraph(organisation_contact, header_style))
        story.append(Spacer(1, 20))
        
        # Title
        story.append(Paragraph("PAYSLIP", title_style))
        story.append(Spacer(1, 15))
        
        # Parse the text to extract data and replace ₹ with Rs.
        lines = text.strip().split('\n')
        employee_data = {}
        earnings_data = []
        deductions_data = []
        
        for line in lines:
            line = line.strip()
            if not line or line.startswith('==='):
                continue
            elif ':' in line:
                parts = line.split(':', 1)
                if len(parts) == 2:
                    key, value = parts[0].strip(), parts[1].strip()
                    # Replace ₹ with Rs. for better PDF compatibility
                    value = value.replace('₹', 'Rs.')
                    
                    if key == "Employee ID":
                        employee_data['employee_id'] = value
                    elif key == "Pay Period":
                        employee_data['pay_period'] = value
                    elif key == "Tax Year":
                        employee_data['tax_year'] = value
                    elif key == "Working Days":
                        employee_data['working_days'] = value
                    elif key == "Status":
                        employee_data['status'] = value
                    elif key == "Generated on":
                        employee_data['generated_on'] = value
                    elif value.startswith('Rs.') and key not in ["Total Earnings", "Total Deductions", "Net Pay"]:
                        if "Earnings" in key or any(earning in key for earning in ["Basic Salary", "Dearness Allowance", "HRA", "Special Allowance", "Transport", "Medical", "Bonus", "Commission", "Other Allowances", "Arrears"]):
                            earnings_data.append((key, value))
                        elif "Deductions" in key or any(deduction in key for deduction in ["EPF", "ESI", "Professional Tax", "TDS", "Advance", "Loan", "Other Deductions"]):
                            deductions_data.append((key, value))
        
        # Employee and Bank Details Table - use user_details if available
        if user_details:
            employee_table_data = [
                ["Employee Details", "Bank Details"],
                ["Employee ID: " + employee_data.get('employee_id', 'N/A'), "Bank Name: " + user_details.get('bank_name', 'N/A')],
                ["Name: " + user_details.get('name', 'N/A'), "Account No: " + user_details.get('account_no', 'N/A')],
                ["Department: " + user_details.get('department', 'N/A'), "IFSC Code: " + user_details.get('ifsc_code', 'N/A')],
                ["Designation: " + user_details.get('designation', 'N/A'), "Branch: " + user_details.get('branch', 'N/A')],
                ["Pay Period: " + employee_data.get('pay_period', 'N/A'), ""],
                ["Tax Year: " + employee_data.get('tax_year', 'N/A'), ""],
            ]
        else:
            # Fallback to parsing from text
            employee_table_data = [
                ["Employee Details", "Bank Details"],
                ["Employee ID: " + employee_data.get('employee_id', 'N/A'), "Bank Name: " + employee_data.get('bank_name', 'N/A')],
                ["Name: " + employee_data.get('name', 'N/A'), "Account No: " + employee_data.get('account_no', 'N/A')],
                ["Department: " + employee_data.get('department', 'N/A'), "IFSC Code: " + employee_data.get('ifsc_code', 'N/A')],
                ["Designation: " + employee_data.get('designation', 'N/A'), "Branch: " + employee_data.get('branch', 'N/A')],
                ["Pay Period: " + employee_data.get('pay_period', 'N/A'), ""],
                ["Tax Year: " + employee_data.get('tax_year', 'N/A'), ""],
            ]
        
        employee_table = Table(employee_table_data, colWidths=[doc.width/2, doc.width/2])
        employee_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        story.append(employee_table)
        story.append(Spacer(1, 20))
        
        # Summary Grid - extract actual values from parsed data
        gross_pay = "Rs.0.00"
        total_deductions = "Rs.0.00"
        net_pay = "Rs.0.00"
        
        # Find totals in the parsed data
        for line in lines:
            line = line.strip()
            if ':' in line:
                parts = line.split(':', 1)
                if len(parts) == 2:
                    key, value = parts[0].strip(), parts[1].strip()
                    value = value.replace('₹', 'Rs.')  # Replace ₹ with Rs.
                    if key == "Total Earnings":
                        gross_pay = value
                    elif key == "Total Deductions":
                        total_deductions = value
                    elif key == "Net Pay":
                        net_pay = value
        
        summary_data = [
            ["Gross Pay", "Total Deductions", "Net Pay"],
            [gross_pay, total_deductions, net_pay]
        ]
        
        summary_table = Table(summary_data, colWidths=[doc.width/3, doc.width/3, doc.width/3])
        summary_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('FONTSIZE', (0, 1), (-1, 1), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # Earnings and Deductions side by side
        # Prepare earnings data
        earnings_rows = [["Earnings", "Amount"]]
        for key, value in earnings_data:
            earnings_rows.append([key, value])
        
        # Prepare deductions data
        deductions_rows = [["Deductions", "Amount"]]
        for key, value in deductions_data:
            deductions_rows.append([key, value])
        
        # Create side-by-side table
        combined_data = []
        max_rows = max(len(earnings_rows), len(deductions_rows))
        
        for i in range(max_rows):
            earnings_row = earnings_rows[i] if i < len(earnings_rows) else ["", ""]
            deductions_row = deductions_rows[i] if i < len(deductions_rows) else ["", ""]
            combined_data.append([earnings_row[0], earnings_row[1], "", deductions_row[0], deductions_row[1]])
        
        combined_table = Table(combined_data, colWidths=[doc.width*0.35, doc.width*0.15, doc.width*0.1, doc.width*0.35, doc.width*0.15])
        combined_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),  # Amount columns
            ('ALIGN', (4, 0), (4, -1), 'RIGHT'),  # Amount columns
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (3, 0), (4, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 0), (1, 0), colors.lightblue),
            ('BACKGROUND', (3, 0), (4, 0), colors.lightcoral),
            ('GRID', (0, 0), (1, -1), 1, colors.black),
            ('GRID', (3, 0), (4, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('SPAN', (0, 0), (1, 0)),  # Merge earnings header
            ('SPAN', (3, 0), (4, 0)),  # Merge deductions header
        ]))
        story.append(combined_table)
        story.append(Spacer(1, 20))
        
        # Footer
        footer_data = [
            ["Working Days: " + employee_data.get('working_days', 'N/A')],
            ["Generated on: " + employee_data.get('generated_on', datetime.now().strftime('%Y-%m-%d %H:%M:%S')), ""]
        ]
        
        footer_table = Table(footer_data, colWidths=[doc.width/2, doc.width/2])
        footer_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.grey),
        ]))
        story.append(footer_table)
        
        doc.build(story)
        pdf_bytes = buffer.getvalue()
        buffer.close()
        return pdf_bytes
 