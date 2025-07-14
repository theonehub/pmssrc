"""
File Generation Service Implementation
Handles generation of various export file formats
"""

import csv
import io
import logging
import re
from typing import List, Dict, Any, Optional
from datetime import datetime
from decimal import Decimal

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.application.interfaces.services.file_generation_service import FileGenerationService
from app.application.interfaces.repositories.organisation_repository import OrganisationRepository
from app.application.interfaces.repositories.user_repository import UserRepository
from app.auth.auth_dependencies import CurrentUser


logger = logging.getLogger(__name__)


class FileGenerationServiceImpl(FileGenerationService):
    """
    Implementation of file generation service for various export formats.
    """
    
    def __init__(
        self,
        organisation_repository: OrganisationRepository,
        user_repository: UserRepository
    ):
        self.organisation_repository = organisation_repository
        self.user_repository = user_repository
    
    async def generate_salary_csv(
        self,
        salary_data: List[Dict[str, Any]],
        current_user: CurrentUser,
        filters: Optional[Dict[str, Any]] = None
    ) -> bytes:
        """Generate CSV file for salary data"""
        try:
            output = io.StringIO()
            writer = csv.writer(output)
            
            # Header
            headers = [
                'Employee ID', 'Employee Name', 'Department', 'Designation',
                'Month', 'Year', 'Basic Salary', 'HRA', 'DA', 'Other Allowances',
                'Gross Salary', 'PF', 'PT', 'TDS', 'Net Salary', 'Status'
            ]
            writer.writerow(headers)
            
            # Data rows
            for salary in salary_data:
                row = [
                    salary.get('employee_id', ''),
                    salary.get('employee_name', ''),
                    salary.get('department', ''),
                    salary.get('designation', ''),
                    salary.get('month', ''),
                    salary.get('year', ''),
                    salary.get('basic_salary', 0),
                    salary.get('hra', 0),
                    salary.get('da', 0),
                    salary.get('other_allowances', 0),
                    salary.get('gross_salary', 0),
                    salary.get('pf', 0),
                    salary.get('pt', 0),
                    salary.get('tds', 0),
                    salary.get('net_salary', 0),
                    salary.get('status', '')
                ]
                writer.writerow(row)
            
            return output.getvalue().encode('utf-8')
            
        except Exception as e:
            logger.error(f"Error generating salary CSV: {e}")
            raise
    
    async def generate_salary_excel(
        self,
        salary_data: List[Dict[str, Any]],
        current_user: CurrentUser,
        filters: Optional[Dict[str, Any]] = None
    ) -> bytes:
        """Generate Excel file for salary data"""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Salary Data"
            
            # Styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Header
            headers = [
                'Employee ID', 'Employee Name', 'Department', 'Designation',
                'Month', 'Year', 'Basic Salary', 'HRA', 'DA', 'Other Allowances',
                'Gross Salary', 'PF', 'PT', 'TDS', 'Net Salary', 'Status'
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Data rows
            for row_idx, salary in enumerate(salary_data, 2):
                ws.cell(row=row_idx, column=1, value=salary.get('employee_id', ''))
                ws.cell(row=row_idx, column=2, value=salary.get('employee_name', ''))
                ws.cell(row=row_idx, column=3, value=salary.get('department', ''))
                ws.cell(row=row_idx, column=4, value=salary.get('designation', ''))
                ws.cell(row=row_idx, column=5, value=salary.get('month', ''))
                ws.cell(row=row_idx, column=6, value=salary.get('year', ''))
                ws.cell(row=row_idx, column=7, value=salary.get('basic_salary', 0))
                ws.cell(row=row_idx, column=8, value=salary.get('hra', 0))
                ws.cell(row=row_idx, column=9, value=salary.get('da', 0))
                ws.cell(row=row_idx, column=10, value=salary.get('other_allowances', 0))
                ws.cell(row=row_idx, column=11, value=salary.get('gross_salary', 0))
                ws.cell(row=row_idx, column=12, value=salary.get('pf', 0))
                ws.cell(row=row_idx, column=13, value=salary.get('pt', 0))
                ws.cell(row=row_idx, column=14, value=salary.get('tds', 0))
                ws.cell(row=row_idx, column=15, value=salary.get('net_salary', 0))
                ws.cell(row=row_idx, column=16, value=salary.get('status', ''))
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()
            
        except Exception as e:
            logger.error(f"Error generating salary Excel: {e}")
            raise
    
    async def generate_salary_bank_transfer(
        self,
        salary_data: List[Dict[str, Any]],
        current_user: CurrentUser,
        filters: Optional[Dict[str, Any]] = None
    ) -> bytes:
        """Generate bank transfer format file for salary data"""
        try:
            output = io.StringIO()
            writer = csv.writer(output)
            
            # Bank transfer header
            headers = [
                'Employee ID', 'Employee Name', 'Bank Account Number', 'IFSC Code',
                'Bank Name', 'Branch', 'Net Salary', 'Month', 'Year'
            ]
            writer.writerow(headers)
            
            # Data rows
            for salary in salary_data:
                # Get user bank details
                user = await self.user_repository.get_by_id(salary.get('employee_id'), current_user.hostname)
                bank_details = user.bank_details if user else None
                
                row = [
                    salary.get('employee_id', ''),
                    salary.get('employee_name', ''),
                    bank_details.account_number if bank_details else '',
                    bank_details.ifsc_code if bank_details else '',
                    bank_details.bank_name if bank_details else '',
                    bank_details.branch_name if bank_details else '',
                    salary.get('net_salary', 0),
                    salary.get('month', ''),
                    salary.get('year', '')
                ]
                writer.writerow(row)
            
            return output.getvalue().encode('utf-8')
            
        except Exception as e:
            logger.error(f"Error generating bank transfer file: {e}")
            raise
    
    async def generate_tds_csv(
        self,
        tds_data: List[Dict[str, Any]],
        current_user: CurrentUser,
        filters: Optional[Dict[str, Any]] = None
    ) -> bytes:
        """Generate CSV file for TDS data"""
        try:
            output = io.StringIO()
            writer = csv.writer(output)
            
            # Header
            headers = [
                'Employee ID', 'Employee Name', 'PAN Number', 'Department',
                'Month', 'Year', 'Gross Salary', 'TDS Amount', 'Tax Regime',
                'Status'
            ]
            writer.writerow(headers)
            
            # Data rows
            for tds in tds_data:
                row = [
                    tds.get('employee_id', ''),
                    tds.get('employee_name', ''),
                    tds.get('pan_number', ''),
                    tds.get('department', ''),
                    tds.get('month', ''),
                    tds.get('year', ''),
                    tds.get('gross_salary', 0),
                    tds.get('tds', 0),
                    tds.get('tax_regime', ''),
                    tds.get('status', '')
                ]
                writer.writerow(row)
            
            return output.getvalue().encode('utf-8')
            
        except Exception as e:
            logger.error(f"Error generating TDS CSV: {e}")
            raise
    
    async def generate_tds_excel(
        self,
        tds_data: List[Dict[str, Any]],
        current_user: CurrentUser,
        filters: Optional[Dict[str, Any]] = None
    ) -> bytes:
        """Generate Excel file for TDS data"""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "TDS Data"
            
            # Styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Header
            headers = [
                'Employee ID', 'Employee Name', 'PAN Number', 'Department',
                'Month', 'Year', 'Gross Salary', 'TDS Amount', 'Tax Regime',
                'Status'
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Data rows
            for row_idx, tds in enumerate(tds_data, 2):
                ws.cell(row=row_idx, column=1, value=tds.get('employee_id', ''))
                ws.cell(row=row_idx, column=2, value=tds.get('employee_name', ''))
                ws.cell(row=row_idx, column=3, value=tds.get('pan_number', ''))
                ws.cell(row=row_idx, column=4, value=tds.get('department', ''))
                ws.cell(row=row_idx, column=5, value=tds.get('month', ''))
                ws.cell(row=row_idx, column=6, value=tds.get('year', ''))
                ws.cell(row=row_idx, column=7, value=tds.get('gross_salary', 0))
                ws.cell(row=row_idx, column=8, value=tds.get('tds', 0))
                ws.cell(row=row_idx, column=9, value=tds.get('tax_regime', ''))
                ws.cell(row=row_idx, column=10, value=tds.get('status', ''))
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()
            
        except Exception as e:
            logger.error(f"Error generating TDS Excel: {e}")
            raise
    
    async def generate_form_16(
        self,
        tds_data: List[Dict[str, Any]],
        current_user: CurrentUser,
        employee_id: Optional[str] = None,
        tax_year: Optional[str] = None
    ) -> bytes:
        """Generate Form 16 format file"""
        try:
            # Get organisation details
            organisation = await self.organisation_repository.get_by_id(current_user.hostname)
            
            output = io.StringIO()
            writer = csv.writer(output)
            
            # Form 16 header
            writer.writerow(['FORM 16 - CERTIFICATE OF DEDUCTION OF TAX AT SOURCE'])
            writer.writerow([''])
            writer.writerow(['Deductor Details:'])
            writer.writerow(['Name:', organisation.name if organisation else 'MISSING'])
            writer.writerow(['TAN:', organisation.tax_info.tan_number if organisation and organisation.tax_info else 'MISSING'])
            writer.writerow(['PAN:', organisation.tax_info.pan_number if organisation and organisation.tax_info else 'MISSING'])
            writer.writerow(['Address:', organisation.address.street_address if organisation and organisation.address else 'MISSING'])
            writer.writerow([''])
            
            # Employee details
            if employee_id:
                employee_data = [tds for tds in tds_data if tds.get('employee_id') == employee_id]
            else:
                employee_data = tds_data
            
            if employee_data:
                emp = employee_data[0]
                writer.writerow(['Employee Details:'])
                writer.writerow(['Name:', emp.get('employee_name', 'MISSING')])
                writer.writerow(['PAN:', emp.get('pan_number', 'MISSING')])
                writer.writerow([''])
            
            # TDS summary
            writer.writerow(['TDS Summary:'])
            total_tds = sum(tds.get('tds', 0) for tds in employee_data)
            writer.writerow(['Total TDS Deducted:', total_tds])
            writer.writerow([''])
            
            # Monthly breakdown
            writer.writerow(['Monthly Breakdown:'])
            headers = ['Month', 'Year', 'Gross Salary', 'TDS Amount']
            writer.writerow(headers)
            
            for tds in employee_data:
                row = [
                    tds.get('month', ''),
                    tds.get('year', ''),
                    tds.get('gross_salary', 0),
                    tds.get('tds', 0)
                ]
                writer.writerow(row)
            
            return output.getvalue().encode('utf-8')
            
        except Exception as e:
            logger.error(f"Error generating Form 16: {e}")
            raise
    
    async def generate_form_24q(
        self,
        tds_data: List[Dict[str, Any]],
        current_user: CurrentUser,
        quarter: int,
        tax_year: int
    ) -> bytes:
        """Generate Form 24Q format file"""
        try:
            # Get organisation details
            organisation = await self.organisation_repository.get_by_id(current_user.hostname)
            
            output = io.StringIO()
            writer = csv.writer(output)
            
            # Form 24Q header
            writer.writerow(['FORM 24Q - STATEMENT OF TDS FROM SALARIES'])
            writer.writerow([''])
            writer.writerow(['Deductor Details:'])
            writer.writerow(['Name:', organisation.name if organisation else 'MISSING'])
            writer.writerow(['TAN:', organisation.tax_info.tan_number if organisation and organisation.tax_info else 'MISSING'])
            writer.writerow(['PAN:', organisation.tax_info.pan_number if organisation and organisation.tax_info else 'MISSING'])
            writer.writerow(['Quarter:', f'Q{quarter}'])
            writer.writerow(['Financial Year:', f'{tax_year}-{tax_year+1}'])
            writer.writerow([''])
            
            # Deductee details
            writer.writerow(['Deductee Details:'])
            headers = ['Sr No', 'PAN', 'Name', 'Gross Salary', 'TDS Amount', 'Month', 'Year']
            writer.writerow(headers)
            
            for idx, tds in enumerate(tds_data, 1):
                row = [
                    idx,
                    tds.get('pan_number', 'MISSING'),
                    tds.get('employee_name', 'MISSING'),
                    tds.get('gross_salary', 0),
                    tds.get('tds', 0),
                    tds.get('month', ''),
                    tds.get('year', '')
                ]
                writer.writerow(row)
            
            # Summary
            writer.writerow([''])
            total_tds = sum(tds.get('tds', 0) for tds in tds_data)
            writer.writerow(['Total TDS:', total_tds])
            
            return output.getvalue().encode('utf-8')
            
        except Exception as e:
            logger.error(f"Error generating Form 24Q: {e}")
            raise
    
    async def generate_fvu_form_24q(
        self,
        tds_data: List[Dict[str, Any]],
        current_user: CurrentUser,
        quarter: int,
        tax_year: int
    ) -> bytes:
        """Generate FVU (File Validation Utility) format for Form 24Q"""
        try:
            # Get organisation details
            organisation = await self.organisation_repository.get_by_id(current_user.hostname)
            
            lines = []
            
            # FVU fields
            TAN = organisation.tax_info.tan_number if organisation and organisation.tax_info else 'MISSING'
            FVU_VERSION = '7.9'
            FIN_YEAR = f'{tax_year}-{str(tax_year + 1)[-2:]}'
            QTR = f'Q{quarter}'
            DEDUCTOR_NAME = organisation.name if organisation else 'MISSING'
            DEDUCTOR_ADDR = organisation.address.street_address if organisation and organisation.address else 'MISSING'
            DEDUCTOR_PIN = organisation.address.pin_code if organisation and organisation.address else 'MISSING'
            DEDUCTOR_STATE = organisation.address.state if organisation and organisation.address else 'MISSING'
            DEDUCTOR_PAN = organisation.tax_info.pan_number if organisation and organisation.tax_info else 'MISSING'
            CHALLAN_NO = 'MISSING'
            BSR_CODE = 'MISSING'
            CHALLAN_DATE = 'MISSING'
            CHALLAN_AMOUNT = sum(tds.get('tds', 0) for tds in tds_data)
            
            # 1. File Header
            lines.append('|'.join(['FH', FVU_VERSION, TAN, FIN_YEAR, QTR]))
            
            # 2. Batch Header
            lines.append('|'.join(['BH', TAN, DEDUCTOR_NAME, DEDUCTOR_ADDR, DEDUCTOR_PIN, DEDUCTOR_STATE, DEDUCTOR_PAN, FIN_YEAR, QTR, '24Q', 'Regular']))
            
            # 3. Challan Details
            lines.append('|'.join(['CD', BSR_CODE, CHALLAN_DATE, CHALLAN_NO, str(CHALLAN_AMOUNT), 'MISSING', 'MISSING', 'MISSING', 'MISSING', 'MISSING']))
            
            # 4. Deductee Details
            for idx, tds in enumerate(tds_data, 1):
                lines.append('|'.join([
                    'DD',
                    str(idx),
                    tds.get('pan_number', 'MISSING'),
                    tds.get('employee_name', 'N/A'),
                    str(tds.get('gross_salary', 0)),
                    str(tds.get('tds', 0)),
                    str(tds.get('month', '')),
                    str(tds.get('year', '')),
                    tds.get('department', 'N/A'),
                    tds.get('designation', 'N/A')
                ]))
            
            # 5. End of File
            lines.append('EOF')
            
            return '\r\n'.join(lines).encode('utf-8')
            
        except Exception as e:
            logger.error(f"Error generating FVU Form 24Q: {e}")
            raise
    
    async def generate_processed_salaries_export(
        self,
        salary_data: List[Dict[str, Any]],
        current_user: CurrentUser,
        format_type: str,
        filters: Optional[Dict[str, Any]] = None
    ) -> bytes:
        """Generate processed salaries export in specified format"""
        try:
            if format_type.lower() == 'csv':
                return await self.generate_salary_csv(salary_data, current_user, filters)
            elif format_type.lower() == 'excel':
                return await self.generate_salary_excel(salary_data, current_user, filters)
            elif format_type.lower() == 'bank_transfer':
                return await self.generate_salary_bank_transfer(salary_data, current_user, filters)
            else:
                raise ValueError(f"Unsupported format: {format_type}")
                
        except Exception as e:
            logger.error(f"Error generating processed salaries export: {e}")
            raise
    
    async def generate_tds_report_export(
        self,
        tds_data: List[Dict[str, Any]],
        current_user: CurrentUser,
        format_type: str,
        quarter: Optional[int] = None,
        tax_year: Optional[int] = None,
        filters: Optional[Dict[str, Any]] = None
    ) -> bytes:
        """Generate TDS report export in specified format"""
        try:
            if format_type.lower() == 'csv':
                return await self.generate_tds_csv(tds_data, current_user, filters)
            elif format_type.lower() == 'excel':
                return await self.generate_tds_excel(tds_data, current_user, filters)
            elif format_type.lower() == 'form_16':
                return await self.generate_form_16(tds_data, current_user, None, str(tax_year) if tax_year else None)
            elif format_type.lower() == 'form_24q':
                if not quarter or not tax_year:
                    raise ValueError("Quarter and tax_year are required for Form 24Q")
                return await self.generate_form_24q(tds_data, current_user, quarter, tax_year)
            elif format_type.lower() == 'fvu':
                if not quarter or not tax_year:
                    raise ValueError("Quarter and tax_year are required for FVU")
                return await self.generate_fvu_form_24q(tds_data, current_user, quarter, tax_year)
            else:
                raise ValueError(f"Unsupported format: {format_type}")
                
        except Exception as e:
            logger.error(f"Error generating TDS report export: {e}")
            raise

    async def generate_pf_report_export(
        self,
        pf_data: List[Dict[str, Any]],
        current_user: CurrentUser,
        format_type: str,
        quarter: Optional[int] = None,
        tax_year: Optional[int] = None,
        filters: Optional[Dict[str, Any]] = None
    ) -> bytes:
        """Generate PF report export in specified format"""
        try:
            if format_type.lower() == 'csv':
                return await self.generate_pf_csv(pf_data, current_user, filters)
            elif format_type.lower() == 'excel':
                return await self.generate_pf_excel(pf_data, current_user, filters)
            elif format_type.lower() == 'challan':
                if not quarter or not tax_year:
                    raise ValueError("Quarter and tax_year are required for PF Challan")
                return await self.generate_pf_challan(pf_data, current_user, quarter, tax_year)
            elif format_type.lower() == 'return':
                if not quarter or not tax_year:
                    raise ValueError("Quarter and tax_year are required for PF Return")
                return await self.generate_pf_return(pf_data, current_user, quarter, tax_year)
            else:
                raise ValueError(f"Unsupported format: {format_type}")
                
        except Exception as e:
            logger.error(f"Error generating PF report export: {e}")
            raise

    async def generate_pf_csv(
        self,
        pf_data: List[Dict[str, Any]],
        current_user: CurrentUser,
        filters: Optional[Dict[str, Any]] = None
    ) -> bytes:
        """Generate PF report in CSV format"""
        try:
            # Get organisation details
            organisation = await self.organisation_repository.get_by_id(current_user.hostname)
            
            output = io.StringIO()
            writer = csv.writer(output)
            
            # Header
            writer.writerow(['PROVIDENT FUND (PF) REPORT'])
            writer.writerow([''])
            writer.writerow(['Organisation:', organisation.name if organisation else 'MISSING'])
            writer.writerow(['Generated On:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
            if filters:
                writer.writerow(['Month:', filters.get('month', 'N/A')])
                writer.writerow(['Year:', filters.get('year', 'N/A')])
                if filters.get('department'):
                    writer.writerow(['Department:', filters.get('department')])
            writer.writerow([''])
            
            # PF data
            headers = [
                'Sr No', 'Employee ID', 'Employee Name', 'Department', 'Designation',
                'Month', 'Year', 'Basic + DA', 'Employee PF (12%)', 'Employer PF (12%)',
                'Total PF', 'Status'
            ]
            writer.writerow(headers)
            
            total_employee_pf = 0
            total_employer_pf = 0
            total_pf = 0
            
            for idx, pf in enumerate(pf_data, 1):
                employee_pf = pf.get('epf_employee', 0)
                employer_pf = pf.get('epf_employer', employee_pf)  # Default to employee PF if not specified
                total_pf_amount = employee_pf + employer_pf
                
                row = [
                    idx,
                    pf.get('employee_id', 'MISSING'),
                    pf.get('employee_name', 'MISSING'),
                    pf.get('department', 'N/A'),
                    pf.get('designation', 'N/A'),
                    pf.get('month', ''),
                    pf.get('year', ''),
                    pf.get('gross_salary', 0),
                    employee_pf,
                    employer_pf,
                    total_pf_amount,
                    pf.get('status', 'computed')
                ]
                writer.writerow(row)
                
                total_employee_pf += employee_pf
                total_employer_pf += employer_pf
                total_pf += total_pf_amount
            
            # Summary
            writer.writerow([''])
            writer.writerow(['SUMMARY'])
            writer.writerow(['Total Employees with PF:', len(pf_data)])
            writer.writerow(['Total Employee PF:', total_employee_pf])
            writer.writerow(['Total Employer PF:', total_employer_pf])
            writer.writerow(['Total PF Contribution:', total_pf])
            
            return output.getvalue().encode('utf-8')
            
        except Exception as e:
            logger.error(f"Error generating PF CSV: {e}")
            raise

    async def generate_pf_excel(
        self,
        pf_data: List[Dict[str, Any]],
        current_user: CurrentUser,
        filters: Optional[Dict[str, Any]] = None
    ) -> bytes:
        """Generate PF report in Excel format"""
        try:
            # Get organisation details
            organisation = await self.organisation_repository.get_by_id(current_user.hostname)
            
            # Create workbook and worksheet
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "PF Report"
            
            # Header
            worksheet['A1'] = 'PROVIDENT FUND (PF) REPORT'
            worksheet.merge_cells('A1:L1')
            worksheet['A1'].font = Font(bold=True, size=14)
            worksheet['A1'].alignment = Alignment(horizontal='center')
            
            worksheet['A3'] = 'Organisation:'
            worksheet['B3'] = organisation.name if organisation else 'MISSING'
            worksheet['A4'] = 'Generated On:'
            worksheet['B4'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            if filters:
                worksheet['A5'] = 'Month:'
                worksheet['B5'] = filters.get('month', 'N/A')
                worksheet['A6'] = 'Year:'
                worksheet['B6'] = filters.get('year', 'N/A')
                if filters.get('department'):
                    worksheet['A7'] = 'Department:'
                    worksheet['B7'] = filters.get('department')
            
            # Headers
            headers = [
                'Sr No', 'Employee ID', 'Employee Name', 'Department', 'Designation',
                'Month', 'Year', 'Basic + DA', 'Employee PF (12%)', 'Employer PF (12%)',
                'Total PF', 'Status'
            ]
            
            start_row = 10
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=start_row, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # Data
            total_employee_pf = 0
            total_employer_pf = 0
            total_pf = 0
            
            for idx, pf in enumerate(pf_data, 1):
                row = start_row + idx
                employee_pf = pf.get('epf_employee', 0)
                employer_pf = pf.get('epf_employer', employee_pf)
                total_pf_amount = employee_pf + employer_pf
                
                worksheet.cell(row=row, column=1, value=idx)
                worksheet.cell(row=row, column=2, value=pf.get('employee_id', 'MISSING'))
                worksheet.cell(row=row, column=3, value=pf.get('employee_name', 'MISSING'))
                worksheet.cell(row=row, column=4, value=pf.get('department', 'N/A'))
                worksheet.cell(row=row, column=5, value=pf.get('designation', 'N/A'))
                worksheet.cell(row=row, column=6, value=pf.get('month', ''))
                worksheet.cell(row=row, column=7, value=pf.get('year', ''))
                worksheet.cell(row=row, column=8, value=pf.get('gross_salary', 0))
                worksheet.cell(row=row, column=9, value=employee_pf)
                worksheet.cell(row=row, column=10, value=employer_pf)
                worksheet.cell(row=row, column=11, value=total_pf_amount)
                worksheet.cell(row=row, column=12, value=pf.get('status', 'computed'))
                
                total_employee_pf += employee_pf
                total_employer_pf += employer_pf
                total_pf += total_pf_amount
            
            # Summary
            summary_row = start_row + len(pf_data) + 2
            worksheet.cell(row=summary_row, column=1, value='SUMMARY')
            worksheet.cell(row=summary_row, column=1).font = Font(bold=True)
            
            worksheet.cell(row=summary_row + 1, column=1, value='Total Employees with PF:')
            worksheet.cell(row=summary_row + 1, column=2, value=len(pf_data))
            
            worksheet.cell(row=summary_row + 2, column=1, value='Total Employee PF:')
            worksheet.cell(row=summary_row + 2, column=2, value=total_employee_pf)
            
            worksheet.cell(row=summary_row + 3, column=1, value='Total Employer PF:')
            worksheet.cell(row=summary_row + 3, column=2, value=total_employer_pf)
            
            worksheet.cell(row=summary_row + 4, column=1, value='Total PF Contribution:')
            worksheet.cell(row=summary_row + 4, column=2, value=total_pf)
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Save to bytes
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)
            
            return output.getvalue()
            
        except Exception as e:
            logger.error(f"Error generating PF Excel: {e}")
            raise

    async def generate_pf_challan(
        self,
        pf_data: List[Dict[str, Any]],
        current_user: CurrentUser,
        quarter: int,
        tax_year: int
    ) -> bytes:
        """Generate PF Challan format"""
        try:
            # Get organisation details
            organisation = await self.organisation_repository.get_by_id(current_user.hostname)
            
            output = io.StringIO()
            writer = csv.writer(output)
            
            # PF Challan header
            writer.writerow(['PROVIDENT FUND CHALLAN'])
            writer.writerow([''])
            writer.writerow(['Employer Details:'])
            writer.writerow(['Name:', organisation.name if organisation else 'MISSING'])
            writer.writerow(['Establishment Code:', 'MISSING'])
            writer.writerow(['Quarter:', f'Q{quarter}'])
            writer.writerow(['Financial Year:', f'{tax_year}-{tax_year+1}'])
            writer.writerow([''])
            
            # Calculate totals
            total_employee_pf = sum(pf.get('epf_employee', 0) for pf in pf_data)
            total_employer_pf = sum(pf.get('epf_employer', pf.get('epf_employee', 0)) for pf in pf_data)
            total_pf = total_employee_pf + total_employer_pf
            
            # Challan details
            writer.writerow(['Challan Details:'])
            writer.writerow(['Employee PF Contribution:', total_employee_pf])
            writer.writerow(['Employer PF Contribution:', total_employer_pf])
            writer.writerow(['Total PF Amount:', total_pf])
            writer.writerow([''])
            
            # Employee details
            writer.writerow(['Employee Details:'])
            headers = ['Sr No', 'Employee ID', 'Name', 'Employee PF', 'Employer PF', 'Total PF']
            writer.writerow(headers)
            
            for idx, pf in enumerate(pf_data, 1):
                employee_pf = pf.get('epf_employee', 0)
                employer_pf = pf.get('epf_employer', employee_pf)
                total_pf_amount = employee_pf + employer_pf
                
                row = [
                    idx,
                    pf.get('employee_id', 'MISSING'),
                    pf.get('employee_name', 'MISSING'),
                    employee_pf,
                    employer_pf,
                    total_pf_amount
                ]
                writer.writerow(row)
            
            return output.getvalue().encode('utf-8')
            
        except Exception as e:
            logger.error(f"Error generating PF Challan: {e}")
            raise

    async def generate_pf_return(
        self,
        pf_data: List[Dict[str, Any]],
        current_user: CurrentUser,
        quarter: int,
        tax_year: int
    ) -> bytes:
        """Generate PF Return format"""
        try:
            # Get organisation details
            organisation = await self.organisation_repository.get_by_id(current_user.hostname)
            
            output = io.StringIO()
            writer = csv.writer(output)
            
            # PF Return header
            writer.writerow(['PROVIDENT FUND RETURN'])
            writer.writerow([''])
            writer.writerow(['Employer Details:'])
            writer.writerow(['Name:', organisation.name if organisation else 'MISSING'])
            writer.writerow(['Establishment Code:', 'MISSING'])
            writer.writerow(['Quarter:', f'Q{quarter}'])
            writer.writerow(['Financial Year:', f'{tax_year}-{tax_year+1}'])
            writer.writerow([''])
            
            # Return details
            writer.writerow(['Return Details:'])
            headers = ['Sr No', 'Employee ID', 'Name', 'UAN', 'Employee PF', 'Employer PF', 'Total PF', 'Month']
            writer.writerow(headers)
            
            for idx, pf in enumerate(pf_data, 1):
                employee_pf = pf.get('epf_employee', 0)
                employer_pf = pf.get('epf_employer', employee_pf)
                total_pf_amount = employee_pf + employer_pf
                uan = re.sub(r'\D', '', pf.get('employee_id', '')).zfill(12)  # Generate UAN from employee ID
                
                row = [
                    idx,
                    pf.get('employee_id', 'MISSING'),
                    pf.get('employee_name', 'MISSING'),
                    uan,
                    employee_pf,
                    employer_pf,
                    total_pf_amount,
                    pf.get('month', '')
                ]
                writer.writerow(row)
            
            # Summary
            writer.writerow([''])
            total_employee_pf = sum(pf.get('epf_employee', 0) for pf in pf_data)
            total_employer_pf = sum(pf.get('epf_employer', pf.get('epf_employee', 0)) for pf in pf_data)
            total_pf = total_employee_pf + total_employer_pf
            
            writer.writerow(['Total Employee PF:', total_employee_pf])
            writer.writerow(['Total Employer PF:', total_employer_pf])
            writer.writerow(['Total PF Contribution:', total_pf])
            
            return output.getvalue().encode('utf-8')
            
        except Exception as e:
            logger.error(f"Error generating PF Return: {e}")
            raise