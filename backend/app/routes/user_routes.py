from datetime import datetime
import logging
import uuid
import os
from fastapi import APIRouter, HTTPException, Depends, status, Query, UploadFile, File, Form, Body
from openpyxl import load_workbook, Workbook
from fastapi.security import OAuth2PasswordBearer
from auth.jwt_handler import decode_access_token
from models.activity_tracker import ActivityTracker
from models.user_model import UserInfo, UserCreate, Token, User
from auth.auth import extract_emp_id, extract_hostname, get_current_user, extract_role
from auth.dependencies import role_checker
from auth.password_handler import hash_password
from fastapi.responses import JSONResponse, FileResponse, StreamingResponse
import services.user_service as us
from bson import ObjectId
from services.organisation_service import increment_used_employee_strength, user_creation_allowed
from io import BytesIO
from services.activity_tracker_service import track_activity
import json
from utils.file_handler import validate_file, save_file
from utils.json_encoder import mongodb_jsonable_encoder

# Create upload directory if it doesn't exist
UPLOAD_DIR = "uploads"
if not os.path.exists(UPLOAD_DIR):
    os.makedirs(UPLOAD_DIR)
    os.makedirs(os.path.join(UPLOAD_DIR, "pan"))
    os.makedirs(os.path.join(UPLOAD_DIR, "aadhar"))
    os.makedirs(os.path.join(UPLOAD_DIR, "photos"))

def serialize_user(user):
    # Handle MongoDB ObjectId serialization
    user_copy = user.copy()
    if '_id' in user_copy:
        user_copy['_id'] = str(user_copy['_id'])
    
    # Convert any other ObjectId fields to strings
    for key, value in user_copy.items():
        if isinstance(value, ObjectId):
            user_copy[key] = str(value)
    
    # Handle date serialization
    if user_copy.get("created_at"):
        user_copy["created_at"] = user_copy["created_at"].isoformat() if hasattr(user_copy["created_at"], "isoformat") else user_copy["created_at"]
    
    if user_copy.get("updated_at"):
        user_copy["updated_at"] = user_copy["updated_at"].isoformat() if hasattr(user_copy["updated_at"], "isoformat") else user_copy["updated_at"]
    
    return user_copy

def save_uploaded_file(file: UploadFile, directory: str) -> str:
    if not file:
        return None
    
    file_extension = os.path.splitext(file.filename)[1].lower()
    allowed_extensions = {'.jpg', '.jpeg', '.png', '.pdf'}
    
    if file_extension not in allowed_extensions:
        raise HTTPException(
            status_code=400,
            detail=f"File type not allowed. Allowed types: {', '.join(allowed_extensions)}"
        )
    
    filename = f"{uuid.uuid4()}{file_extension}"
    file_path = os.path.join(directory, filename)
    
    try:
        with open(file_path, "wb") as buffer:
            buffer.write(file.file.read())
        return file_path
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Could not save file: {str(e)}")

logger = logging.getLogger(__name__)
router = APIRouter()
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="login")

@router.post("/users", status_code=status.HTTP_201_CREATED)
async def create_user(
    data: UserInfo = Body(...),
    current_emp_id: str = Depends(extract_emp_id),
    hostname: str = Depends(extract_hostname)
):
    """
    Create a new user without file uploads (JSON only)
    """
    try:
        if not user_creation_allowed(hostname):
            raise HTTPException(status_code=403, detail="User limit reached!!! Contact your administrator")
        
        logger.info("Creating user: %s", data.name)
        us.validate_user_data(data)
        
        result = us.create_user(data, hostname)
        logger.info(result)
        increment_used_employee_strength(hostname)
        return result
    except Exception as e:
        logger.error(f"Error creating user: {str(e)}")
        raise HTTPException(status_code=400, detail=str(e))

@router.post("/users/with-files", status_code=status.HTTP_201_CREATED)
async def create_user_with_files(
    user_data: str = Form(...),
    pan_file: UploadFile = File(None),
    aadhar_file: UploadFile = File(None),
    photo: UploadFile = File(None),
    current_emp_id: str = Depends(extract_emp_id),
    hostname: str = Depends(extract_hostname)
):
    """
    Create a new user with document uploads
    """
    try:
        if not user_creation_allowed(hostname):
            raise HTTPException(status_code=403, detail="User limit reached!!! Contact your administrator")
        
        # Parse the JSON string into UserInfo model
        try:
            user_data_dict = json.loads(user_data)
            user_info = UserInfo(**user_data_dict)
        except json.JSONDecodeError:
            raise HTTPException(status_code=400, detail="Invalid JSON data")
        
        logger.info("Creating user with files: %s", user_info.name)
        us.validate_user_data(user_info)

        # Validate and save files
        if photo:
            is_valid, error = validate_file(photo, allowed_types=["image/jpeg", "image/png"], max_size=2*1024*1024)
            if not is_valid:
                raise HTTPException(status_code=400, detail=f"Invalid photo: {error}")
            user_info.photo_path = await save_file(photo, "photos")
            
        if pan_file:
            is_valid, error = validate_file(pan_file, allowed_types=["image/jpeg", "image/png", "application/pdf"], max_size=5*1024*1024)
            if not is_valid:
                raise HTTPException(status_code=400, detail=f"Invalid PAN file: {error}")
            user_info.pan_file_path = await save_file(pan_file, "pan")
            
        if aadhar_file:
            is_valid, error = validate_file(aadhar_file, allowed_types=["image/jpeg", "image/png", "application/pdf"], max_size=5*1024*1024)
            if not is_valid:
                raise HTTPException(status_code=400, detail=f"Invalid Aadhar file: {error}")
            user_info.aadhar_file_path = await save_file(aadhar_file, "aadhar")

        result = us.create_user(user_info, hostname)
        logger.info(result)
        increment_used_employee_strength(hostname)
        return result
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creating user with files: {str(e)}")
        raise HTTPException(status_code=400, detail=str(e))

@router.post("/users/import/with-file", status_code=status.HTTP_201_CREATED)
async def import_users_with_file(
    file: UploadFile = File(...), 
    hostname: str = Depends(extract_hostname),
    current_emp_id: str = Depends(extract_emp_id)
):
    """
    Imports users from an Excel (.xlsx) file.
    """
    logger.info("Received file for user import: %s", file.filename)

    # Validate file type
    if not file.filename.endswith('.xlsx'):
        logger.warning("Invalid file type uploaded: %s", file.filename)
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported")

    # Validate file size and type
    is_valid, error = validate_file(
        file, 
        allowed_types=["application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"],
        max_size=10*1024*1024
    )
    if not is_valid:
        raise HTTPException(status_code=400, detail=error)

    contents = await file.read()

    try:
        logger.info("Attempting to load Excel workbook")
        wb = load_workbook(BytesIO(contents), data_only=True)
        sheet = wb.active
    except Exception as e:
        logger.error("Failed to read Excel file: %s", str(e))
        raise HTTPException(status_code=500, detail="Error reading Excel file")

    headers = [cell.value for cell in sheet[1]]
    expected_headers = ["emp_id", "email","name", "gender", "dob", "doj", "mobile", "manager_id", "password", "role"]
    logger.info("Parsed headers from Excel: %s", headers)

    if headers != expected_headers:
        logger.warning("Invalid headers in uploaded Excel: %s", headers)
        raise HTTPException(status_code=400, detail=f"Invalid headers. Expected: {expected_headers}")

    created, failed = 0, 0
    errors = []

    logger.info("Starting user creation loop")
    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        try:
            logger.info("Processing row %d: %s", idx, row)
            user = UserInfo(
                emp_id=row[0],
                email=row[1],
                name=row[2],
                gender=row[3],
                dob=row[4],
                doj=row[5], 
                mobile=row[6],
                manager_id=row[7],
                password=row[8],
                role=row[9]
            )
            us.create_user(user, hostname)
            created += 1
            activity = ActivityTracker(
                emp_id=current_emp_id,
                activity="importUsersSuccess",
                date=datetime.now(),
                metadata=user.model_dump()
            )
            track_activity(activity)
        except Exception as e:
            failed += 1
            error_msg = f"Row {idx}: {str(e)}"
            logger.error(error_msg)
            errors.append(error_msg)
            activity = ActivityTracker(
                emp_id=current_emp_id,
                activity="importUsersFailed",
                date=datetime.now(),
                metadata={}
            )
            track_activity(activity)

    logger.info("User import completed: %d created, %d failed", created, failed)
    return {
        "created": created,
        "failed": failed,
        "errors": errors
    }

@router.get("/users/me")
def read_users_me(hostname: str = Depends(extract_hostname),
                        current_emp_id: str = Depends(extract_emp_id)):
    """
    Returns the username of the current logged-in user.
    """
    logger.info("read_users_me successful for username: %s", current_emp_id)
    user = us.get_user_by_emp_id(current_emp_id, hostname)
    return mongodb_jsonable_encoder(user)

@router.get("/users/emp/{emp_id}")
def read_users_me(emp_id: str, hostname: str = Depends(extract_hostname), 
                  role: str = Depends(role_checker(["admin", "superadmin", "manager"]))):
    """
    Returns the username of the current logged-in user.
    """
    logger.info("read_users_me successful for username: %s", emp_id)
    user = us.get_user_by_emp_id(emp_id, hostname)
    return mongodb_jsonable_encoder(user)

@router.get("/users")
def list_users(
    skip: int = Query(0, ge=0),
    limit: int = Query(10, ge=1),
    role: str = Depends(role_checker(["admin", "superadmin", "manager"])),
    hostname: str = Depends(extract_hostname),
    current_emp_id: str = Depends(extract_emp_id)
):
    """
    Lists all login info entries, paginated.
    Only accessible by admin or superadmin.
    """
    logger.info("Listing users with skip=%d, limit=%d", skip, limit)
    if role == "manager":
        logger.info("Listing users for manager: %s", current_emp_id)
        users = us.get_users_by_manager_id(current_emp_id, hostname)
    else:
        logger.info("Listing all users")
        users = us.get_all_users(hostname)
    
    # Use our custom encoder to safely handle MongoDB ObjectId
    encoded_users = mongodb_jsonable_encoder(users[skip:skip + limit])

    logger.info("Returning %d users out of %d", len(encoded_users), len(users))
    return {
        "total": len(users),
        "users": encoded_users
    }
    
@router.get("/users/stats")
def get_user_stats(hostname: str = Depends(extract_hostname)):
    stats = us.get_users_stats(hostname)
    return mongodb_jsonable_encoder(stats)

@router.get("/users/my/directs")
def get_my_directs(hostname: str = Depends(extract_hostname),
                        current_emp_id: str = Depends(extract_emp_id)):
    users = us.get_user_by_manager_id(current_emp_id, hostname)
    return mongodb_jsonable_encoder(users)

@router.get("/users/manager/directs")
def get_user_by_manager_id(manager_id: str, hostname: str = Depends(extract_hostname)):
    users = us.get_user_by_manager_id(manager_id, hostname)
    return mongodb_jsonable_encoder(users)

@router.get("/users/template")
def download_template():
    """
    Endpoint to download the user import template file.
    """
    try:
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        
        # Add headers
        headers = ["emp_id", "email", "name", "gender", "dob", "doj", "mobile", "manager_id", "password", "role"]
        ws.append(headers)
        
        # Add example row
        example_row = [
            "EMP001",
            "example@company.com",
            "John Doe",
            "male",
            "1990-01-01",
            "2023-01-01",
            "1234567890",
            "MAN001",
            "password123",
            "user"
        ]
        ws.append(example_row)
        
        # Save to BytesIO
        template_file = BytesIO()
        wb.save(template_file)
        template_file.seek(0)
        
        return StreamingResponse(
            template_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=user_import_template.xlsx"
            }
        )
    except Exception as e:
        logger.error(f"Error generating template: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to generate template file")
