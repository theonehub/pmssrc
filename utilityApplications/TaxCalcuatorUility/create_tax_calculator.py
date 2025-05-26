import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

def create_tax_calculator_excel():
    # Create a new workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create worksheets
    input_sheet = wb.create_sheet("INPUT_DATA")
    deductions_sheet = wb.create_sheet("DEDUCTIONS")
    calculations_sheet = wb.create_sheet("CALCULATIONS")
    tax_calc_sheet = wb.create_sheet("TAX_CALCULATION")
    validation_sheet = wb.create_sheet("VALIDATION")
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    input_fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
    formula_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # Setup INPUT_DATA worksheet
    setup_input_data_sheet(input_sheet, header_font, header_fill, input_fill)
    
    # Setup DEDUCTIONS worksheet
    setup_deductions_sheet(deductions_sheet, header_font, header_fill, input_fill, formula_fill)
    
    # Setup CALCULATIONS worksheet
    setup_calculations_sheet(calculations_sheet, header_font, header_fill, formula_fill)
    
    # Setup TAX_CALCULATION worksheet
    setup_tax_calculation_sheet(tax_calc_sheet, header_font, header_fill, formula_fill)
    
    # Setup VALIDATION worksheet
    setup_validation_sheet(validation_sheet, header_font, header_fill, formula_fill)
    
    # Save the workbook
    wb.save("Tax_Calculator.xlsx")
    print("Tax Calculator Excel file created successfully!")

def setup_input_data_sheet(ws, header_font, header_fill, input_fill):
    """Setup the INPUT_DATA worksheet"""
    
    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    
    # Basic Information Section
    ws['A1'] = "BASIC INFORMATION"
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws.merge_cells('A1:B1')
    
    basic_info = [
        ("Employee ID", ""),
        ("Age", ""),
        ("Tax Regime", ""),
        ("Is Government Employee", ""),
        ("Tax Year", "2024-25"),
        ("Date of Joining", ""),
        ("Date of Leaving", ""),
        ("Service Years", "=IF(AND(B6<>\"\",B7<>\"\"),DATEDIF(B6,B7,\"Y\"),\"\")")
    ]
    
    for i, (label, value) in enumerate(basic_info, 2):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
        if not value.startswith('='):
            ws[f'B{i}'].fill = input_fill
    
    # Add data validation for Tax Regime
    regime_validation = DataValidation(type="list", formula1='"old,new"')
    ws.add_data_validation(regime_validation)
    regime_validation.add('B3')
    
    # Add data validation for Government Employee
    govt_validation = DataValidation(type="list", formula1='"TRUE,FALSE"')
    ws.add_data_validation(govt_validation)
    govt_validation.add('B4')
    
    # Salary Components Section
    ws['A10'] = "SALARY COMPONENTS"
    ws['A10'].font = header_font
    ws['A10'].fill = header_fill
    ws.merge_cells('A10:B10')
    
    salary_components = [
        ("Basic Salary", ""),
        ("Dearness Allowance", ""),
        ("HRA", ""),
        ("Actual Rent Paid", ""),
        ("HRA City", ""),
        ("Special Allowance", ""),
        ("Bonus", ""),
        ("Commission", ""),
        ("Transport Allowance", ""),
        ("Children Education Allowance", ""),
        ("Children Count", ""),
        ("Education Months", ""),
        ("Entertainment Allowance", ""),
        ("Medical Allowance", ""),
        ("Any Other Allowance", "")
    ]
    
    for i, (label, value) in enumerate(salary_components, 11):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
        ws[f'B{i}'].fill = input_fill
    
    # Add data validation for HRA City
    city_validation = DataValidation(type="list", formula1='"Delhi,Mumbai,Kolkata,Chennai,Others"')
    ws.add_data_validation(city_validation)
    city_validation.add('B15')
    
    # Other Income Sources Section
    ws['A27'] = "OTHER INCOME SOURCES"
    ws['A27'].font = header_font
    ws['A27'].fill = header_fill
    ws.merge_cells('A27:B27')
    
    other_income = [
        ("Interest on Savings", ""),
        ("Interest on FD", ""),
        ("Interest on RD", ""),
        ("Dividend Income", ""),
        ("Gift Income", ""),
        ("Other Interest", ""),
        ("Business Income", ""),
        ("Other Income", "")
    ]
    
    for i, (label, value) in enumerate(other_income, 28):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
        ws[f'B{i}'].fill = input_fill
    
    # House Property Section
    ws['A37'] = "HOUSE PROPERTY"
    ws['A37'].font = header_font
    ws['A37'].fill = header_fill
    ws.merge_cells('A37:B37')
    
    house_property = [
        ("Property Status", ""),
        ("Rent Income", ""),
        ("Property Tax", ""),
        ("Home Loan Interest", ""),
        ("Pre-Construction Interest", "")
    ]
    
    for i, (label, value) in enumerate(house_property, 38):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
        ws[f'B{i}'].fill = input_fill
    
    # Add validation for Property Status
    property_validation = DataValidation(type="list", formula1='"Self-Occupied,Let-Out"')
    ws.add_data_validation(property_validation)
    property_validation.add('B38')
    
    # Capital Gains Section
    ws['A44'] = "CAPITAL GAINS"
    ws['A44'].font = header_font
    ws['A44'].fill = header_fill
    ws.merge_cells('A44:B44')
    
    capital_gains = [
        ("STCG 111A", ""),
        ("STCG Other Assets", ""),
        ("STCG Debt MF", ""),
        ("LTCG 112A", ""),
        ("LTCG Other Assets", ""),
        ("LTCG Debt MF", "")
    ]
    
    for i, (label, value) in enumerate(capital_gains, 45):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
        ws[f'B{i}'].fill = input_fill
    
    # Leave Encashment & Others Section
    ws['A52'] = "LEAVE ENCASHMENT & OTHERS"
    ws['A52'].font = header_font
    ws['A52'].fill = header_fill
    ws.merge_cells('A52:B52')
    
    leave_others = [
        ("Leave Encashment Amount", ""),
        ("Leave Days Encashed", ""),
        ("VRS Amount", ""),
        ("Pension Income", ""),
        ("Gratuity Amount", ""),
        ("Retrenchment Compensation", ""),
        ("Average Monthly Salary", "")
    ]
    
    for i, (label, value) in enumerate(leave_others, 53):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
        ws[f'B{i}'].fill = input_fill

def setup_deductions_sheet(ws, header_font, header_fill, input_fill, formula_fill):
    """Setup the DEDUCTIONS worksheet"""
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    
    # Section 80C Components
    ws['A1'] = "SECTION 80C COMPONENTS"
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws.merge_cells('A1:B1')
    
    section_80c = [
        ("LIC Premium", ""),
        ("EPF Contribution", ""),
        ("PPF Investment", ""),
        ("NSC Investment", ""),
        ("ELSS Investment", ""),
        ("Home Loan Principal", ""),
        ("Tuition Fees", ""),
        ("Others 80C", ""),
        ("Total 80C", "=SUM(B2:B9)"),
        ("80C Deduction", "=MIN(B10,150000)")
    ]
    
    for i, (label, value) in enumerate(section_80c, 2):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
        if value.startswith('='):
            ws[f'B{i}'].fill = formula_fill
        else:
            ws[f'B{i}'].fill = input_fill
    
    # Section 80CCD
    ws['A13'] = "SECTION 80CCD (NPS)"
    ws['A13'].font = header_font
    ws['A13'].fill = header_fill
    ws.merge_cells('A13:B13')
    
    section_80ccd = [
        ("NPS Self Contribution", ""),
        ("NPS Additional 1B", ""),
        ("NPS Employer Contribution", ""),
        ("80CCD(1) Included in 80C", "Already in B11"),
        ("80CCD(1B) Additional", "=MIN(B15,50000)"),
        ("80CCD(2) Employer", "=MIN(B16,IF(INPUT_DATA.B4,(INPUT_DATA.B11+INPUT_DATA.B12)*0.14,(INPUT_DATA.B11+INPUT_DATA.B12)*0.10))")
    ]
    
    for i, (label, value) in enumerate(section_80ccd, 14):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
        if value.startswith('='):
            ws[f'B{i}'].fill = formula_fill
        elif "Already" not in value:
            ws[f'B{i}'].fill = input_fill
    
    # Section 80D Health Insurance
    ws['A21'] = "SECTION 80D HEALTH INSURANCE"
    ws['A21'].font = header_font
    ws['A21'].fill = header_fill
    ws.merge_cells('A21:B21')
    
    section_80d = [
        ("Self & Family Premium", ""),
        ("Preventive Checkup", ""),
        ("Parent Premium", ""),
        ("Parent Age", ""),
        ("80D Self & Family", "=MIN(B22+MIN(B23,5000),IF(INPUT_DATA.B2>=60,50000,25000))"),
        ("80D Parents", "=MIN(B24,IF(B25>=60,50000,25000))"),
        ("Total 80D", "=B26+B27")
    ]
    
    for i, (label, value) in enumerate(section_80d, 22):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
        if value.startswith('='):
            ws[f'B{i}'].fill = formula_fill
        else:
            ws[f'B{i}'].fill = input_fill
    
    # Other Deductions
    ws['A30'] = "OTHER DEDUCTIONS"
    ws['A30'].font = header_font
    ws['A30'].fill = header_fill
    ws.merge_cells('A30:B30')
    
    other_deductions = [
        ("80DD Relation", ""),
        ("80DD Disability %", ""),
        ("80DD Deduction", "=IF(AND(B31<>\"\",B32<>\"\"),IF(B32=\"Between 40%-80%\",75000,125000),0)"),
        ("80DDB Medical Expenses", ""),
        ("80DDB Patient Age", ""),
        ("80DDB Deduction", "=MIN(B34,IF(B35>=60,100000,40000))"),
        ("80E Education Loan Interest", ""),
        ("80E Deduction", "=B37"),
        ("80U Self Disability %", ""),
        ("80U Deduction", "=IF(B39=\"Between 40%-80%\",75000,IF(B39=\"More than 80%\",125000,0))"),
        ("80G Donations", ""),
        ("80G Deduction", "=B41")
    ]
    
    for i, (label, value) in enumerate(other_deductions, 31):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
        if value.startswith('='):
            ws[f'B{i}'].fill = formula_fill
        else:
            ws[f'B{i}'].fill = input_fill
    
    # Add validations for disability percentages
    disability_validation = DataValidation(type="list", formula1='"Between 40%-80%,More than 80%"')
    ws.add_data_validation(disability_validation)
    disability_validation.add('B32')
    disability_validation.add('B39')

def setup_calculations_sheet(ws, header_font, header_fill, formula_fill):
    """Setup the CALCULATIONS worksheet"""
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 25
    
    # Salary Income Calculations
    ws['A1'] = "SALARY INCOME CALCULATIONS"
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws.merge_cells('A1:C1')
    
    salary_calcs = [
        ("Basic Components", "", "=SUM(INPUT_DATA.B11:B25)"),
        ("HRA Exemption", "", "=IF(INPUT_DATA.B3=\"old\",MIN(INPUT_DATA.B13,IF(INPUT_DATA.B15=\"Delhi\",(INPUT_DATA.B11+INPUT_DATA.B12)*0.5,(INPUT_DATA.B11+INPUT_DATA.B12)*0.4),MAX(0,INPUT_DATA.B14-(INPUT_DATA.B11+INPUT_DATA.B12)*0.1)),0)"),
        ("Other Exemptions", "", "0"),
        ("Standard Deduction", "", "=IF(INPUT_DATA.B3=\"new\",MIN(75000,C2),MIN(50000,C2))"),
        ("Net Salary Income", "", "=MAX(0,C2-C3-C4-C5)")
    ]
    
    for i, (desc, col_b, formula) in enumerate(salary_calcs, 2):
        ws[f'A{i}'] = desc
        ws[f'C{i}'] = formula
        ws[f'C{i}'].fill = formula_fill
    
    # Other Sources Income
    ws['A8'] = "OTHER SOURCES INCOME"
    ws['A8'].font = header_font
    ws['A8'].fill = header_fill
    ws.merge_cells('A8:C8')
    
    other_sources = [
        ("Total Interest", "", "=SUM(INPUT_DATA.B28:B30)"),
        ("Interest Exemption", "", "=IF(INPUT_DATA.B3=\"old\",IF(INPUT_DATA.B2>=60,MIN(50000,C9),MIN(10000,INPUT_DATA.B28)),0)"),
        ("Taxable Interest", "", "=MAX(0,C9-C10)"),
        ("Other Income", "", "=SUM(INPUT_DATA.B31:B35)"),
        ("Total Other Sources", "", "=C11+C12")
    ]
    
    for i, (desc, col_b, formula) in enumerate(other_sources, 9):
        ws[f'A{i}'] = desc
        ws[f'C{i}'] = formula
        ws[f'C{i}'].fill = formula_fill
    
    # House Property Income
    ws['A15'] = "HOUSE PROPERTY INCOME"
    ws['A15'].font = header_font
    ws['A15'].fill = header_fill
    ws.merge_cells('A15:C15')
    
    house_property = [
        ("Annual Value", "", "=IF(INPUT_DATA.B38=\"Self-Occupied\",0,INPUT_DATA.B39)"),
        ("Property Tax", "", "=INPUT_DATA.B40"),
        ("Net Annual Value", "", "=MAX(0,C16-C17)"),
        ("Standard Deduction", "", "=IF(INPUT_DATA.B38=\"Let-Out\",C18*0.3,0)"),
        ("Interest Deduction", "", "=IF(INPUT_DATA.B38=\"Self-Occupied\",MIN(200000,INPUT_DATA.B41),INPUT_DATA.B41)"),
        ("Pre-Construction", "", "=INPUT_DATA.B42/5"),
        ("Net HP Income", "", "=C18-C19-C20-C21")
    ]
    
    for i, (desc, col_b, formula) in enumerate(house_property, 16):
        ws[f'A{i}'] = desc
        ws[f'C{i}'] = formula
        ws[f'C{i}'].fill = formula_fill
    
    # Capital Gains
    ws['A24'] = "CAPITAL GAINS"
    ws['A24'].font = header_font
    ws['A24'].fill = header_fill
    ws.merge_cells('A24:C24')
    
    capital_gains = [
        ("STCG Slab Rate", "", "=INPUT_DATA.B46+INPUT_DATA.B47"),
        ("STCG Special Tax", "", "=INPUT_DATA.B45*0.20"),
        ("LTCG 112A Taxable", "", "=MAX(0,INPUT_DATA.B48-125000)"),
        ("LTCG 112A Tax", "", "=C27*0.125"),
        ("LTCG Other Tax", "", "=(INPUT_DATA.B49+INPUT_DATA.B50)*0.125")
    ]
    
    for i, (desc, col_b, formula) in enumerate(capital_gains, 25):
        ws[f'A{i}'] = desc
        ws[f'C{i}'] = formula
        ws[f'C{i}'].fill = formula_fill
    
    # Total Income Summary
    ws['A31'] = "TOTAL INCOME SUMMARY"
    ws['A31'].font = header_font
    ws['A31'].fill = header_fill
    ws.merge_cells('A31:C31')
    
    income_summary = [
        ("Gross Income", "", "=C6+C13+C22+C25"),
        ("Total Deductions", "", "=IF(INPUT_DATA.B3=\"old\",SUM(DEDUCTIONS.B11,DEDUCTIONS.B18,DEDUCTIONS.B19,DEDUCTIONS.B28,DEDUCTIONS.B33,DEDUCTIONS.B36,DEDUCTIONS.B38,DEDUCTIONS.B40,DEDUCTIONS.B42),0)"),
        ("Net Taxable Income", "", "=MAX(0,C32-C33)")
    ]
    
    for i, (desc, col_b, formula) in enumerate(income_summary, 32):
        ws[f'A{i}'] = desc
        ws[f'C{i}'] = formula
        ws[f'C{i}'].fill = formula_fill

def setup_tax_calculation_sheet(ws, header_font, header_fill, formula_fill):
    """Setup the TAX_CALCULATION worksheet"""
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 25
    
    # Tax Slab Calculation
    ws['A1'] = "TAX SLAB CALCULATION"
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws.merge_cells('A1:C1')
    
    tax_info = [
        ("Age Category", "", "=IF(INPUT_DATA.B2>=80,\"Super Senior\",IF(INPUT_DATA.B2>=60,\"Senior\",\"Regular\"))"),
        ("Regime", "", "=INPUT_DATA.B3"),
        ("", "", ""),
        ("", "", ""),
        ("Tax Slab 1", "", "=IF(INPUT_DATA.B3=\"old\",IF(C1=\"Regular\",IF(CALCULATIONS.C34<=250000,0,MIN(50000,(CALCULATIONS.C34-250000)*0.05)),IF(C1=\"Senior\",IF(CALCULATIONS.C34<=300000,0,MIN(50000,(CALCULATIONS.C34-300000)*0.05)),IF(CALCULATIONS.C34<=500000,0,MIN(50000,(CALCULATIONS.C34-500000)*0.05)))),IF(CALCULATIONS.C34<=300000,0,MIN(25000,(CALCULATIONS.C34-300000)*0.05)))"),
        ("Tax Slab 2", "", "=IF(INPUT_DATA.B3=\"old\",IF(C1=\"Regular\",IF(CALCULATIONS.C34<=500000,0,MIN(50000,(CALCULATIONS.C34-500000)*0.20)),IF(C1=\"Senior\",IF(CALCULATIONS.C34<=500000,0,MIN(40000,(CALCULATIONS.C34-500000)*0.20)),IF(CALCULATIONS.C34<=500000,0,0))),IF(CALCULATIONS.C34<=600000,0,MIN(25000,(CALCULATIONS.C34-600000)*0.10)))"),
        ("Tax Slab 3", "", "=IF(INPUT_DATA.B3=\"old\",IF(CALCULATIONS.C34<=1000000,0,(CALCULATIONS.C34-1000000)*0.30),IF(CALCULATIONS.C34<=900000,0,MIN(50000,(CALCULATIONS.C34-900000)*0.15)))"),
        ("Tax Slab 4", "", "=IF(INPUT_DATA.B3=\"new\",IF(CALCULATIONS.C34<=1200000,0,MIN(150000,(CALCULATIONS.C34-1200000)*0.20)),0)"),
        ("Tax Slab 5", "", "=IF(INPUT_DATA.B3=\"new\",IF(CALCULATIONS.C34<=1500000,0,(CALCULATIONS.C34-1500000)*0.30),0)"),
        ("Regular Tax", "", "=SUM(C5:C9)")
    ]
    
    for i, (desc, col_b, formula) in enumerate(tax_info, 2):
        ws[f'A{i}'] = desc
        ws[f'C{i}'] = formula
        if formula:
            ws[f'C{i}'].fill = formula_fill
    
    # Final Tax Calculation
    ws['A13'] = "FINAL TAX CALCULATION"
    ws['A13'].font = header_font
    ws['A13'].fill = header_fill
    ws.merge_cells('A13:C13')
    
    final_tax = [
        ("Base Tax", "", "=C10+CALCULATIONS.C26+CALCULATIONS.C28+CALCULATIONS.C29"),
        ("87A Rebate", "", "=IF(AND(INPUT_DATA.B3=\"old\",CALCULATIONS.C34<=500000),MIN(12500,C14),IF(AND(INPUT_DATA.B3=\"new\",CALCULATIONS.C34<=1200000),MIN(60000,C14),0))"),
        ("Tax After Rebate", "", "=MAX(0,C14-C15)"),
        ("Surcharge Rate", "", "=IF(CALCULATIONS.C34>50000000,0.37,IF(CALCULATIONS.C34>20000000,0.25,IF(CALCULATIONS.C34>10000000,0.15,IF(CALCULATIONS.C34>5000000,0.10,0))))"),
        ("Surcharge", "", "=C16*C17"),
        ("Tax + Surcharge", "", "=C16+C18"),
        ("Health & Education Cess", "", "=C19*0.04"),
        ("FINAL TAX", "", "=C19+C20")
    ]
    
    for i, (desc, col_b, formula) in enumerate(final_tax, 14):
        ws[f'A{i}'] = desc
        ws[f'C{i}'] = formula
        ws[f'C{i}'].fill = formula_fill
        if desc == "FINAL TAX":
            ws[f'A{i}'].font = Font(bold=True)
            ws[f'C{i}'].font = Font(bold=True)

def setup_validation_sheet(ws, header_font, header_fill, formula_fill):
    """Setup the VALIDATION worksheet"""
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 25
    
    # Input Validations
    ws['A1'] = "INPUT VALIDATIONS"
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws.merge_cells('A1:C1')
    
    input_validations = [
        ("Age Valid", "", "=IF(AND(INPUT_DATA.B2>=18,INPUT_DATA.B2<=100),\"✓\",\"✗\")"),
        ("Salary Components ≥ 0", "", "=IF(COUNTIF(INPUT_DATA.B11:B25,\"<0\")=0,\"✓\",\"✗\")"),
        ("HRA City Valid", "", "=IF(INPUT_DATA.B15<>\"\",\"✓\",\"✗\")"),
        ("80C Limit Check", "", "=IF(DEDUCTIONS.B10<=150000,\"✓\",\"Exceeds Limit\")"),
        ("80D Age-based Check", "", "=IF(DEDUCTIONS.B26<=IF(INPUT_DATA.B2>=60,50000,25000),\"✓\",\"Exceeds Limit\")")
    ]
    
    for i, (desc, col_b, formula) in enumerate(input_validations, 2):
        ws[f'A{i}'] = desc
        ws[f'C{i}'] = formula
        ws[f'C{i}'].fill = formula_fill
    
    # Calculation Cross-checks
    ws['A8'] = "CALCULATION CROSS-CHECKS"
    ws['A8'].font = header_font
    ws['A8'].fill = header_fill
    ws.merge_cells('A8:C8')
    
    cross_checks = [
        ("Tax Regime Applied", "", "=INPUT_DATA.B3"),
        ("Deductions Applied", "", "=IF(INPUT_DATA.B3=\"old\",\"Yes\",\"No\")"),
        ("Age-based Exemption", "", "=IF(INPUT_DATA.B2>=60,\"Applied\",\"Not Applied\")"),
        ("Capital Gains Rate", "", "\"STCG: 20%, LTCG: 12.5%\"")
    ]
    
    for i, (desc, col_b, formula) in enumerate(cross_checks, 9):
        ws[f'A{i}'] = desc
        ws[f'C{i}'] = formula
        ws[f'C{i}'].fill = formula_fill

if __name__ == "__main__":
    create_tax_calculator_excel() 